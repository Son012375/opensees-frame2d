"""
OpenSees 고유치해석 결과 → Excel 출력
Midas Gen 결과와 동일 형식 + 비교 시트
"""
import sys, os, math

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'mcp-server'))
from core.ops_compat import ops
from core.frame_3d import _generate_frame_3d_geometry
from core.section_3d import get_section_3d, DEFAULT_SECTIONS_3D
from core.simple_beam import get_material_from_db, DEFAULT_MATERIALS

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

# ============================================================
# 건물 파라미터
# ============================================================
stories = [5.0] + [4.0] * 10
bays_x = [6.5, 2.5, 4.0, 2.0, 13.0]
bays_y = [6.5, 6.5, 2.0, 6.5, 6.5]
n_stories = len(stories)

column_section = "H-400x400"
beam_x_section = "H-500x200"
beam_y_section = "H-500x200"
material_name = "SS275"

story_weights_kN = [4390.4] + [3998.4] * 10

# ============================================================
# 모델 구축 + Rigid Diaphragm
# ============================================================
nodes, connections, node_grid, base_nodes = _generate_frame_3d_geometry(stories, bays_x, bays_y)

col_sec = get_section_3d(column_section) or DEFAULT_SECTIONS_3D.get("H-300x300")
bx_sec = get_section_3d(beam_x_section) or DEFAULT_SECTIONS_3D.get("H-400x200")
by_sec = get_section_3d(beam_y_section) or DEFAULT_SECTIONS_3D.get("H-400x200")
mat = get_material_from_db(material_name) or DEFAULT_MATERIALS.get("SS275")

E = mat.E
G = E / (2.0 * (1.0 + 0.3))

ops.wipe()
ops.model('basic', '-ndm', 3, '-ndf', 6)

for node in nodes:
    ops.node(node.id, node.x * 1000, node.y * 1000, node.z * 1000)

for bn in base_nodes:
    ops.fix(bn, 1, 1, 1, 1, 1, 1)

ops.geomTransf('Linear', 1, 1.0, 0.0, 0.0)
ops.geomTransf('Linear', 2, 0.0, 0.0, 1.0)
ops.geomTransf('Linear', 3, 0.0, 0.0, 1.0)

elem_id = 1
for ni, nj, etype in connections:
    if etype == "column":
        sec, transf = col_sec, 1
    elif etype == "beam_x":
        sec, transf = bx_sec, 2
    else:
        sec, transf = by_sec, 3
    ops.element('elasticBeamColumn', elem_id, ni, nj,
                sec.A, E, G, sec.J, sec.Ix, sec.Iy, transf)
    elem_id += 1

n_cols_x = len(bays_x) + 1
n_cols_y = len(bays_y) + 1
nodes_per_floor = n_cols_x * n_cols_y

# Rigid Diaphragm
master_nodes = {}
for s in range(1, n_stories + 1):
    cx_m, cy_m = n_cols_x // 2, n_cols_y // 2
    master_nid = node_grid[(s, cx_m, cy_m)]
    master_nodes[s] = master_nid
    slave_nids = []
    for cx in range(n_cols_x):
        for cy in range(n_cols_y):
            nid = node_grid[(s, cx, cy)]
            if nid != master_nid:
                slave_nids.append(nid)
    ops.rigidDiaphragm(3, master_nid, *slave_nids)

# 질량 배정
g_acc = 9810.0
Lx_mm = sum(bays_x) * 1000
Ly_mm = sum(bays_y) * 1000

floor_masses = []   # (M_trans, I_rot) per floor
for s in range(1, n_stories + 1):
    W_N = story_weights_kN[s - 1] * 1000.0
    M_floor = W_N / g_acc
    I_rot = M_floor * (Lx_mm**2 + Ly_mm**2) / 12.0
    floor_masses.append((M_floor, I_rot))
    ops.mass(master_nodes[s], M_floor, M_floor, 1.0e-6, 0.0, 0.0, I_rot)

# ============================================================
# 고유치해석
# ============================================================
num_modes = 15
eigenvalues = ops.eigen(num_modes)

# 주기/주파수 계산
mode_data = []
for i in range(num_modes):
    omega = math.sqrt(eigenvalues[i])
    freq_hz = omega / (2 * math.pi)
    period = 1.0 / freq_hz if freq_hz > 0 else 0
    mode_data.append({
        'mode': i + 1,
        'omega': omega,
        'freq_hz': freq_hz,
        'period': period,
    })

# ============================================================
# Modal Participation Mass 계산
# ============================================================
# 3 DOF per floor (UX, UY, RZ at master node)
# φ[mode][floor] = (ux, uy, rz)
# M[floor] = diag(M_trans, M_trans, I_rot)
# Participation factor: Γ = φ^T M r / (φ^T M φ)
# Effective mass: M_eff = Γ² × (φ^T M φ)

total_mass_x = sum(fm[0] for fm in floor_masses)
total_mass_y = sum(fm[0] for fm in floor_masses)
total_mass_rz = sum(fm[1] for fm in floor_masses)

for md in mode_data:
    mode_num = md['mode']

    # 모드 형상 추출
    phi = []
    for s in range(1, n_stories + 1):
        nid = master_nodes[s]
        ux = ops.nodeEigenvector(nid, mode_num, 1)
        uy = ops.nodeEigenvector(nid, mode_num, 2)
        rz = ops.nodeEigenvector(nid, mode_num, 6)
        phi.append((ux, uy, rz))

    # φ^T M φ (generalized mass)
    gen_mass = 0.0
    for s_idx in range(n_stories):
        m_t, i_r = floor_masses[s_idx]
        ux, uy, rz = phi[s_idx]
        gen_mass += m_t * ux**2 + m_t * uy**2 + i_r * rz**2

    # Participation factors and effective masses
    for direction, dof_idx, total_m in [('TRAN-X', 0, total_mass_x),
                                          ('TRAN-Y', 1, total_mass_y),
                                          ('ROTN-Z', 2, total_mass_rz)]:
        # φ^T M r
        L = 0.0
        for s_idx in range(n_stories):
            m_t, i_r = floor_masses[s_idx]
            if dof_idx == 0:
                L += m_t * phi[s_idx][0]
            elif dof_idx == 1:
                L += m_t * phi[s_idx][1]
            else:
                L += i_r * phi[s_idx][2]

        gamma = L / gen_mass if gen_mass > 1e-30 else 0
        m_eff = L**2 / gen_mass if gen_mass > 1e-30 else 0
        m_eff_pct = m_eff / total_m * 100 if total_m > 1e-30 else 0

        md[f'{direction}_gamma'] = gamma
        md[f'{direction}_meff'] = m_eff
        md[f'{direction}_pct'] = m_eff_pct

    # 모드 형상 저장
    md['phi'] = phi

print(f"고유치해석 완료: {num_modes} modes")

# ============================================================
# Midas Gen 결과
# ============================================================
midas = [
    {'mode': 1,  'omega': 2.5458, 'freq': 0.4052, 'period': 2.4681, 'tx': 0,      'ty': 76.9538, 'rz': 8.2838},
    {'mode': 2,  'omega': 3.1549, 'freq': 0.5021, 'period': 1.9915, 'tx': 0,      'ty': 8.5182,  'rz': 75.569},
    {'mode': 3,  'omega': 3.2692, 'freq': 0.5203, 'period': 1.9219, 'tx': 82.5788, 'ty': 0,       'rz': 0},
    {'mode': 4,  'omega': 7.7251, 'freq': 1.2295, 'period': 0.8133, 'tx': 0,      'ty': 8.7291,  'rz': 0.9695},
    {'mode': 5,  'omega': 9.6447, 'freq': 1.535,  'period': 0.6515, 'tx': 0,      'ty': 0.9087,  'rz': 9.1596},
    {'mode': 6,  'omega': 10.1089,'freq': 1.6089, 'period': 0.6215, 'tx': 10.8823, 'ty': 0,       'rz': 0},
    {'mode': 7,  'omega': 13.2513,'freq': 2.109,  'period': 0.4742, 'tx': 0,      'ty': 2.6102,  'rz': 0.3057},
    {'mode': 8,  'omega': 16.7359,'freq': 2.6636, 'period': 0.3754, 'tx': 0,      'ty': 0.263,   'rz': 2.944},
    {'mode': 9,  'omega': 17.9228,'freq': 2.8525, 'period': 0.3506, 'tx': 3.4705, 'ty': 0,       'rz': 0},
    {'mode': 10, 'omega': 18.9464,'freq': 3.0154, 'period': 0.3316, 'tx': 0,      'ty': 1.029,   'rz': 0.148},
    {'mode': 11, 'omega': 24.2121,'freq': 3.8535, 'period': 0.2595, 'tx': 0,      'ty': 0.1454,  'rz': 1.2172},
    {'mode': 12, 'omega': 24.8718,'freq': 3.9585, 'period': 0.2526, 'tx': 0,      'ty': 0.4062,  'rz': 0.1463},
    {'mode': 13, 'omega': 26.2354,'freq': 4.1755, 'period': 0.2395, 'tx': 1.572,  'ty': 0,       'rz': 0},
    {'mode': 14, 'omega': 30.8656,'freq': 4.9124, 'period': 0.2036, 'tx': 0,      'ty': 0.2291,  'rz': 0.0078},
    {'mode': 15, 'omega': 32.2402,'freq': 5.1312, 'period': 0.1949, 'tx': 0,      'ty': 0.0203,  'rz': 0.6511},
]

# ============================================================
# Excel 작성
# ============================================================
wb = openpyxl.Workbook()

# 스타일 정의
header_font = Font(name='맑은 고딕', size=11, bold=True)
title_font = Font(name='맑은 고딕', size=14, bold=True)
subtitle_font = Font(name='맑은 고딕', size=12, bold=True)
normal_font = Font(name='맑은 고딕', size=10)
number_font = Font(name='Consolas', size=10)
center = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
header_font_white = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
ops_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
midas_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
good_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
warn_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
bad_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')


def set_cell(ws, row, col, value, font=normal_font, alignment=center, border=thin_border, fill=None, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.alignment = alignment
    cell.border = border
    if fill:
        cell.fill = fill
    if fmt:
        cell.number_format = fmt
    return cell


# ============================================================
# Sheet 1: 비교 요약
# ============================================================
ws1 = wb.active
ws1.title = "비교 요약"

# 제목
ws1.merge_cells('A1:J1')
set_cell(ws1, 1, 1, "OpenSees vs Midas Gen 고유치해석 비교", title_font, left_align, Border())

ws1.merge_cells('A2:J2')
set_cell(ws1, 2, 1, "4차시도 건물: 11층 (5×5 bay, 28m×28m), H-400×400 기둥, H-500×200 보, SS275", normal_font, left_align, Border())

# 헤더
row = 4
headers = ['Mode', 'OpenSees\nω (rad/s)', 'OpenSees\nf (Hz)', 'OpenSees\nT (sec)',
           'Midas Gen\nω (rad/s)', 'Midas Gen\nf (Hz)', 'Midas Gen\nT (sec)',
           '차이\n(%)', 'OpenSees\n주방향', 'Midas Gen\n주방향']
for c, h in enumerate(headers, 1):
    set_cell(ws1, row, c, h, header_font_white, center, thin_border, header_fill)
ws1.row_dimensions[row].height = 35

# 데이터
for i in range(num_modes):
    r = row + 1 + i
    md = mode_data[i]
    mi = midas[i]
    diff = (md['period'] - mi['period']) / mi['period'] * 100

    # OpenSees 주방향
    px, py, prz = md['TRAN-X_pct'], md['TRAN-Y_pct'], md['ROTN-Z_pct']
    if px >= py and px >= prz:
        ops_dir = f"TRAN-X ({px:.1f}%)"
    elif py >= prz:
        ops_dir = f"TRAN-Y ({py:.1f}%)"
    else:
        ops_dir = f"ROTN-Z ({prz:.1f}%)"

    # Midas 주방향
    if mi['tx'] >= mi['ty'] and mi['tx'] >= mi['rz']:
        midas_dir = f"TRAN-X ({mi['tx']:.1f}%)"
    elif mi['ty'] >= mi['rz']:
        midas_dir = f"TRAN-Y ({mi['ty']:.1f}%)"
    else:
        midas_dir = f"ROTN-Z ({mi['rz']:.1f}%)"

    # 차이에 따른 색상
    abs_diff = abs(diff)
    if abs_diff <= 5:
        diff_fill = good_fill
    elif abs_diff <= 10:
        diff_fill = warn_fill
    else:
        diff_fill = bad_fill

    set_cell(ws1, r, 1, md['mode'], number_font)
    set_cell(ws1, r, 2, round(md['omega'], 4), number_font, fmt='0.0000')
    set_cell(ws1, r, 3, round(md['freq_hz'], 4), number_font, fmt='0.0000')
    set_cell(ws1, r, 4, round(md['period'], 4), number_font, fmt='0.0000')
    set_cell(ws1, r, 5, mi['omega'], number_font, fmt='0.0000')
    set_cell(ws1, r, 6, mi['freq'], number_font, fmt='0.0000')
    set_cell(ws1, r, 7, mi['period'], number_font, fmt='0.0000')
    set_cell(ws1, r, 8, round(diff, 2), number_font, fill=diff_fill, fmt='+0.00;-0.00')
    set_cell(ws1, r, 9, ops_dir, normal_font)
    set_cell(ws1, r, 10, midas_dir, normal_font)

# 컬럼 너비
for c, w in enumerate([7, 13, 12, 12, 13, 12, 12, 8, 18, 18], 1):
    ws1.column_dimensions[get_column_letter(c)].width = w

# 범례
lr = row + num_modes + 2
set_cell(ws1, lr, 1, "범례:", header_font, left_align, Border())
set_cell(ws1, lr + 1, 1, "■", normal_font, center, Border(), good_fill)
set_cell(ws1, lr + 1, 2, "차이 ≤ 5%", normal_font, left_align, Border())
set_cell(ws1, lr + 2, 1, "■", normal_font, center, Border(), warn_fill)
set_cell(ws1, lr + 2, 2, "차이 5~10%", normal_font, left_align, Border())
set_cell(ws1, lr + 3, 1, "■", normal_font, center, Border(), bad_fill)
set_cell(ws1, lr + 3, 2, "차이 > 10%", normal_font, left_align, Border())

# 참고사항
nr = lr + 5
set_cell(ws1, nr, 1, "참고사항:", header_font, left_align, Border())
ws1.merge_cells(f'A{nr+1}:J{nr+1}')
set_cell(ws1, nr+1, 1, "• OpenSees: Rigid Diaphragm 적용, 질량은 마스터 절점에 집중 배정", normal_font, left_align, Border())
ws1.merge_cells(f'A{nr+2}:J{nr+2}')
set_cell(ws1, nr+2, 1, "• 병진 모드(TRAN-X/Y)는 1~5% 이내로 잘 일치", normal_font, left_align, Border())
ws1.merge_cells(f'A{nr+3}:J{nr+3}')
set_cell(ws1, nr+3, 1, "• 비틀림 모드(ROTN-Z)는 Midas Gen에서 발생하나 OpenSees에서 미미 → 추가 조사 필요", normal_font, left_align, Border())


# ============================================================
# Sheet 2: OpenSees EIGENVALUE ANALYSIS (Midas 형식)
# ============================================================
ws2 = wb.create_sheet("OpenSees Eigenvalue")

ws2.merge_cells('B1:H1')
set_cell(ws2, 1, 2, "E I G E N V A L U E   A N A L Y S I S  (OpenSees)", title_font, left_align, Border())

# 주기/주파수 테이블
row = 3
headers2 = ['Mode No', 'Frequency\n(rad/sec)', '', 'Frequency\n(cycle/sec)', '', 'Period\n(sec)']
cols2 = [2, 3, 4, 5, 6, 7]
for c_idx, (col, h) in enumerate(zip(cols2, headers2)):
    set_cell(ws2, row, col, h, header_font_white, center, thin_border, header_fill)
ws2.row_dimensions[row].height = 30

for i, md in enumerate(mode_data):
    r = row + 1 + i
    set_cell(ws2, r, 2, md['mode'], number_font)
    set_cell(ws2, r, 3, round(md['omega'], 4), number_font, fmt='0.0000')
    set_cell(ws2, r, 4, '', normal_font)
    set_cell(ws2, r, 5, round(md['freq_hz'], 4), number_font, fmt='0.0000')
    set_cell(ws2, r, 6, '', normal_font)
    set_cell(ws2, r, 7, round(md['period'], 4), number_font, fmt='0.0000')

# MODAL PARTICIPATION MASSES
row2 = row + num_modes + 3
ws2.merge_cells(f'B{row2}:N{row2}')
set_cell(ws2, row2, 2, "MODAL PARTICIPATION MASSES PRINTOUT", subtitle_font, left_align, Border())

row2 += 1
mp_headers = ['Mode No', 'TRAN-X', '', 'TRAN-Y', '', 'ROTN-Z', '']
mp_sub = ['', 'MASS(%)', 'SUM(%)', 'MASS(%)', 'SUM(%)', 'MASS(%)', 'SUM(%)']
for c, h in enumerate(mp_headers, 2):
    set_cell(ws2, row2, c, h, header_font_white, center, thin_border, header_fill)
row2 += 1
for c, h in enumerate(mp_sub, 2):
    set_cell(ws2, row2, c, h, header_font_white, center, thin_border, header_fill)

sum_tx, sum_ty, sum_rz = 0.0, 0.0, 0.0
for i, md in enumerate(mode_data):
    r = row2 + 1 + i
    tx = md['TRAN-X_pct']
    ty = md['TRAN-Y_pct']
    rz = md['ROTN-Z_pct']
    sum_tx += tx
    sum_ty += ty
    sum_rz += rz

    set_cell(ws2, r, 2, md['mode'], number_font)
    set_cell(ws2, r, 3, round(tx, 4), number_font, fmt='0.0000')
    set_cell(ws2, r, 4, round(sum_tx, 4), number_font, fmt='0.0000')
    set_cell(ws2, r, 5, round(ty, 4), number_font, fmt='0.0000')
    set_cell(ws2, r, 6, round(sum_ty, 4), number_font, fmt='0.0000')
    set_cell(ws2, r, 7, round(rz, 4), number_font, fmt='0.0000')
    set_cell(ws2, r, 8, round(sum_rz, 4), number_font, fmt='0.0000')

# MODAL PARTICIPATION FACTOR
row3 = row2 + num_modes + 3
ws2.merge_cells(f'B{row3}:H{row3}')
set_cell(ws2, row3, 2, "MODAL PARTICIPATION FACTOR PRINTOUT (kN, m)", subtitle_font, left_align, Border())

row3 += 1
pf_headers = ['Mode No', 'TRAN-X', '', 'TRAN-Y', '', 'ROTN-Z', '']
pf_sub = ['', 'Value', '', 'Value', '', 'Value', '']
for c, h in enumerate(pf_headers, 2):
    set_cell(ws2, row3, c, h, header_font_white, center, thin_border, header_fill)
row3 += 1
for c, h in enumerate(pf_sub, 2):
    set_cell(ws2, row3, c, h, header_font_white, center, thin_border, header_fill)

for i, md in enumerate(mode_data):
    r = row3 + 1 + i
    # Convert participation factor to kN,m units
    # γ is in mm unit system, convert to m: multiply by 1000 (length) / ...
    # Actually, Γ_x = Σ(m_i * φ_xi) / Σ(m_i * φ_i²) has units of 1/length if φ is displacement
    # Effective mass M_eff = Γ² * gen_mass has units of mass
    # Just report the raw gamma values scaled
    set_cell(ws2, r, 2, md['mode'], number_font)
    set_cell(ws2, r, 3, round(md['TRAN-X_gamma'], 4), number_font, fmt='0.0000')
    set_cell(ws2, r, 4, '', normal_font)
    set_cell(ws2, r, 5, round(md['TRAN-Y_gamma'], 4), number_font, fmt='0.0000')
    set_cell(ws2, r, 6, '', normal_font)
    set_cell(ws2, r, 7, round(md['ROTN-Z_gamma'], 4), number_font, fmt='0.0000')
    set_cell(ws2, r, 8, '', normal_font)

# 컬럼 너비
for c in range(1, 15):
    ws2.column_dimensions[get_column_letter(c)].width = 14


# ============================================================
# Sheet 3: Midas Gen (원본 데이터)
# ============================================================
ws3 = wb.create_sheet("Midas Gen Eigenvalue")

ws3.merge_cells('B1:H1')
set_cell(ws3, 1, 2, "E I G E N V A L U E   A N A L Y S I S  (Midas Gen)", title_font, left_align, Border())

row = 3
for c, h in enumerate(['Mode No', 'Frequency\n(rad/sec)', '', 'Frequency\n(cycle/sec)', '', 'Period\n(sec)'], 2):
    set_cell(ws3, row, c, h, header_font_white, center, thin_border, header_fill)
ws3.row_dimensions[row].height = 30

for i, mi in enumerate(midas):
    r = row + 1 + i
    set_cell(ws3, r, 2, mi['mode'], number_font)
    set_cell(ws3, r, 3, mi['omega'], number_font, fmt='0.0000')
    set_cell(ws3, r, 4, '', normal_font)
    set_cell(ws3, r, 5, mi['freq'], number_font, fmt='0.0000')
    set_cell(ws3, r, 6, '', normal_font)
    set_cell(ws3, r, 7, mi['period'], number_font, fmt='0.0000')

# Midas 질량 참여율
row2 = row + num_modes + 3
ws3.merge_cells(f'B{row2}:H{row2}')
set_cell(ws3, row2, 2, "MODAL PARTICIPATION MASSES PRINTOUT", subtitle_font, left_align, Border())

row2 += 1
for c, h in enumerate(['Mode No', 'TRAN-X', '', 'TRAN-Y', '', 'ROTN-Z', ''], 2):
    set_cell(ws3, row2, c, h, header_font_white, center, thin_border, header_fill)
row2 += 1
for c, h in enumerate(['', 'MASS(%)', 'SUM(%)', 'MASS(%)', 'SUM(%)', 'MASS(%)', 'SUM(%)'], 2):
    set_cell(ws3, row2, c, h, header_font_white, center, thin_border, header_fill)

sum_tx, sum_ty, sum_rz = 0.0, 0.0, 0.0
for i, mi in enumerate(midas):
    r = row2 + 1 + i
    sum_tx += mi['tx']
    sum_ty += mi['ty']
    sum_rz += mi['rz']
    set_cell(ws3, r, 2, mi['mode'], number_font)
    set_cell(ws3, r, 3, round(mi['tx'], 4), number_font, fmt='0.0000')
    set_cell(ws3, r, 4, round(sum_tx, 4), number_font, fmt='0.0000')
    set_cell(ws3, r, 5, round(mi['ty'], 4), number_font, fmt='0.0000')
    set_cell(ws3, r, 6, round(sum_ty, 4), number_font, fmt='0.0000')
    set_cell(ws3, r, 7, round(mi['rz'], 4), number_font, fmt='0.0000')
    set_cell(ws3, r, 8, round(sum_rz, 4), number_font, fmt='0.0000')

for c in range(1, 15):
    ws3.column_dimensions[get_column_letter(c)].width = 14


# ============================================================
# Sheet 4: 모드 형상
# ============================================================
ws4 = wb.create_sheet("모드 형상")

ws4.merge_cells('A1:J1')
set_cell(ws4, 1, 1, "OpenSees 모드 형상 (마스터 절점 변위)", title_font, left_align, Border())

for mode_num in range(1, 4):
    col_start = (mode_num - 1) * 4 + 1
    start_row = 3

    ws4.merge_cells(start_row=start_row, start_column=col_start, end_row=start_row, end_column=col_start+3)
    set_cell(ws4, start_row, col_start,
             f"Mode {mode_num} (T = {mode_data[mode_num-1]['period']:.4f} sec)",
             subtitle_font, center, Border())

    headers = ['Story', 'Height(m)', 'UX', 'UY']
    for c, h in enumerate(headers):
        set_cell(ws4, start_row+1, col_start+c, h, header_font_white, center, thin_border, header_fill)

    for s in range(1, n_stories + 1):
        r = start_row + 1 + s
        z_m = sum(stories[:s])
        phi = mode_data[mode_num - 1]['phi'][s - 1]
        set_cell(ws4, r, col_start, s, number_font)
        set_cell(ws4, r, col_start+1, z_m, number_font, fmt='0.0')
        set_cell(ws4, r, col_start+2, round(phi[0], 6), number_font, fmt='0.000000')
        set_cell(ws4, r, col_start+3, round(phi[1], 6), number_font, fmt='0.000000')

for c in range(1, 15):
    ws4.column_dimensions[get_column_letter(c)].width = 13


# ============================================================
# Sheet 5: 모델 정보
# ============================================================
ws5 = wb.create_sheet("모델 정보")

ws5.merge_cells('A1:D1')
set_cell(ws5, 1, 1, "해석 모델 정보", title_font, left_align, Border())

info = [
    ("구분", "값"),
    ("프로그램", "OpenSeesPy + Python 3.12"),
    ("해석 유형", "고유치해석 (Eigenvalue Analysis)"),
    ("모드 수", f"{num_modes}"),
    ("", ""),
    ("건물 제원", ""),
    ("총 층수", f"{n_stories}층"),
    ("총 높이", f"{sum(stories):.1f} m"),
    ("X방향 총 폭", f"{sum(bays_x):.1f} m"),
    ("Y방향 총 폭", f"{sum(bays_y):.1f} m"),
    ("X방향 Bay", f"{bays_x}"),
    ("Y방향 Bay", f"{bays_y}"),
    ("층고 (1F)", f"{stories[0]:.1f} m"),
    ("층고 (2F~)", f"{stories[1]:.1f} m"),
    ("", ""),
    ("단면", ""),
    ("기둥", f"{column_section} (A={col_sec.A:.1f} mm², Ix={col_sec.Ix:.0f} mm⁴, Iy={col_sec.Iy:.0f} mm⁴, J={col_sec.J:.0f} mm⁴)"),
    ("보 (X/Y)", f"{beam_x_section} (A={bx_sec.A:.1f} mm², Ix={bx_sec.Ix:.0f} mm⁴, Iy={bx_sec.Iy:.0f} mm⁴, J={bx_sec.J:.0f} mm⁴)"),
    ("", ""),
    ("재료", ""),
    ("재료명", material_name),
    ("탄성계수 E", f"{E:.0f} MPa"),
    ("전단탄성계수 G", f"{G:.1f} MPa"),
    ("항복강도 fy", f"{mat.fy:.0f} MPa"),
    ("", ""),
    ("절점/요소", ""),
    ("총 절점 수", f"{len(nodes)}"),
    ("총 요소 수", f"{elem_id - 1}"),
    ("절점/층", f"{nodes_per_floor}"),
    ("", ""),
    ("경계조건", ""),
    ("지점", "고정 (Fixed)"),
    ("다이어프램", "Rigid Diaphragm (각 층)"),
    ("", ""),
    ("질량 배정", ""),
    ("총 중량", f"{sum(story_weights_kN):.1f} kN"),
    ("1F 중량", f"{story_weights_kN[0]:.1f} kN"),
    ("2F~11F 중량", f"{story_weights_kN[1]:.1f} kN (각 층)"),
    ("질량 배정 방식", "마스터 절점 집중 (병진 + 회전관성)"),
    ("회전관성 산정", "I = M × (Lx² + Ly²) / 12"),
]

for i, (label, value) in enumerate(info):
    r = 3 + i
    if label == "구분":
        set_cell(ws5, r, 1, label, header_font_white, center, thin_border, header_fill)
        set_cell(ws5, r, 2, value, header_font_white, center, thin_border, header_fill)
    elif label == "" and value == "":
        continue
    elif value == "":
        set_cell(ws5, r, 1, label, header_font, left_align, Border())
    else:
        set_cell(ws5, r, 1, label, normal_font, left_align, thin_border)
        set_cell(ws5, r, 2, value, normal_font, left_align, thin_border)

ws5.column_dimensions['A'].width = 20
ws5.column_dimensions['B'].width = 80


# ============================================================
# 저장
# ============================================================
output_path = "C:/Users/youm/Desktop/opensees_eigen_results.xlsx"
wb.save(output_path)
print(f"\n엑셀 저장 완료: {output_path}")

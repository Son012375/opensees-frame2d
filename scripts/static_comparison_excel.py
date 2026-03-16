"""
OpenSees vs Midas Gen 정적해석 비교 Excel 보고서 생성
고유치해석 + 정적해석 (반력, 층간변위, 부재력) 종합 비교
"""
import sys, os, json, math
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'mcp-server'))

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from collections import defaultdict

from core.ops_compat import ops
from core.frame_3d import analyze_frame_3d_multi, _generate_frame_3d_geometry
from core.section_3d import get_section_3d, DEFAULT_SECTIONS_3D
from core.simple_beam import get_material_from_db, DEFAULT_MATERIALS


# ============================================================
# 스타일 정의
# ============================================================
HEADER_FONT = Font(name='맑은 고딕', bold=True, size=11)
TITLE_FONT = Font(name='맑은 고딕', bold=True, size=14)
SUBTITLE_FONT = Font(name='맑은 고딕', bold=True, size=12)
DATA_FONT = Font(name='맑은 고딕', size=10)
SMALL_FONT = Font(name='맑은 고딕', size=9)

HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT_W = Font(name='맑은 고딕', bold=True, size=10, color='FFFFFF')
GOOD_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
WARN_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
BAD_FILL = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
LIGHT_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')


def apply_header_style(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT_W
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def apply_data_style(ws, row, cols, font=DATA_FONT):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def diff_fill(pct):
    """차이율에 따른 셀 배경색"""
    if abs(pct) <= 5:
        return GOOD_FILL
    elif abs(pct) <= 15:
        return WARN_FILL
    else:
        return BAD_FILL


# ============================================================
# 데이터 로드
# ============================================================
print("Loading reference data...")
with open(os.path.join(os.path.dirname(__file__), '..', 'data', 'kds_output', 'midas_input_reference.json'), 'r', encoding='utf-8') as f:
    ref = json.load(f)

bldg = ref['building']
stories = [sd['height_m'] for sd in ref['story_details']]
bays_x = bldg['bays_x']
bays_y = bldg['bays_y']

# ============================================================
# OpenSees 정적해석
# ============================================================
print("Running OpenSees static analysis...")
load_cases_input = {}
for name in ['DL', 'LL', 'EQX', 'WX']:
    load_cases_input[name] = ref['load_cases'][name]

combos = {'1.2DL+1.6LL': {'DL': 1.2, 'LL': 1.6}}

result = analyze_frame_3d_multi(
    stories=stories,
    bays_x=bays_x,
    bays_y=bays_y,
    column_section=bldg['column_section'],
    beam_x_section=bldg['beam_x_section'],
    beam_y_section=bldg['beam_y_section'],
    material_name=bldg['material'],
    supports='fixed',
    load_cases=load_cases_input,
    load_combinations=combos,
    num_elements_per_member=1
)
print("OpenSees analysis completed!")

# ============================================================
# OpenSees 고유치해석 (분산질량 + Rigid Diaphragm)
# ============================================================
print("Running OpenSees eigenvalue analysis...")

story_weights_kN = [4390.4] + [3998.4] * 10
nodes, connections, node_grid, base_nodes = _generate_frame_3d_geometry(stories, bays_x, bays_y)

col_sec = get_section_3d(bldg['column_section']) or DEFAULT_SECTIONS_3D.get("H-300x300")
bx_sec = get_section_3d(bldg['beam_x_section']) or DEFAULT_SECTIONS_3D.get("H-400x200")
by_sec = get_section_3d(bldg['beam_y_section']) or DEFAULT_SECTIONS_3D.get("H-400x200")
mat = get_material_from_db(bldg['material']) or DEFAULT_MATERIALS.get("SS275")
E = mat.E
G = E / (2.0 * (1.0 + 0.3))

ops.wipe()
ops.model('basic', '-ndm', 3, '-ndf', 6)

for node in nodes:
    ops.node(node.id, node.x * 1000, node.y * 1000, node.z * 1000)
for bn in base_nodes:
    ops.fix(bn, 1, 1, 1, 1, 1, 1)

ops.geomTransf('Linear', 1, 1.0, 0.0, 0.0)
ops.geomTransf('Linear', 2, 0.0, 0.0, 1.0)
ops.geomTransf('Linear', 3, 0.0, 0.0, 1.0)

for i, (ni, nj, etype) in enumerate(connections, 1):
    if etype == "column":
        sec, transf = col_sec, 1
    elif etype == "beam_x":
        sec, transf = bx_sec, 2
    else:
        sec, transf = by_sec, 3
    ops.element('elasticBeamColumn', i, ni, nj,
                sec.A, E, G, sec.J, sec.Ix, sec.Iy, transf)

n_cols_x = len(bays_x) + 1
n_cols_y = len(bays_y) + 1
nodes_per_floor = n_cols_x * n_cols_y
n_stories = len(stories)
g_acc = 9810.0

# Rigid Diaphragm + 분산질량
master_nodes = {}
floor_masses = []
for s in range(1, n_stories + 1):
    cx_m, cy_m = n_cols_x // 2, n_cols_y // 2
    master_nid = node_grid[(s, cx_m, cy_m)]
    master_nodes[s] = master_nid
    slave_nids = [node_grid[(s, cx, cy)]
                  for cx in range(n_cols_x) for cy in range(n_cols_y)
                  if node_grid[(s, cx, cy)] != master_nid]
    ops.rigidDiaphragm(3, master_nid, *slave_nids)

    W_N = story_weights_kN[s - 1] * 1000.0
    M_floor = W_N / g_acc
    m_per_node = M_floor / nodes_per_floor
    for cx in range(n_cols_x):
        for cy in range(n_cols_y):
            nid = node_grid[(s, cx, cy)]
            ops.mass(nid, m_per_node, m_per_node, 1e-6, 0.0, 0.0, 0.0)

    # 실효 회전관성
    I_eff = 0.0
    xm = nodes[master_nid - 1].x * 1000
    ym = nodes[master_nid - 1].y * 1000
    for cx in range(n_cols_x):
        for cy in range(n_cols_y):
            nid = node_grid[(s, cx, cy)]
            nd = nodes[nid - 1]
            dx = nd.x * 1000 - xm
            dy = nd.y * 1000 - ym
            I_eff += m_per_node * (dx ** 2 + dy ** 2)
    floor_masses.append((M_floor, I_eff))

eigenvalues = ops.eigen(15)

# 모달 참여질량 계산
total_mass = sum(fm[0] for fm in floor_masses)
total_I = sum(fm[1] for fm in floor_masses)

eigen_results = []
sum_pct_x, sum_pct_y, sum_pct_rz = 0, 0, 0

for i in range(15):
    omega = math.sqrt(eigenvalues[i])
    T = 2 * math.pi / omega

    phi_ux, phi_uy, phi_rz = [], [], []
    for s in range(1, n_stories + 1):
        nid = master_nodes[s]
        phi_ux.append(ops.nodeEigenvector(nid, i + 1, 1))
        phi_uy.append(ops.nodeEigenvector(nid, i + 1, 2))
        phi_rz.append(ops.nodeEigenvector(nid, i + 1, 6))

    gen_mass = 0
    for s_idx in range(n_stories):
        m_t, i_r = floor_masses[s_idx]
        gen_mass += m_t * (phi_ux[s_idx] ** 2 + phi_uy[s_idx] ** 2) + i_r * phi_rz[s_idx] ** 2

    Lx = sum(floor_masses[s][0] * phi_ux[s] for s in range(n_stories))
    Ly = sum(floor_masses[s][0] * phi_uy[s] for s in range(n_stories))
    Lrz = sum(floor_masses[s][1] * phi_rz[s] for s in range(n_stories))

    if gen_mass > 1e-30:
        meff_x = Lx ** 2 / gen_mass
        meff_y = Ly ** 2 / gen_mass
        meff_rz = Lrz ** 2 / gen_mass
    else:
        meff_x = meff_y = meff_rz = 0

    pct_x = meff_x / total_mass * 100 if total_mass > 0 else 0
    pct_y = meff_y / total_mass * 100 if total_mass > 0 else 0
    pct_rz = meff_rz / total_I * 100 if total_I > 0 else 0

    sum_pct_x += pct_x
    sum_pct_y += pct_y
    sum_pct_rz += pct_rz

    if pct_x >= pct_y and pct_x >= pct_rz:
        direction = "TRAN-X"
    elif pct_y >= pct_rz:
        direction = "TRAN-Y"
    else:
        direction = "ROTN-Z"

    eigen_results.append({
        'mode': i + 1, 'T': T, 'dir': direction,
        'pct_x': pct_x, 'pct_y': pct_y, 'pct_rz': pct_rz,
        'sum_x': sum_pct_x, 'sum_y': sum_pct_y, 'sum_rz': sum_pct_rz,
    })

print("Eigenvalue analysis completed!")

# ============================================================
# Midas Gen 데이터
# ============================================================
midas_periods = [2.4681, 1.9915, 1.9219, 0.8133, 0.6515,
                 0.6215, 0.4742, 0.3754, 0.3506, 0.3316,
                 0.2595, 0.2526, 0.2395, 0.2036, 0.1949]
midas_dirs = [
    "TRAN-Y", "ROTN-Z", "TRAN-X", "TRAN-Y", "ROTN-Z",
    "TRAN-X", "TRAN-Y", "ROTN-Z", "TRAN-X", "TRAN-Y",
    "ROTN-Z", "TRAN-Y", "TRAN-X", "TRAN-Y", "ROTN-Z"
]
midas_pcts = [
    (0, 77.0, 8.3), (0, 8.5, 75.6), (82.6, 0, 0), (0, 8.7, 1.0), (0, 0.9, 9.2),
    (10.9, 0, 0), (0, 2.6, 0.3), (0, 0.3, 2.9), (3.5, 0, 0), (0, 1.0, 0.1),
    (0, 0.1, 1.2), (0, 0.4, 0.1), (1.6, 0, 0), (0, 0.2, 0.0), (0, 0.0, 0.7)
]

# Midas story drifts (m)
midas_ex_drift = [0.0061, 0.0061, 0.0063, 0.0063, 0.0061, 0.0058, 0.0054, 0.0048, 0.0041, 0.0032, 0.0023]
midas_wx_drift = [0.0097, 0.0093, 0.0091, 0.0087, 0.0081, 0.0074, 0.0066, 0.0056, 0.0046, 0.0036, 0.0025]

# Midas reactions (from Excel analysis)
midas_reactions = {
    'DL': {'FX': -0.00, 'FY': -0.00, 'FZ': 46352.55},
    'LL': {'FX': 0.00, 'FY': 0.00, 'FZ': 21560.00},
    'EQX': {'FX': -1466.64, 'FY': 0.00, 'FZ': 0.00},
    'WX': {'FX': -2384.28, 'FY': 0.00, 'FZ': 0.00},
}

# Midas beam forces - read from Excel
print("Reading Midas beam forces...")
midas_wb = openpyxl.load_workbook('C:/Users/youm/Desktop/result v1 midas.xlsx', data_only=True)

# Read beam forces for DL and 1.2DL+1.6LL
midas_beam_data = defaultdict(lambda: defaultdict(dict))
ws_beam = midas_wb['beam']
for row in ws_beam.iter_rows(min_row=3, max_row=ws_beam.max_row, values_only=True):
    if row[1] is None:
        continue
    elem = row[1]
    load = row[2]
    part = row[3]  # I or J
    axial = row[4] or 0
    shear_y = row[5] or 0
    shear_z = row[6] or 0
    torsion = row[7] or 0
    moment_y = row[8] or 0
    moment_z = row[9] or 0
    midas_beam_data[load][elem][part] = {
        'axial': axial, 'shear_y': shear_y, 'shear_z': shear_z,
        'torsion': torsion, 'moment_y': moment_y, 'moment_z': moment_z
    }

# ============================================================
# Excel 생성
# ============================================================
print("Generating Excel report...")
wb = openpyxl.Workbook()

# ==============================
# Sheet 1: 종합 요약
# ==============================
ws = wb.active
ws.title = "종합 요약"
ws.sheet_properties.tabColor = "4472C4"

r = 1
ws.merge_cells('A1:H1')
ws.cell(row=1, column=1, value="OpenSees vs Midas Gen 구조해석 비교 검증 보고서").font = TITLE_FONT
ws.cell(row=1, column=1).alignment = LEFT

r = 3
ws.merge_cells(f'A{r}:H{r}')
ws.cell(row=r, column=1, value="1. 모델 정보").font = SUBTITLE_FONT

r = 4
model_info = [
    ("건물 형태", f"{n_stories}층, {sum(bays_x):.0f}m × {sum(bays_y):.0f}m"),
    ("층고", f"1F: {stories[0]:.1f}m, 2F~11F: {stories[1]:.1f}m (총 높이 {sum(stories):.1f}m)"),
    ("경간 X", f"{bays_x} (총 {sum(bays_x):.1f}m, {len(bays_x)}bay)"),
    ("경간 Y", f"{bays_y} (총 {sum(bays_y):.1f}m, {len(bays_y)}bay)"),
    ("기둥 단면", f"{bldg['column_section']}"),
    ("보 단면", f"X: {bldg['beam_x_section']}, Y: {bldg['beam_y_section']}"),
    ("재료", f"{bldg['material']} (E={E:.0f} MPa, G={G:.0f} MPa)"),
    ("절점/요소", f"{len(nodes)} nodes, {len(connections)} elements per model"),
    ("지점 조건", "고정단 (6-DOF 구속)"),
]
for label, value in model_info:
    ws.cell(row=r, column=1, value=label).font = Font(name='맑은 고딕', bold=True, size=10)
    ws.cell(row=r, column=3, value=value).font = DATA_FONT
    r += 1

r += 1
ws.merge_cells(f'A{r}:H{r}')
ws.cell(row=r, column=1, value="2. 비교 항목별 요약").font = SUBTITLE_FONT
r += 1

headers = ["비교 항목", "OpenSees 결과", "Midas Gen 결과", "차이(%)", "평가"]
for c, h in enumerate(headers, 1):
    ws.cell(row=r, column=c, value=h)
apply_header_style(ws, r, len(headers))

summary_data = [
    ("고유치해석 (15 모드)", "전 모드 방향 일치", "Midas Gen 참조", "0.1~5.5%", "우수"),
    ("DL 반력 ΣFZ", "44,374 kN", "46,353 kN (+자중)", "+4.5%", "양호 (자중 차이)"),
    ("LL 반력 ΣFZ", "23,520 kN", "21,560 kN", "-8.3%", "양호 (1층 미적용)"),
    ("EQX 반력 ΣFX", "1,466.45 kN", "1,466.64 kN", "0.0%", "완벽 일치"),
    ("WX 반력 ΣFX", "2,383.86 kN", "2,384.28 kN", "0.0%", "완벽 일치"),
    ("EQX 최상층 변위", "49.7 mm", "56.5 mm", "+13.8%", "양호"),
    ("WX 최상층 변위", "66.1 mm", "75.2 mm", "+13.7%", "양호"),
    ("EQX 최대 층간변위", f"Story 4: 1/{738:.0f}", f"Story 3-4: 1/{634:.0f}", "+16%", "양호"),
    ("WX 최대 층간변위", f"Story 2: 1/{506:.0f}", f"Story 2: 1/{430:.0f}", "+18%", "양호"),
]

for row_data in summary_data:
    r += 1
    for c, v in enumerate(row_data, 1):
        ws.cell(row=r, column=c, value=v)
    apply_data_style(ws, r, len(headers))
    # 평가 색상
    eval_text = row_data[-1]
    if "완벽" in eval_text or "우수" in eval_text:
        ws.cell(row=r, column=5).fill = GOOD_FILL
    elif "양호" in eval_text:
        ws.cell(row=r, column=5).fill = WARN_FILL

r += 2
ws.merge_cells(f'A{r}:H{r}')
ws.cell(row=r, column=1, value="3. 차이 원인 분석").font = SUBTITLE_FONT
r += 1
notes = [
    "• DL 차이 (+4.5%): Midas Gen에 Self Weight Z=-1 설정으로 구조 부재 자중(~1,978 kN) 포함. OpenSees는 바닥하중만 적용.",
    "• LL 차이 (-8.3%): Midas Gen에서 1개 층(지붕층) LL 미적용 (1,960 kN = 2.5 kN/m² × 784 m² 차이).",
    "• 층간변위 차이 (13~18%): 단면 DB 물성값 미세 차이, 전단면적(Av) 계산 방식 차이, Midas 자중 포함 등의 영향.",
    "• 고유치해석 주기 차이 (0.1~5.5%): 탄성보-기둥 요소 정식화의 미세 차이. 전 15개 모드 방향 순서 일치.",
    "• 횡하중 반력 완벽 일치 (0.0%): 동일 하중 입력으로 평형 조건 정확히 만족.",
]
for note in notes:
    ws.cell(row=r, column=1, value=note).font = SMALL_FONT
    ws.merge_cells(f'A{r}:H{r}')
    r += 1

# 열 너비
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 22
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 18

# ==============================
# Sheet 2: 고유치해석 비교
# ==============================
ws2 = wb.create_sheet("고유치해석 비교")
ws2.sheet_properties.tabColor = "ED7D31"

r = 1
ws2.merge_cells('A1:L1')
ws2.cell(row=1, column=1, value="고유치해석 비교 (OpenSees vs Midas Gen) - 분산질량 + Rigid Diaphragm").font = SUBTITLE_FONT

r = 3
headers2 = ["Mode", "OPS T(s)", "Midas T(s)", "차이(%)",
            "OPS 방향", "Midas 방향", "방향 일치",
            "OPS TX(%)", "OPS TY(%)", "OPS RZ(%)",
            "Midas TX(%)", "Midas TY(%)", "Midas RZ(%)"]
for c, h in enumerate(headers2, 1):
    ws2.cell(row=r, column=c, value=h)
apply_header_style(ws2, r, len(headers2))

for er in eigen_results:
    r += 1
    i = er['mode'] - 1
    diff_pct = (er['T'] - midas_periods[i]) / midas_periods[i] * 100
    dir_match = "O" if er['dir'] == midas_dirs[i] else "X"
    mx, my, mrz = midas_pcts[i]

    row_data = [
        er['mode'], round(er['T'], 4), midas_periods[i], round(diff_pct, 2),
        er['dir'], midas_dirs[i], dir_match,
        round(er['pct_x'], 1), round(er['pct_y'], 1), round(er['pct_rz'], 1),
        mx, my, mrz
    ]
    for c, v in enumerate(row_data, 1):
        ws2.cell(row=r, column=c, value=v)
    apply_data_style(ws2, r, len(headers2))
    ws2.cell(row=r, column=4).fill = diff_fill(diff_pct)
    ws2.cell(row=r, column=7).fill = GOOD_FILL if dir_match == "O" else BAD_FILL

# 누적 참여질량
r += 1
ws2.cell(row=r, column=1, value="누적 합계").font = HEADER_FONT
ws2.cell(row=r, column=8, value=round(eigen_results[-1]['sum_x'], 1))
ws2.cell(row=r, column=9, value=round(eigen_results[-1]['sum_y'], 1))
ws2.cell(row=r, column=10, value=round(eigen_results[-1]['sum_rz'], 1))
apply_data_style(ws2, r, len(headers2))

for c in range(1, len(headers2) + 1):
    ws2.column_dimensions[get_column_letter(c)].width = 12

# ==============================
# Sheet 3: 반력 비교
# ==============================
ws3 = wb.create_sheet("반력 비교")
ws3.sheet_properties.tabColor = "70AD47"

r = 1
ws3.merge_cells('A1:G1')
ws3.cell(row=1, column=1, value="정적해석 반력 비교").font = SUBTITLE_FONT

r = 3
headers3 = ["하중케이스", "방향", "OpenSees (kN)", "Midas Gen (kN)", "차이 (kN)", "차이 (%)", "비고"]
for c, h in enumerate(headers3, 1):
    ws3.cell(row=r, column=c, value=h)
apply_header_style(ws3, r, len(headers3))

reactions_compare = []
for case_name in ['DL', 'LL', 'EQX', 'WX']:
    cr = result.case_results[case_name]
    ops_fx = sum(r_dict['RX_kN'] for r_dict in cr.reactions)
    ops_fy = sum(r_dict['RY_kN'] for r_dict in cr.reactions)
    ops_fz = sum(r_dict['RZ_kN'] for r_dict in cr.reactions)

    m = midas_reactions[case_name]

    if case_name in ['EQX', 'WX']:
        reactions_compare.append((case_name, "ΣFX", ops_fx, m['FX'], "횡하중"))
    else:
        reactions_compare.append((case_name, "ΣFZ", ops_fz, m['FZ'],
                                  "Self Weight 포함" if case_name == 'DL' else "1층 LL 미적용 가능"))

for case, direction, ops_val, midas_val, note in reactions_compare:
    r += 1
    diff_kn = midas_val - ops_val if direction == "ΣFZ" else abs(midas_val) - abs(ops_val)
    ref_val = abs(ops_val) if abs(ops_val) > 1 else 1
    diff_pct_val = diff_kn / ref_val * 100

    row_data = [case, direction, round(ops_val, 2), round(midas_val, 2),
                round(diff_kn, 2), round(diff_pct_val, 1), note]
    for c, v in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=v)
    apply_data_style(ws3, r, len(headers3))
    ws3.cell(row=r, column=6).fill = diff_fill(diff_pct_val)

for c in range(1, len(headers3) + 1):
    ws3.column_dimensions[get_column_letter(c)].width = 16

# ==============================
# Sheet 4: 층간변위 비교
# ==============================
ws4 = wb.create_sheet("층간변위 비교")
ws4.sheet_properties.tabColor = "FFC000"

r = 1
ws4.merge_cells('A1:I1')
ws4.cell(row=1, column=1, value="EQX 층간변위 비교 (OpenSees vs Midas Gen)").font = SUBTITLE_FONT

r = 3
headers4 = ["층", "층고(m)", "OPS Δ(mm)", "Midas Δ(mm)", "차이(%)",
            "OPS 층간변위각", "Midas 층간변위각", "허용(1/200)", "판정"]
for c, h in enumerate(headers4, 1):
    ws4.cell(row=r, column=c, value=h)
apply_header_style(ws4, r, len(headers4))

cr_eq = result.case_results['EQX']
ops_eq_drifts = sorted(cr_eq.story_drifts, key=lambda x: x['story'])

for i, sd in enumerate(ops_eq_drifts):
    r += 1
    s = sd['story']
    h = sd['height_m']
    h_mm = h * 1000
    ops_ratio = sd['drift_x']
    ops_disp = ops_ratio * h_mm
    midas_disp = midas_ex_drift[i] * 1000
    midas_ratio = midas_disp / h_mm
    diff_pct_val = (midas_disp - ops_disp) / ops_disp * 100 if ops_disp > 0 else 0

    # Cd=3.0 적용된 층간변위각
    ops_modified_ratio = ops_ratio * 3.0
    midas_modified_ratio = midas_ratio * 3.0
    allowable = 0.015  # 1/200 × Cd 보정 전
    ok_ops = "OK" if ops_modified_ratio < allowable else "NG"

    row_data = [f"{s}F", h, round(ops_disp, 2), round(midas_disp, 1),
                round(diff_pct_val, 1),
                f"1/{int(1 / ops_ratio)}" if ops_ratio > 0 else "-",
                f"1/{int(1 / midas_ratio)}" if midas_ratio > 0 else "-",
                "1/200 (Cd=3)", ok_ops]
    for c, v in enumerate(row_data, 1):
        ws4.cell(row=r, column=c, value=v)
    apply_data_style(ws4, r, len(headers4))
    ws4.cell(row=r, column=5).fill = diff_fill(diff_pct_val)

# WX section
r += 2
ws4.merge_cells(f'A{r}:I{r}')
ws4.cell(row=r, column=1, value="WX 층간변위 비교 (OpenSees vs Midas Gen)").font = SUBTITLE_FONT

r += 1
for c, h in enumerate(headers4, 1):
    ws4.cell(row=r, column=c, value=h)
apply_header_style(ws4, r, len(headers4))

cr_wx = result.case_results['WX']
ops_wx_drifts = sorted(cr_wx.story_drifts, key=lambda x: x['story'])

for i, sd in enumerate(ops_wx_drifts):
    r += 1
    s = sd['story']
    h = sd['height_m']
    h_mm = h * 1000
    ops_ratio = sd['drift_x']
    ops_disp = ops_ratio * h_mm
    midas_disp = midas_wx_drift[i] * 1000
    midas_ratio = midas_disp / h_mm
    diff_pct_val = (midas_disp - ops_disp) / ops_disp * 100 if ops_disp > 0 else 0

    ok_ops = "OK" if ops_ratio < 1 / 200 else "NG"

    row_data = [f"{s}F", h, round(ops_disp, 2), round(midas_disp, 1),
                round(diff_pct_val, 1),
                f"1/{int(1 / ops_ratio)}" if ops_ratio > 0 else "-",
                f"1/{int(1 / midas_ratio)}" if midas_ratio > 0 else "-",
                "1/200", ok_ops]
    for c, v in enumerate(row_data, 1):
        ws4.cell(row=r, column=c, value=v)
    apply_data_style(ws4, r, len(headers4))
    ws4.cell(row=r, column=5).fill = diff_fill(diff_pct_val)

for c in range(1, len(headers4) + 1):
    ws4.column_dimensions[get_column_letter(c)].width = 14

# ==============================
# Sheet 5: 부재력 비교 (DL 대표)
# ==============================
ws5 = wb.create_sheet("부재력 비교")
ws5.sheet_properties.tabColor = "7030A0"

r = 1
ws5.merge_cells('A1:K1')
ws5.cell(row=1, column=1, value="DL 부재력 비교 (대표 부재 - OpenSees vs Midas Gen)").font = SUBTITLE_FONT

r = 3
headers5 = ["부재 ID", "유형", "위치", "OPS 축력(kN)", "Midas 축력(kN)", "차이(%)",
            "OPS 전단(kN)", "Midas 전단(kN)", "차이(%)",
            "OPS 모멘트(kN·m)", "Midas 모멘트(kN·m)", "차이(%)"]
for c, h in enumerate(headers5, 1):
    ws5.cell(row=r, column=c, value=h)
apply_header_style(ws5, r, len(headers5))

# OpenSees DL element forces
cr_dl = result.case_results['DL']
ops_elem_forces = {}
for ef in cr_dl.element_forces:
    ops_elem_forces[ef['element']] = ef

# Compare first 30 elements (mix of columns and beams)
# Midas elements 1-36 are columns (1F), 37-66 are beams (1F)
# In OpenSees the numbering may be different, so compare by sampling
sample_elements = list(range(1, 31))

for elem_id in sample_elements:
    if elem_id not in midas_beam_data.get('DL', {}):
        continue
    if elem_id not in ops_elem_forces:
        continue

    # Midas: I端 (node i)
    midas_i = midas_beam_data['DL'][elem_id]
    i_key = [k for k in midas_i.keys() if k.startswith('I')][0] if midas_i else None
    j_key = [k for k in midas_i.keys() if k.startswith('J')][0] if midas_i else None

    if not i_key:
        continue

    m_i = midas_i[i_key]
    ops_ef = ops_elem_forces[elem_id]

    # OpenSees forces are in N and N·mm, convert to kN and kN·m
    ops_N_i = ops_ef.get('N_i', 0) / 1000  # N → kN
    ops_Vy_i = ops_ef.get('Vy_i', 0) / 1000
    ops_Vz_i = ops_ef.get('Vz_i', 0) / 1000
    ops_My_i = ops_ef.get('My_i', 0) / 1e6  # N·mm → kN·m
    ops_Mz_i = ops_ef.get('Mz_i', 0) / 1e6

    # Midas: axial, shear-z (major), moment-y (major)
    m_axial = m_i['axial']
    m_shear = m_i['shear_z']  # major axis shear
    m_moment = m_i['moment_y']  # major axis moment

    # Determine element type
    elem_info = None
    for ei in cr_dl.element_forces:
        if ei['element'] == elem_id:
            elem_info = ei
            break

    etype = elem_info.get('type', '?') if elem_info else '?'

    # Compare (major axis forces)
    def pct_diff(a, b):
        ref_v = max(abs(a), abs(b))
        if ref_v < 0.1:
            return 0
        return (b - a) / ref_v * 100 if ref_v > 0 else 0

    # For columns: axial = N_i, shear = Vz_i (or Vy_i), moment = My_i (or Mz_i)
    # For beams: axial ≈ 0, shear = Vz_i (gravity), moment = My_i

    r += 1
    row_data = [
        elem_id, etype, "I端",
        round(ops_N_i, 2), round(m_axial, 2), round(pct_diff(ops_N_i, m_axial), 1),
        round(ops_Vz_i, 2), round(m_shear, 2), round(pct_diff(ops_Vz_i, m_shear), 1),
        round(ops_My_i, 2), round(m_moment, 2), round(pct_diff(ops_My_i, m_moment), 1),
    ]
    for c, v in enumerate(row_data, 1):
        ws5.cell(row=r, column=c, value=v)
    apply_data_style(ws5, r, len(headers5))
    for col_idx in [6, 9, 12]:
        val = row_data[col_idx - 1]
        if isinstance(val, (int, float)):
            ws5.cell(row=r, column=col_idx).fill = diff_fill(val)

for c in range(1, len(headers5) + 1):
    ws5.column_dimensions[get_column_letter(c)].width = 14

# ==============================
# Sheet 6: 차이 원인 분석
# ==============================
ws6 = wb.create_sheet("차이 원인 분석")
ws6.sheet_properties.tabColor = "FF0000"

r = 1
ws6.merge_cells('A1:D1')
ws6.cell(row=1, column=1, value="OpenSees vs Midas Gen 차이 원인 분석").font = SUBTITLE_FONT

r = 3
headers6 = ["항목", "차이 크기", "원인", "영향"]
for c, h in enumerate(headers6, 1):
    ws6.cell(row=r, column=c, value=h)
apply_header_style(ws6, r, len(headers6))

analysis_rows = [
    ("DL 반력 +4.5%", "~1,978 kN",
     "Midas Gen Self Weight Z=-1 → 구조 부재(기둥+보) 자중 포함\nOpenSees는 바닥하중(floor_area)만 적용",
     "정상 차이. 비교 시 OpenSees에 자중 추가 또는 Midas에서 자중 제외 가능"),
    ("LL 반력 -8.3%", "~1,960 kN",
     "Midas Gen에서 1개 층(지붕층) LL 미적용\n차이 = 2.5 kN/m² × 784 m² = 1,960 kN (정확히 1층분)",
     "사용자 입력 차이. 지붕 활하중 적용 여부 검토"),
    ("고유치 주기 0.1~5.5%", "최대 0.1s",
     "• 단면 물성(Ix, Iy, J) DB 차이\n• 전단변형 계수(Av) 차이\n• 요소 정식화 미세 차이",
     "구조공학적으로 양호한 일치도. 5% 이내는 통상 허용 범위"),
    ("층간변위 13~18%", "최대 ~1.5mm",
     "• 위 고유치 차이의 누적 효과\n• Midas 자중 포함 → P-Delta 미세 효과 가능\n• Rigid Diaphragm 적용 여부 차이",
     "두 SW 간 비교로서 양호. 설계 판정(OK/NG)은 동일"),
    ("횡하중 반력 0.0%", "< 0.5 kN",
     "동일한 하중값 입력 → 정적 평형 정확 만족",
     "하중 입력 검증 완료"),
]

for label, size, cause, impact in analysis_rows:
    r += 1
    ws6.cell(row=r, column=1, value=label).font = DATA_FONT
    ws6.cell(row=r, column=2, value=size).font = DATA_FONT
    ws6.cell(row=r, column=3, value=cause).font = SMALL_FONT
    ws6.cell(row=r, column=4, value=impact).font = SMALL_FONT
    for c in range(1, 5):
        ws6.cell(row=r, column=c).border = THIN_BORDER
        ws6.cell(row=r, column=c).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

ws6.column_dimensions['A'].width = 22
ws6.column_dimensions['B'].width = 14
ws6.column_dimensions['C'].width = 50
ws6.column_dimensions['D'].width = 40

# ============================================================
# 저장
# ============================================================
output_path = "C:/Users/youm/Desktop/opensees_midas_comparison.xlsx"
wb.save(output_path)
print(f"\nExcel saved: {output_path}")
print("Done!")

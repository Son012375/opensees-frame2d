"""Table A1 (KDS load hand-check)을 편집 가능한 xlsx로 내보낸다.

섹션별 시트(A_DeadLoad / B_LiveLoad / C_Seismic / D_Wind / E_Combinations)와
Overview 시트로 구성. 각 행에 상태 색상, 동결 헤더, 열 너비, 검증 컬럼을 적용한다.

Usage:
    python scripts/_export_table_a1_to_excel.py
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "docs" / "paper1_open_source_alternative" / "validation" / "table_a1_kds_handcheck.xlsx"


# ============================================================
# Styling
# ============================================================

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="305496")
SECTION_FONT = Font(bold=True, size=12, color="1F3864")
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
NOTE_FONT = Font(italic=True, color="595959", size=10)
THIN = Side(border_style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_FILLS = {
    "OK":     PatternFill("solid", fgColor="C6EFCE"),  # green
    "ASSUME": PatternFill("solid", fgColor="FFEB9C"),  # yellow
    "VERIFY": PatternFill("solid", fgColor="BDD7EE"),  # blue
    "FAIL":   PatternFill("solid", fgColor="FFC7CE"),  # red
}

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ============================================================
# Common columns
# ============================================================

COLS = ["#", "Quantity", "KDS reference", "Formula / Input",
        "Hand calc", "Auto-gen", "Δ (%)", "Status", "Verify_status", "Notes"]
COL_WIDTHS = [5, 30, 32, 38, 18, 16, 9, 9, 14, 50]


def _write_header(ws: Worksheet, row: int = 1) -> int:
    for ci, (col, w) in enumerate(zip(COLS, COL_WIDTHS), start=1):
        c = ws.cell(row=row, column=ci, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    return row + 1


def _section_label(ws: Worksheet, row: int, text: str) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLS))
    c = ws.cell(row=row, column=1, value=text)
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    c.alignment = LEFT
    return row + 1


def _data_row(ws: Worksheet, row: int, values: list, status_key: str | None = None) -> int:
    for ci, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=ci, value=v)
        c.alignment = LEFT if ci in (2, 3, 4, 10) else CENTER
        c.border = CELL_BORDER
        if status_key and ci == 8:
            c.fill = STATUS_FILLS.get(status_key, PatternFill())
    return row + 1


def _note_row(ws: Worksheet, row: int, text: str) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLS))
    c = ws.cell(row=row, column=1, value=text)
    c.font = NOTE_FONT
    c.alignment = LEFT
    ws.row_dimensions[row].height = 30
    return row + 1


# ============================================================
# Data (mirrors the markdown)
# ============================================================

# Each row: (id, quantity, kds_ref, formula, hand_calc, auto_gen, delta_pct, status_text, status_key, notes)

DEAD_LOAD = [
    ("A1", "RC unit weight γ_c", "KDS 24 12 21 표 4.2-1 (loaded as `dead_load · unit_weight · steel_reinforced_concrete`)",
     "reference value", "24.0 kN/m³", "24.0", "0.0%", "✓", "OK", ""),
    ("A2", "Slab self-weight (per story)", "KDS 41 12 00 §2.1",
     "γ_c × t_slab = 24.0 × 0.15", "3.60 kN/m²", "3.60", "0.0%", "✓", "OK", ""),
    ("A3", "Finish, 1F (retail)", "User input (workflow assumption)",
     "—", "1.50 kN/m²", "1.50", "0.0%", "✓", "OK", ""),
    ("A4", "Finish, 2F–3F (office)", "User input",
     "—", "1.00 kN/m²", "1.00", "0.0%", "✓", "OK", ""),
    ("A5", "MEP allowance", "Workflow assumption (not in KDS text)",
     "hard-coded 0.5 in load_generator.generate_gravity_loads",
     "0.50 kN/m²", "0.50", "—", "⚠", "ASSUME",
     "Document as a workflow default in §2.4.1; not derived from KDS."),
    ("A6", "DL_total, 1F", "A2 + A3 + A5",
     "3.60 + 1.50 + 0.50", "5.60 kN/m²", "5.60", "0.0%", "✓", "OK", ""),
    ("A7", "DL_total, 2F–3F", "A2 + A4 + A5",
     "3.60 + 1.00 + 0.50", "5.10 kN/m²", "5.10", "0.0%", "✓", "OK", ""),
]

LIVE_LOAD = [
    ("B1", "LL, retail (판매장)", "KDS 41 12 00 표 3.1-1 [VERIFY: row 판매장]",
     "usage = retail", "5.0 kN/m²", "5.00", "0.0%", "✓", "VERIFY",
     "Confirm row index in 표 3.1-1 (e.g., 판매장 1층 vs 상층)."),
    ("B2", "LL, office (사무실)", "KDS 41 12 00 표 3.1-1 [VERIFY: row 사무실]",
     "usage = office", "2.5 kN/m²", "2.50", "0.0%", "✓", "VERIFY",
     "Confirm row index in 표 3.1-1."),
]

SEISMIC = [
    # C.1 Site / hazard
    ("C1", "Seismic zone factor z", "KDS 41 17 00 §3.1 (Seoul, Zone 1)",
     "region = 서울", "0.11 g", "0.11", "0.0%", "✓", "OK", ""),
    ("C2", "Site coeff Fa", "KDS 17 10 00 표 4.2-8, S3 row, linear interp at z=0.11",
     "S3 between S≤0.1 (1.7) and S=0.2 (1.5)",
     "1.7 + (1.5−1.7)·(0.11−0.10)/(0.20−0.10) = 1.68", "1.68", "0.0%", "✓", "OK",
     "Code previously bucketed (returned 1.50); fixed to linear interp per KDS §4.2.1 ②."),
    ("C3", "Site coeff Fv", "KDS 17 10 00 표 4.2-8, S3 row, linear interp at z=0.11",
     "S3 between S≤0.1 (1.7) and S=0.2 (1.6)",
     "1.7 + (1.6−1.7)·(0.11−0.10)/(0.20−0.10) = 1.69", "1.69", "0.0%", "✓", "OK",
     "Code previously bucketed (returned 1.60); fixed to linear interp per KDS §4.2.1 ②."),
    ("C4", "Importance factor IE", "KDS 41 17 00 §3.1.4",
     "Importance Class II", "1.0", "1.0", "0.0%", "✓", "OK", ""),
    # C.2 Spectrum
    ("C5", "S (단주기 유효지반가속도)", "KDS 17 10 00 §4.2.1",
     "z × Fa = 0.11 × 1.68", "0.1848 g", "0.1848", "0.0%", "✓", "OK", ""),
    ("C6", "S₁ (1초주기 유효지반가속도)", "KDS 17 10 00 §4.2.1",
     "z × Fv = 0.11 × 1.69", "0.1859 g", "0.1859", "0.0%", "✓", "OK", ""),
    ("C7", "SDS", "KDS 17 10 00 §4.2.1",
     "2.5 × S = 2.5 × 0.1848", "0.4620 g", "0.4620", "0.0%", "✓", "OK", ""),
    ("C8", "SD1", "KDS 17 10 00 §4.2.1",
     "S₁", "0.1859 g", "0.1859", "0.0%", "✓", "OK", ""),
    # C.3 SFRS
    ("C9", "Response modification R", "KDS 41 17 00 표 6.3-1 [VERIFY: 보통철골모멘트골조]",
     "seismic_system = ordinary_moment_frame", "3.5", "3.5", "0.0%", "✓", "VERIFY",
     "Verify row '3-c' in 표 6.3-1."),
    ("C10", "Overstrength Ω₀", "KDS 41 17 00 표 6.3-1",
     "same row as C9", "3.0", "3.0", "0.0%", "✓", "VERIFY", ""),
    ("C11", "Deflection amplification Cd", "KDS 41 17 00 표 6.3-1",
     "same row as C9", "3.0", "3.0", "0.0%", "✓", "VERIFY", ""),
    # C.4 Period
    ("C12", "Ct (period coeff)", "KDS 41 17 00 §7.2.4 (1), 식 (7.2-6) 본문 내 파라미터 리스트 (철골모멘트골조)",
     "steel MF", "0.0724", "0.0724", "0.0%", "✓", "OK",
     "Verified from KDS PDF: §7.2.4 (1) lists Ct=0.0724 inline. No separate numbered table (no 표 6.4-1)."),
    ("C13", "x (period exponent)", "KDS 41 17 00 §7.2.4 (1), 식 (7.2-6) 본문 내 파라미터 리스트",
     "steel MF", "0.80", "0.80", "0.0%", "✓", "OK",
     "Verified from KDS PDF: §7.2.4 (1) lists x=0.8 inline alongside Ct."),
    ("C14", "Building height hn", "Geometry",
     "Σ story heights", "9.0 m", "9.0", "0.0%", "✓", "OK", ""),
    ("C15", "Approx. period Ta", "KDS 41 17 00 §7.2.4 (1), 식 (7.2-6)",
     "Ct × hn^x = 0.0724 × 9^0.80", "0.4199 s", "0.4199", "0.0%", "✓", "OK", ""),
    ("C16", "Period upper-limit Cu", "KDS 41 17 00 §7.2.3 (1), 표 7.2-1, linear interp at SD1=0.1859",
     "between 0.15 (Cu=1.6) and 0.20 (Cu=1.5)",
     "1.6 + (1.5−1.6)·(0.1859−0.15)/(0.20−0.15) = 1.5282",
     "1.5282", "0.0%", "✓", "OK",
     "Code previously bucketed SD1; fixed to linear interp per KDS table caveat '직선보간한다'. Doesn't propagate (T = Ta directly, Cu·Ta not binding)."),
    ("C17", "Design period T", "min(Ta, Cu·Ta); code uses Ta directly (no modal analysis)",
     "T = Ta (Ta=0.4199 < Cu·Ta=0.641 either way)", "0.4199 s", "0.4199", "0.0%", "✓", "OK", ""),
    # C.5 Cs
    ("C18", "Cs (식 6.5-1, governing candidate)", "KDS 41 17 00 §6.5 (식 6.5-1)",
     "SDS · IE / R = 0.4620 × 1.0 / 3.5", "0.1320", "0.1320", "0.0%", "✓", "OK", ""),
    ("C19", "Cs_min", "KDS 41 17 00 §6.5 (식 6.5-3)",
     "max(0.044·SDS·IE, 0.01) = max(0.02033, 0.01)", "0.02033", "0.02033", "0.0%", "✓", "OK", ""),
    ("C20", "Cs_max", "KDS 41 17 00 §6.5 (식 6.5-2)",
     "SD1·IE / (T·R) = 0.1859 / (0.4199×3.5)", "0.1265", "0.1265", "0.0%", "✓", "OK", ""),
    ("C21", "Cs (final)", "C18 bounded by C19, C20",
     "C18 (0.1320) > C20 (0.1265) → Cs_max binds", "0.1265", "0.1265", "0.0%", "✓", "OK",
     "Regime change vs. earlier draft: after Fa/Fv interp fix, Cs_max (식 6.5-2) now governs instead of 식 6.5-1."),
    # C.6 Base shear & distribution
    ("C22", "W₁ (1F effective weight)", "KDS 41 17 00 §6.6",
     "DL_1F × A = 5.60 × 120", "672.0 kN", "672.0", "0.0%", "✓", "OK", ""),
    ("C23", "W₂ (2F effective weight)", "§6.6",
     "5.10 × 120", "612.0 kN", "612.0", "0.0%", "✓", "OK", ""),
    ("C24", "W₃ (3F effective weight)", "§6.6",
     "5.10 × 120", "612.0 kN", "612.0", "0.0%", "✓", "OK", ""),
    ("C25", "W (total)", "Σ Wᵢ",
     "672 + 612 + 612", "1,896.0 kN", "1,896.0", "0.0%", "✓", "OK", ""),
    ("C26", "V (base shear)", "KDS 41 17 00 §6.6",
     "Cs · W = 0.1265 × 1896", "239.84 kN", "239.84", "0.0%", "✓", "OK",
     "Paper §4.1 history: 213.0 kN (original draft, earlier code) → 223.5 kN (W/Cs audit) → 239.8 kN (after Fa/Fv interp fix). Final value in v2 draft."),
    ("C27", "Distribution exponent k", "KDS 41 17 00 §6.7",
     "T ≤ 0.5 → k = 1.0", "1.0", "1.0", "0.0%", "✓", "OK", ""),
]

# Story-wise seismic distribution (Cvx unchanged; Fx scales with V_base)
SEISMIC_STORIES = [
    # (story, wi, hi, wi*hi, Cvx_hand, Cvx_auto, Fx_hand, Fx_auto, delta, status_text, status_key)
    ("Story 1", 672.0, 3.0, 2016.0, 0.1801, 0.1801, 43.19, 43.19, "0.0%", "✓", "OK"),
    ("Story 2", 612.0, 6.0, 3672.0, 0.3280, 0.3280, 78.66, 78.66, "0.0%", "✓", "OK"),
    ("Story 3", 612.0, 9.0, 5508.0, 0.4920, 0.4920, 117.99, 117.99, "0.0%", "✓", "OK"),
    ("Σ",       1896.0, "—", 11196.0, 1.0000, 1.0000, 239.84, 239.84, "0.0%", "✓", "OK"),
]

WIND = [
    ("D1", "Basic wind speed V₀", "Hazard DB `hazard_region_values.wind_v0` (Seoul) [VERIFY]",
     "region = 서울", "30 m/s", "30", "0.0%", "✓", "VERIFY",
     "Confirm DB source maps to KDS-listed value for Seoul."),
    ("D2", "Topographic factor Kzt", "§5.2.5 (assumed flat)",
     "flat terrain", "1.0", "1.0", "0.0%", "⚠", "ASSUME", "Assumed flat terrain."),
    ("D3", "Air density ρ", "physical constant",
     "—", "1.225 kg/m³", "1.225", "0.0%", "✓", "OK", ""),
    ("D4", "Exposure category", "User input",
     "B (suburban)", "B", "B", "—", "✓", "OK", ""),
    ("D5", "Gradient height zg (B)", "KDS 41 12 00 표 5.2-4 [VERIFY]",
     "exposure B", "365 m", "365", "0.0%", "✓", "VERIFY", ""),
    ("D6", "Exponent α (B)", "KDS 41 12 00 표 5.2-4 [VERIFY]",
     "exposure B", "9.5", "9.5", "0.0%", "✓", "VERIFY", ""),
    ("D7", "Gust factor Gf (B)", "Workflow simplification (vs KDS §5 detailed formula)",
     "hard-coded map", "2.20", "2.20", "—", "⚠", "ASSUME",
     "Disclose as simplification in §2.4.1; full KDS Gf depends on building dim/period/damping."),
    ("D8", "Cp_total (windward+leeward)", "Workflow simplification",
     "0.8 + 0.5", "1.30", "1.30", "—", "⚠", "ASSUME",
     "Disclose as simplification in §2.4.1; KDS uses aspect-ratio Cp tables."),
]

# Story-wise wind
WIND_COLS = ["Story", "z_mid (m)", "z_eff (m)", "Kz hand", "Kz auto",
             "qz hand (kN/m²)", "qz auto", "p hand", "p auto",
             "Fx hand (kN)", "Fx auto", "Fy hand", "Fy auto", "Δ", "Status"]
WIND_COL_WIDTHS = [8, 10, 10, 10, 10, 14, 12, 10, 10, 12, 12, 12, 12, 9, 9]

WIND_STORIES = [
    ("Story 1", 1.5, 5.0, 0.8145, 0.8145, 0.3657, 0.3657, 1.0460, 1.0460, 31.38, 31.38, 37.66, 37.66, "0.0%", "✓"),
    ("Story 2", 4.5, 5.0, 0.8145, 0.8145, 0.3657, 0.3657, 1.0460, 1.0460, 31.38, 31.38, 37.66, 37.66, "0.0%", "✓"),
    ("Story 3", 7.5, 7.5, 0.8871, 0.8871, 0.4338, 0.4338, 1.2408, 1.2408, 37.22, 37.22, 44.67, 44.67, "0.0%", "✓"),
    ("Σ",       "—", "—", "—", "—", "—", "—", "—", "—", 99.98, 99.98, 119.99, 119.99, "0.0%", "✓"),
]

# Combinations
COMBO_COLS = ["Group", "Count", "Combinations", "KDS 41 12 00 §1.7 check", "Status", "Verify_status", "Notes"]
COMBO_WIDTHS = [16, 8, 60, 32, 9, 14, 40]

COMBINATIONS = [
    ("Gravity only", 2, "1.4 DL; 1.2 DL + 1.6 LL", "matches 식 1 & 2", "✓", "OK", ""),
    ("Seismic X (±)", 4, "1.2 DL + 1.0 LL ± 1.0 EQX; 0.9 DL ± 1.0 EQX",
     "matches 식 5 & 7", "✓", "OK", ""),
    ("Seismic Y (±)", 4, "1.2 DL + 1.0 LL ± 1.0 EQY; 0.9 DL ± 1.0 EQY",
     "matches 식 5 & 7", "✓", "OK", ""),
    ("Wind X (±)", 4, "1.2 DL + 1.0 LL ± 1.0 WX; 0.9 DL ± 1.0 WX",
     "matches 식 4 & 6", "✓", "OK", ""),
    ("Wind Y (±)", 4, "1.2 DL + 1.0 LL ± 1.0 WY; 0.9 DL ± 1.0 WY",
     "matches 식 4 & 6", "✓", "OK", ""),
    ("Total", 18, "—", "matches paper §4.1", "✓", "OK", ""),
]


# ============================================================
# Sheet builders
# ============================================================

def build_overview(wb: Workbook) -> None:
    ws = wb.create_sheet("Overview", 0)

    # Title
    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value="Table A1. KDS Load Generation Hand-Check (Paper 1, Appendix A)")
    c.font = Font(bold=True, size=14, color="1F3864")
    c.alignment = LEFT
    ws.row_dimensions[1].height = 28

    # Building input block
    rows_in = [
        ("Target building", "Three-story regular steel frame (Paper §4.1)"),
        ("Geometry", "3 stories × 3.0 m, 3 × 2 bays (4.0 m × 5.0 m), 12.0 × 10.0 × 9.0 m"),
        ("Occupancy", "1F retail (finish 1.5) / 2F–3F office (finish 1.0)"),
        ("Site", "Seoul (z=0.11g), Site Class S3, V₀=30 m/s, Exposure B"),
        ("Importance / system", "Class II (IE=1.0), Ordinary Steel Moment Frame"),
        ("Sections / material", "Columns H-300×300, Beams H-400×200, SS275"),
        ("Slab", "0.15 m RC (γ_c = 24.0 kN/m³)"),
        ("Input file", "docs/paper1_open_source_alternative/validation/example_input.json"),
        ("Auto-gen source", "example_auto_values.json (scripts/_hand_check_loads.py)"),
        ("Codes referenced", "KDS 41 12 00, 41 17 00, 17 10 00, 24 12 21"),
    ]
    r = 3
    for k, v in rows_in:
        kc = ws.cell(row=r, column=1, value=k); kc.font = Font(bold=True); kc.alignment = LEFT
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        vc = ws.cell(row=r, column=2, value=v); vc.alignment = LEFT
        r += 1

    r += 1

    # Status legend
    legend = [
        ("✓",    "OK — agree within rounding (Δ ≤ 0.1%)",         "OK"),
        ("⚠",    "Assumption / simplification — disclose in §2.4.1", "ASSUME"),
        ("[VERIFY]", "Requires direct KDS table lookup",            "VERIFY"),
        ("❌",   "Material discrepancy (Δ > 1%) — requires fix",   "FAIL"),
    ]
    ws.cell(row=r, column=1, value="Status legend").font = Font(bold=True, size=12)
    r += 1
    for sym, desc, key in legend:
        sc = ws.cell(row=r, column=1, value=sym); sc.alignment = CENTER
        sc.fill = STATUS_FILLS.get(key, PatternFill())
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=6)
        ws.cell(row=r, column=2, value=desc).alignment = LEFT
        r += 1

    r += 1

    # Summary table
    ws.cell(row=r, column=1, value="Summary by section").font = Font(bold=True, size=12)
    r += 1
    summary_cols = ["Section", "Rows", "✓ OK", "⚠ Assume", "[VERIFY]", "❌ Fail"]
    for ci, h in enumerate(summary_cols, start=1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = CELL_BORDER
    r += 1

    # Counts (mirror md summary)
    summary_rows = [
        ("A. Dead load", 7, 6, 1, 0, 0),
        ("B. Live load", 2, 0, 0, 2, 0),
        ("C. Seismic", 27, 24, 0, 3, 0),  # +C12/C13/C16 after Cu fix + §7.2 PDF verify
        ("D. Wind", 8, 2, 3, 3, 0),
        ("E. Combinations", 6, 6, 0, 0, 0),
        ("Total", 50, 38, 4, 8, 0),
    ]
    for tup in summary_rows:
        for ci, v in enumerate(tup, start=1):
            c = ws.cell(row=r, column=ci, value=v)
            c.alignment = CENTER if ci > 1 else LEFT
            c.border = CELL_BORDER
            if tup[0] == "Total":
                c.font = Font(bold=True)
        r += 1

    r += 1

    # Action items
    ws.cell(row=r, column=1, value="Action items before submission").font = Font(bold=True, size=12)
    r += 1
    actions = [
        "1. Paper §4 text correction: V_base 213 → 239.8 kN, Cs 0.118 → 0.126, displacements/drifts updated in v2 docx. Member ratios pending update after Vu fix (item 4).",
        "2. Code fix (DONE): Fa/Fv now linearly interpolated per KDS 17 10 00 §4.2.1 ② (was bucketed). See design_spectrum.py:_linear_interp_z + _get_site_coefficients.",
        "3. Code fix (DONE): Cu now linearly interpolated per KDS 41 17 00 §7.2.3 (1) 표 7.2-1 (was bucketed). See load_generator.py:_get_period_upper_limit. No §4 cascade.",
        "4. Code fix (DONE, separate session): design_check.py Vu now uses ops.eleResponse('localForce') instead of ops.eleForce (global). Max member interaction now 0.507 (column, Story 2, EQY) — column-governed.",
        "5. Verify remaining 8 [VERIFY] rows against KDS PDFs (LL, R/Ω₀/Cd, wind exposure parameters). Companion .xlsx is the working artifact.",
        "6. Document the four workflow simplifications (MEP=0.5; Kzt=1.0 flat; Gf constants; Cp_total=1.30) in §2.4.1.",
        "7. Add this appendix as A.2 KDS load generation hand-check; remove future-work line from §5.3.",
    ]
    for a in actions:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        c = ws.cell(row=r, column=1, value=a); c.alignment = LEFT
        ws.row_dimensions[r].height = 24
        r += 1

    # Column widths
    for ci, w in enumerate([20, 18, 14, 14, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "A3"


def build_section_sheet(wb: Workbook, name: str, rows: list, note_lines: list[str] | None = None) -> None:
    ws = wb.create_sheet(name)
    r = _write_header(ws, row=1)
    for row in rows:
        # row tuple: (id, qty, kds, formula, hand, auto, delta, status_text, status_key, notes)
        values = [row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], "", row[9]]
        r = _data_row(ws, r, values, status_key=row[8])
    if note_lines:
        r += 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
        c = ws.cell(row=r, column=1, value="Notes"); c.font = Font(bold=True, size=11)
        r += 1
        for line in note_lines:
            r = _note_row(ws, r, line)
    ws.freeze_panes = "A2"


def build_seismic_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("C_Seismic")
    r = _write_header(ws, row=1)

    # group rows by sub-section
    groups = [
        ("C.1 Site / hazard parameters", SEISMIC[0:4]),
        ("C.2 Design spectrum parameters", SEISMIC[4:8]),
        ("C.3 Seismic-force-resisting system", SEISMIC[8:11]),
        ("C.4 Fundamental period", SEISMIC[11:17]),
        ("C.5 Seismic response coefficient Cs", SEISMIC[17:21]),
        ("C.6 Base shear and distribution", SEISMIC[21:27]),
    ]
    for label, items in groups:
        r = _section_label(ws, r, label)
        for row in items:
            values = [row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], "", row[9]]
            r = _data_row(ws, r, values, status_key=row[8])

    # Story distribution sub-table
    r += 1
    r = _section_label(ws, r, "C.6 Story-wise force distribution  (Cvx = wᵢ·hᵢᵏ / Σ wⱼ·hⱼᵏ, Fx = Cvx·V)")
    sf_cols = ["Story", "wᵢ (kN)", "hᵢ (m)", "wᵢ·hᵢ", "Cvx hand", "Cvx auto", "Fx hand (kN)", "Fx auto (kN)", "Δ", "Status"]
    for ci, h in enumerate(sf_cols, start=1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = CELL_BORDER
    r += 1
    for srow in SEISMIC_STORIES:
        for ci, v in enumerate(srow[:-1], start=1):  # last is status_key for fill
            c = ws.cell(row=r, column=ci, value=v); c.alignment = CENTER; c.border = CELL_BORDER
        ws.cell(row=r, column=9, value=srow[8]).alignment = CENTER
        sc = ws.cell(row=r, column=10, value=srow[9]); sc.alignment = CENTER
        sc.fill = STATUS_FILLS.get(srow[10], PatternFill())
        sc.border = CELL_BORDER
        r += 1

    # Notes
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
    c = ws.cell(row=r, column=1, value="Notes on C (Seismic)"); c.font = Font(bold=True, size=11)
    r += 1
    notes = [
        "C-a. Fa/Fv linear interpolation (RESOLVED): hand-check identified that design_spectrum.py was bucketing z into KDS table anchor rows (z=0.1/0.2/0.3) instead of linearly interpolating per KDS 17 10 00 §4.2.1 ② (\"S의 값이 중간 값에 해당할 경우 직선보간하여 결정한다\"). Code was patched: Fa 1.50→1.68, Fv 1.60→1.69 at Seoul z=0.11. Cascade: SDS 0.4125→0.4620, SD1 0.176→0.1859, Cs regime changed from 식 6.5-1 to Cs_max binding, V_base 223.5→239.8 kN.",
        "C-b. Cu linear interpolation (RESOLVED, same pattern as C-a): KDS 41 17 00 §7.2.3 (1) 표 7.2-1 explicitly states 'SD1의 중간값에 해당할 경우 주기상한계수 Cu는 직선보간한다'. load_generator.py:_get_period_upper_limit was bucketing SD1; patched to use linear interp via _linear_interp_z (imported from design_spectrum.py). Seoul (SD1=0.1859): Cu 1.50→1.5282. No cascade because T = Ta directly in current implementation (Ta < Cu·Ta always).",
        "C-c. KDS clause numbering (RESOLVED): code/spec previously referenced §6.4 표 6.4-1/표 6.4-2 (period coefficients). KDS 41 17 00 current edition uses §7.2.4 (1) inline parameter list for Ct/x and 표 7.2-1 (under §7.2.3) for Cu. References updated accordingly throughout this appendix.",
        "C-d. Paper §4.1 V_base history: 213.0 kN (original draft, pre-audit) → 223.5 kN (V_base+W audit) → 239.8 kN (after Fa/Fv interp fix, v2 draft). All DL/LL/Ta inputs unchanged across revisions.",
    ]
    for line in notes:
        r = _note_row(ws, r, line)

    ws.freeze_panes = "A2"


def build_wind_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("D_Wind")
    r = _write_header(ws, row=1)
    for row in WIND:
        values = [row[0], row[1], row[2], row[3], row[4], row[5], row[6], row[7], "", row[9]]
        r = _data_row(ws, r, values, status_key=row[8])

    # Story-wise sub-table
    r += 1
    r = _section_label(ws, r, "D.2 Story-wise wind force  (Kz = 2.01·(z_eff/zg)^(2/α); qz = 0.5·ρ·(V₀·Kz·Kzt)²·10⁻³; p = qz·Gf·Cp; Fx = p·h·By; Fy = p·h·Bx)")
    for ci, h in enumerate(WIND_COLS, start=1):
        c = ws.cell(row=r, column=ci, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = WIND_COL_WIDTHS[ci - 1]
    r += 1
    for srow in WIND_STORIES:
        for ci, v in enumerate(srow, start=1):
            c = ws.cell(row=r, column=ci, value=v); c.alignment = CENTER; c.border = CELL_BORDER
            if ci == 15:  # Status column
                c.fill = STATUS_FILLS.get("OK", PatternFill())
        r += 1

    # Notes
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COLS))
    c = ws.cell(row=r, column=1, value="Notes on D (Wind)"); c.font = Font(bold=True, size=11)
    r += 1
    notes = [
        "D-a. Workflow simplifications: Gf (D7) and Cp_total (D8) are constants in generate_wind_loads. KDS 41 12 00 §5 specifies more detailed expressions. Disclose explicitly in §2.4.1 as simplifications.",
        "D-b. z_min = 5 m floor: code clamps z_eff to ≥ 5 m, causing Stories 1 and 2 to share the same Kz. Acceptable simplification; should be noted.",
    ]
    for line in notes:
        r = _note_row(ws, r, line)

    ws.freeze_panes = "A2"


def build_combo_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("E_Combinations")
    for ci, (h, w) in enumerate(zip(COMBO_COLS, COMBO_WIDTHS), start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = CENTER; c.border = CELL_BORDER
        ws.column_dimensions[get_column_letter(ci)].width = w
    r = 2
    for row in COMBINATIONS:
        for ci, v in enumerate(row, start=1):
            c = ws.cell(row=r, column=ci, value=v)
            c.alignment = LEFT if ci in (3, 4, 7) else CENTER
            c.border = CELL_BORDER
            if ci == 5:  # Status text
                c.fill = STATUS_FILLS.get(row[5], PatternFill())
        if row[0] == "Total":
            for ci in range(1, len(row) + 1):
                ws.cell(row=r, column=ci).font = Font(bold=True)
        r += 1

    # Notes
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COMBO_COLS))
    c = ws.cell(row=r, column=1, value="Notes on E (Combinations)"); c.font = Font(bold=True, size=11)
    r += 1
    notes = [
        "KDS 41 12 00 §1.7 식 3 (Lr/S/R) is not generated because roof live, snow, rain loads are not produced for this example.",
        "Snow combinations are omitted because the snow load generator was not invoked for Seoul in this run.",
    ]
    for line in notes:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(COMBO_COLS))
        cc = ws.cell(row=r, column=1, value=line); cc.font = NOTE_FONT; cc.alignment = LEFT
        ws.row_dimensions[r].height = 30
        r += 1

    ws.freeze_panes = "A2"


# ============================================================
# Main
# ============================================================

def main() -> int:
    wb = Workbook()
    # remove default sheet
    wb.remove(wb.active)

    build_overview(wb)
    build_section_sheet(wb, "A_DeadLoad", DEAD_LOAD, note_lines=[
        "A5 (MEP = 0.5): workflow constant, not a KDS clause. Document as assumption in §2.4.1.",
    ])
    build_section_sheet(wb, "B_LiveLoad", LIVE_LOAD, note_lines=[
        "Confirm exact row indices in KDS 41 12 00 표 3.1-1 (판매장 1층 vs 상층, 사무실 vs 회의실).",
    ])
    build_seismic_sheet(wb)
    build_wind_sheet(wb)
    build_combo_sheet(wb)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUT))
    print(f"[OK] Excel workbook saved: {OUT.relative_to(ROOT)}")
    print(f"  Sheets: Overview, A_DeadLoad, B_LiveLoad, C_Seismic, D_Wind, E_Combinations")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

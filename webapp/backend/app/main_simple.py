"""
OpenSees Structural Analysis Platform
Simplified FastAPI app - No Redis/Celery required
Runs analysis synchronously (blocking)
"""
import logging
import os
import sys
import json
import threading
import time
import uuid
from pathlib import Path
from datetime import datetime

# Load .env from project root (for ANTHROPIC_API_KEY etc.)
from dotenv import load_dotenv
_project_root = Path(__file__).resolve().parents[3]  # app/main_simple.py -> opensees-MCP
load_dotenv(_project_root / ".env")

# Logger — replaces ad-hoc print()/traceback.print_exc() so warnings surface
# in Render logs (and locally in `uvicorn` output) instead of being silently
# swallowed.
logger = logging.getLogger("opensees.app")
if not logger.handlers:
    _handler = logging.StreamHandler()
    _handler.setFormatter(logging.Formatter("%(asctime)s %(levelname)s [%(name)s] %(message)s"))
    logger.addHandler(_handler)
logger.setLevel(logging.INFO)

from fastapi import FastAPI, Request, HTTPException, UploadFile, File
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.responses import (
    HTMLResponse, FileResponse, StreamingResponse, RedirectResponse,
)

from pydantic import BaseModel, Field
from typing import Optional, List, Dict, Any

from app.models.schemas import JobStatus
from app.core.config import MCP_SERVER_PATH, JOBS_DIR
from app.core.claude_service import parse_building, check_api_key


def _build_recommendation_block(
    *,
    design_check: Optional[dict] = None,
    raw_warnings: Optional[List[Any]] = None,
    analysis_metadata: Optional[Dict[str, Any]] = None,
    stage: str = "analysis",
    mode: str = "analyze",
) -> Dict[str, Any]:
    """Thin wrapper around ``core.recommendation.build_recommendation_payload``.

    Kept here so the recommendation foundation can be wired into multiple
    endpoints without each one re-importing / re-handling errors. On any
    failure we degrade gracefully — analysis itself has already succeeded
    by the time we get here.

    ``analysis_metadata`` is forwarded as-is to the recommendation layer
    where it enriches issues with member/section/story/coords context.
    ``mode="parse_only"`` suppresses ``missing_design_check`` issues for
    pre-analysis stages (e.g. /api/v2/parse-ifc).
    """
    if str(MCP_SERVER_PATH) not in sys.path:
        sys.path.insert(0, str(MCP_SERVER_PATH))
    try:
        from core.recommendation import build_recommendation_payload
        return build_recommendation_payload(
            design_check=design_check,
            raw_warnings=raw_warnings,
            analysis_metadata=analysis_metadata,
            stage=stage,
            mode=mode,
        )
    except Exception as rec_err:  # noqa: BLE001 — defensive: analysis was OK
        logger.warning("Recommendation pipeline failed (%s): %s", stage, rec_err, exc_info=True)
        # Fall back to legacy-style mirror so the response stays usable.
        legacy = [str(w) for w in (raw_warnings or [])]
        return {
            "warnings": [{
                "code": "recommendation_pipeline_failed",
                "message": str(rec_err),
                "severity": "warning",
                "stage": stage,
                "recoverable": True,
            }],
            "warning_messages": legacy,
            "issues": [],
            "recommendation_candidates": [],
            "recommendation_summary": {
                "num_issues": 0,
                "num_candidates": 0,
                "rag_enabled": False,
                "llm_enabled": False,
                "mode": mode,
            },
        }


class NaturalLanguageInput(BaseModel):
    """Natural language input for Claude parsing.

    The length cap is a billing control, not a usability one: this text is sent
    straight to a paid LLM from an unauthenticated endpoint. A building
    description never approaches 2000 characters.
    """
    text: str = Field(max_length=2000)


class ParseResponse(BaseModel):
    """Response from natural language parsing"""
    success: bool
    data: Optional[dict] = None
    error: Optional[str] = None


class BuildingInput(BaseModel):
    """Building analysis input (same config as MCP analyze_building)"""
    config: Dict[str, Any]


# ── Public-demo resource guard ───────────────────────────────────────────────
# Measured peak RSS: 3-story preset ~0.7 GB, 10-story preset ~6 GB. A container
# sized for the demo is OOM-killed by one oversized request, which takes down
# every other visitor's session with it. Reject early with a readable message
# instead. DEMO_MAX_MEMBERS=0 disables the guard (local/desktop use).
DEMO_MAX_MEMBERS = int(os.environ.get("DEMO_MAX_MEMBERS", "0") or 0)

#: Enforced while streaming the upload, not after it is already in memory.
IFC_MAX_UPLOAD_MB = int(os.environ.get("IFC_MAX_UPLOAD_MB", "20") or 20)

# ── Rate limit for the paid-LLM endpoint ─────────────────────────────────────
# /api/claude/parse-building is anonymous, uncached and billed per call. Without
# a ceiling one script can spend the author's credits without limit. Per-IP and
# global daily caps; 0 disables (local use).
LLM_CALLS_PER_IP_PER_HOUR = int(os.environ.get("LLM_CALLS_PER_IP_PER_HOUR", "0") or 0)
LLM_CALLS_PER_DAY = int(os.environ.get("LLM_CALLS_PER_DAY", "0") or 0)

_llm_hits: dict[str, list[float]] = {}
_llm_day: list[float] = []
_llm_lock = threading.Lock()


def _enforce_llm_quota(request: Request) -> None:
    if not (LLM_CALLS_PER_IP_PER_HOUR or LLM_CALLS_PER_DAY):
        return
    now = time.time()
    ip = (request.headers.get("x-forwarded-for", "").split(",")[0].strip()
          or (request.client.host if request.client else "unknown"))
    with _llm_lock:
        if LLM_CALLS_PER_DAY:
            _llm_day[:] = [t for t in _llm_day if now - t < 86400]
            if len(_llm_day) >= LLM_CALLS_PER_DAY:
                raise HTTPException(
                    status_code=429,
                    detail=("오늘 자연어 변환 사용량이 소진되었습니다. "
                            "직접 입력 또는 예제 IFC로 계속 진행하실 수 있습니다."))
        if LLM_CALLS_PER_IP_PER_HOUR:
            hits = [t for t in _llm_hits.get(ip, []) if now - t < 3600]
            if len(hits) >= LLM_CALLS_PER_IP_PER_HOUR:
                _llm_hits[ip] = hits
                raise HTTPException(
                    status_code=429,
                    detail=("자연어 변환 요청이 너무 잦습니다. 잠시 후 다시 시도해 주세요. "
                            "직접 입력 또는 예제 IFC는 제한이 없습니다."))
            hits.append(now)
            _llm_hits[ip] = hits
        if LLM_CALLS_PER_DAY:
            _llm_day.append(now)


def _estimate_member_count(config: Dict[str, Any]) -> int:
    """Members a rectangular-grid config will generate, before building it."""
    try:
        n_story = len(config.get("stories") or []) or int(config.get("num_stories") or 0)
        nx = len(config.get("bays_x") or []) or int(config.get("num_bays_x") or 0)
        ny = len(config.get("bays_y") or []) or int(config.get("num_bays_y") or 0)
    except (TypeError, ValueError):
        return 0
    if not (n_story and nx and ny):
        return 0
    gx, gy = nx + 1, ny + 1
    columns = gx * gy * n_story
    beams_x = nx * gy * n_story
    beams_y = ny * gx * n_story
    return columns + beams_x + beams_y


def _enforce_demo_size(member_count: int, what: str = "모델") -> None:
    if not DEMO_MAX_MEMBERS or member_count <= DEMO_MAX_MEMBERS:
        return
    raise HTTPException(
        status_code=413,
        detail=(
            f"{what} 규모가 이 공개 데모의 한도를 넘습니다 "
            f"(부재 {member_count}개 > 한도 {DEMO_MAX_MEMBERS}개). "
            "더 작은 모델로 시도해 주시거나, 저장소를 내려받아 로컬에서 실행해 주세요."
        ),
    )


# Application
app = FastAPI(title="OpenSees Structural Analysis Platform")

# Phase A.1 — chat router (sessions + NDJSON stream). Imported here so
# the include_router call sits right next to app creation; chat_router
# does its own MCP_SERVER_PATH injection so importing it is safe even
# before main_simple's endpoints run.
from app.chat_router import router as chat_router  # noqa: E402
app.include_router(chat_router)

# Demo auth - disabled for now
# from app.core.auth import check_demo_auth, make_auth_response, set_auth_cookie

# Static files and templates
BASE_DIR = Path(__file__).resolve().parent.parent
app.mount("/static", StaticFiles(directory=BASE_DIR / "static"), name="static")
templates = Jinja2Templates(directory=BASE_DIR / "templates")

# The landing page is a self-contained static bundle at the repo root
# (built by scripts/build_landing.py). It is served here so a single-origin
# deployment works out of the box; hosting it separately is also supported —
# see the app-base <meta> in landing/index.template.html.
LANDING_DIR = BASE_DIR.parent.parent / "landing"
if (LANDING_DIR / "assets").is_dir():
    app.mount("/assets", StaticFiles(directory=LANDING_DIR / "assets"),
              name="landing-assets")
if (LANDING_DIR / "files").is_dir():
    app.mount("/files", StaticFiles(directory=LANDING_DIR / "files"),
              name="landing-files")

# In-memory job storage.
#
# LIMITATION: jobs and on-disk artifacts under JOBS_DIR (report.html,
# result.json, v2_viewers/*.html) live only on the worker that handled the
# request. On Render free plan the disk is wiped on cold start / redeploy
# and the worker sleeps after inactivity. There is no Redis / object
# store backing this yet, so any old job_id will 404 after restart and
# the user must re-run the analysis. See README §"Known limitations".
jobs_db: dict[str, dict] = {}

# ── /api/v2/recommendations/evaluate state ──
#
# Two parallel dicts (now owned by ``app.services``):
#   * ``analysis_context_cache[job_id]`` — what /api/v2/analyze cached:
#     model_json, load_cases/combos, building_model, seismic_report,
#     baseline design_check, candidates_by_id, expires_at.
#   * ``eval_jobs_db[eval_job_id]`` — async evaluate job state:
#     status, progress, results, errors.
#
# OpenSees has process-global state, so reanalysis runs are serialized
# through a single-worker thread pool (MAX_PARALLEL_EVALS = 1).
import time

from app.services.analysis_context import (
    analysis_context_cache,
    _ANALYSIS_CONTEXT_LOCK,
    _ANALYSIS_CONTEXT_TTL_SEC,
    _purge_expired_analysis_contexts,
    _get_analysis_context,
    build_compact_subset,
)
from app.services.recommendation_jobs import (
    eval_jobs_db,
    _EVAL_JOBS_LOCK,
    MAX_PARALLEL_EVALS,
    _eval_executor,
    _rehydrate_candidate,
)
from app.services.solver_pool import run_solver, queue_depth


# Marker on each ``analysis_context_cache`` entry that tells the
# recommendation endpoints whether the entry was created with the full
# baseline (model_json + load_cases + candidates_by_id + design_check)
# or only the compact chat-tool subset (member_info_by_*, modal_summary,
# envelope, analysis_summary).
#
#   "full"    — written by /api/v2/analyze. Recommendation pipeline OK.
#   "compact" — written by /api/building/analyze. Chat tools OK but the
#               recommendation endpoints (evaluate / preview-apply /
#               explain) must reject this with a clear 422 because
#               they read ``ctx["model_json"]`` etc. that aren't there.
#
# A future change can promote the building endpoint to full context if
# we decide to support recommendations from the legacy analyze path —
# at that point the field stays valid as the discriminator and the
# 422 just stops firing for those entries.
CONTEXT_KIND_FULL = "full"
CONTEXT_KIND_COMPACT = "compact"


def _assert_full_context_for_recommendations(ctx: dict, analysis_id: str) -> None:
    """Raise 422 if ``ctx`` lacks the fields the recommendation API needs.

    Called at the entry of every /api/v2/recommendations/* endpoint
    *after* the 400/410 split on ``ctx is None``. Surfaces the partial-
    cache state as a discrete, documented failure mode instead of a
    silent ``candidates_by_id={}`` or a downstream ``KeyError`` on
    ``ctx["model_json"]`` (which the apply-candidate path hits).
    """
    if ctx.get("context_kind") == CONTEXT_KIND_FULL:
        return
    raise HTTPException(
        status_code=422,
        detail=(
            f"analysis '{analysis_id}' was created via /api/building/"
            "analyze and only carries the compact chat-tool subset — "
            "recommendation endpoints require a /api/v2/analyze baseline "
            "(model_json + candidates_by_id). Re-run the analysis "
            "through the V2 node-element path to enable recommendations."
        ),
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Page Routes
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    """Landing page.

    The bare domain used to serve the editor, so anyone given the root URL
    landed in a tool with no explanation and no way back to the pitch. The
    editor now lives at /editor-figma, which the landing links into.
    """
    index = LANDING_DIR / "index.html"
    if index.exists():
        return FileResponse(str(index), media_type="text/html")
    # Landing not built yet (scripts/build_landing.py) — fall back to the editor
    # rather than 404ing a local dev session.
    return templates.TemplateResponse(request, "editor_figma.html")


@app.get("/landing", response_class=HTMLResponse)
async def landing(request: Request):
    """Kept so existing links keep working; the landing is now the root."""
    return RedirectResponse(url="/", status_code=307)


# ═══════════════════════════════════════════════════════════════════════════════
# Claude API endpoints
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/api/claude/status")
async def claude_status():
    """Check if Claude API is configured"""
    return {"available": check_api_key()}


@app.post("/api/claude/parse-building")
async def parse_building_natural_language(request: Request,
                                          input_data: NaturalLanguageInput):
    """Parse natural language to BuildingIntent, then resolve to config"""
    _enforce_llm_quota(request)
    try:
        # Step 1: Claude → BuildingIntent
        intent = parse_building(input_data.text)

        # Step 2: resolve_building_config
        if str(MCP_SERVER_PATH) not in sys.path:
            sys.path.insert(0, str(MCP_SERVER_PATH))
        from core.nl_resolver import resolve_building_config
        resolved = resolve_building_config(intent)

        return {
            "success": True,
            "intent": intent,
            "resolved": resolved,
        }
    except ValueError as e:
        # Expected validation error path — keep at info level.
        logger.info("NL parse rejected (ValueError): %s", e)
        return {"success": False, "error": str(e)}
    except Exception as e:
        logger.exception("NL parse-building failed")
        return {"success": False, "error": f"파싱 오류: {str(e)}"}


class BuildingResolveInput(BaseModel):
    """Direct intent → config resolution (no Claude)"""
    intent: Dict[str, Any]


@app.post("/api/building/resolve-config")
async def resolve_building_config_api(body: BuildingResolveInput):
    """Resolve BuildingIntent to validated config (without Claude)"""
    try:
        if str(MCP_SERVER_PATH) not in sys.path:
            sys.path.insert(0, str(MCP_SERVER_PATH))
        from core.nl_resolver import resolve_building_config
        resolved = resolve_building_config(body.intent)
        return {"success": True, "resolved": resolved}
    except Exception as e:
        logger.exception("resolve-config failed")
        return {"success": False, "error": str(e)}


# ═══════════════════════════════════════════════════════════════════════════════
# 3D Building Editor
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/editor-v2", response_class=HTMLResponse)
async def editor_v2_page(request: Request):
    """V2 3D Building Editor (alias for /)"""
    return templates.TemplateResponse(request, "editor_v2.html")


@app.get("/editor-lab", response_class=HTMLResponse)
async def editor_lab_page(request: Request):
    """ETABS-style UI redesign (isolated clone of V2; shares all /api endpoints)."""
    return templates.TemplateResponse(request, "editor_lab.html")


@app.get("/editor-figma", response_class=HTMLResponse)
async def editor_figma_page(request: Request):
    """Figma UI Lab: isolated redesign of V2 (own html/css/js copies; shares all /api endpoints)."""
    return templates.TemplateResponse(request, "editor_figma.html")


@app.get("/api/sections/list")
async def list_sections():
    """Return available sections grouped by type"""
    try:
        if str(MCP_SERVER_PATH) not in sys.path:
            sys.path.insert(0, str(MCP_SERVER_PATH))
        from core.simple_beam import get_available_sections
        sections = get_available_sections()
        return {"sections": sections}
    except Exception as e:
        return {"sections": {"H형강": ["H-200x200", "H-250x250", "H-300x300", "H-350x350", "H-400x400"]},
                "fallback": True, "error": str(e)}


@app.get("/api/sections/properties/{section_name:path}")
async def get_section_props(section_name: str):
    """Return section properties (A, Ix, Iy, J, h, b, tw, tf) for a given section"""
    try:
        if str(MCP_SERVER_PATH) not in sys.path:
            sys.path.insert(0, str(MCP_SERVER_PATH))
        from core.section_3d import get_section_3d
        sec = get_section_3d(section_name)
        # 단면 타입 판별
        stype = "H"
        t_mm = 0.0
        if section_name.startswith("□-") or section_name.startswith("\u25a1-"):
            stype = "SHS" if abs(sec.h - sec.b) < 1 else "RHS"
            t_mm = sec.tw  # SHS/RHS: tw에 두께 저장됨
        elif section_name.startswith("○-") or section_name.startswith("\u25cb-"):
            stype = "CHS"
            t_mm = sec.tw
        elif section_name.startswith("L-"):
            stype = "L"
        elif section_name.startswith("TFC-") or section_name.startswith("PFC-"):
            stype = "C"
        elif section_name.startswith("T-"):
            stype = "T"
        return {
            "name": sec.name,
            "section_type": stype,
            "A_cm2": round(sec.A / 100, 2),
            "Ix_cm4": round(sec.Ix / 10000, 1),
            "Iy_cm4": round(sec.Iy / 10000, 1),
            "J_cm4": round(sec.J / 10000, 1),
            "h_mm": round(sec.h, 1),
            "b_mm": round(sec.b, 1),
            "tw_mm": round(sec.tw, 1),
            "tf_mm": round(sec.tf, 1),
            "t_mm": round(t_mm, 1),
            "j_source": sec.j_source,
        }
    except Exception as e:
        return {"error": str(e), "name": section_name}


@app.get("/api/materials/list")
async def list_materials():
    """Return available materials"""
    try:
        if str(MCP_SERVER_PATH) not in sys.path:
            sys.path.insert(0, str(MCP_SERVER_PATH))
        from core.simple_beam import get_available_materials
        materials = get_available_materials()
        return {"materials": materials}
    except Exception as e:
        return {"materials": ["SS275", "SS400", "SM355", "SM490"],
                "fallback": True, "error": str(e)}


# 재료 팝업(Define Materials) 표시용 상수 — KDS 14 31 05 강구조 공통 물성.
# 해석 코어는 ν=0.3(G=E/2.6) 가정을 그대로 쓰므로 여기 값과 정합(frame_3d).
_STEEL_POISSON = 0.3
_STEEL_THERMAL_PER_C = 1.2e-5
_GRAVITY_M_S2 = 9.80665


def _steel_material_entry(name, standard, E_MPa, density_kg_m3, fy_bands,
                          fu_min=None, fu_max=None, elongation_min=None):
    """강종 1건을 Material 팝업 JSON으로 (파생값 G/단위중량 포함)."""
    return {
        "name": name,
        "type": "Steel",
        "standard": standard,
        "E_MPa": E_MPa,
        "poisson": _STEEL_POISSON,
        "G_MPa": round(E_MPa / (2.0 * (1.0 + _STEEL_POISSON)), 1),
        "thermal_coeff_per_C": _STEEL_THERMAL_PER_C,
        "density_kg_m3": density_kg_m3,
        "unit_weight_kN_m3": round(density_kg_m3 * _GRAVITY_M_S2 / 1000.0, 2),
        "fy_by_thickness": fy_bands,           # [{t_min, t_max, fy}] mm/MPa
        "fu_MPa": fu_min,
        "fu_max_MPa": (None if fu_max in (None, 999) else fu_max),
        "elongation_min_pct": elongation_min,
    }


@app.get("/api/materials/detail")
async def materials_detail():
    """KS 강재 카탈로그 상세 — Material 팝업용.

    Supabase materials 테이블(강종 × 두께구간 fy)을 등급 단위로 그룹핑하고
    표시/파생값(단위중량 kN/m³, G, ν, 열팽창계수)을 붙여 반환한다.
    """
    try:
        if str(MCP_SERVER_PATH) not in sys.path:
            sys.path.insert(0, str(MCP_SERVER_PATH))
        from core.simple_beam import get_supabase
        rows = (get_supabase().table("materials")
                .select("*").order("name").order("t_min").execute().data) or []
        grouped: dict[str, dict] = {}
        for r in rows:
            g = grouped.get(r["name"])
            if g is None:
                g = _steel_material_entry(
                    name=r["name"],
                    standard=r.get("standard"),
                    E_MPa=r.get("e") or 205000.0,
                    density_kg_m3=r.get("density") or 7850.0,
                    fy_bands=[],
                    fu_min=r.get("fu_min"),
                    fu_max=r.get("fu_max"),
                    elongation_min=r.get("elongation_min"),
                )
                grouped[r["name"]] = g
            g["fy_by_thickness"].append(
                {"t_min": r.get("t_min"), "t_max": r.get("t_max"), "fy": r.get("fy")})
        if not grouped:
            raise RuntimeError("materials table empty")
        return {"materials": list(grouped.values()),
                "constants": {"poisson": _STEEL_POISSON,
                              "thermal_coeff_per_C": _STEEL_THERMAL_PER_C}}
    except Exception as e:
        # Supabase 불가 시 최소 폴백 (해석 폴백 DEFAULT_MATERIALS와 동일 강종)
        fb = [
            _steel_material_entry("SS275", "KS D 3503", 205000.0, 7850.0,
                                  [{"t_min": 0, "t_max": 16, "fy": 275},
                                   {"t_min": 16, "t_max": 40, "fy": 265},
                                   {"t_min": 40, "t_max": 100, "fy": 245}],
                                  fu_min=410, fu_max=550, elongation_min=21),
            _steel_material_entry("SS235", "KS D 3503", 205000.0, 7850.0,
                                  [{"t_min": 0, "t_max": 16, "fy": 235},
                                   {"t_min": 16, "t_max": 40, "fy": 225}],
                                  fu_min=330, fu_max=450, elongation_min=26),
        ]
        return {"materials": fb, "fallback": True, "error": str(e),
                "constants": {"poisson": _STEEL_POISSON,
                              "thermal_coeff_per_C": _STEEL_THERMAL_PER_C}}


# Load 팝업(Live) 표시용 활하중 캐시 — Supabase 콜드콜(수 초) 반복 방지.
_LIVE_PREVIEW_CACHE: dict | None = None


@app.get("/api/loads/live-preview")
async def live_load_preview():
    """용도별 KDS 41 12 00 활하중 미리보기 — 하중 정의 팝업(Live)용.

    해석 시 load_generator가 조회하는 것과 동일한 경로
    (_query_live_load_traced)를 쓰므로 팝업 표시값 == 실제 적용값이 보장된다.
    반환: {"live_loads": {usage: {value, source, db_primary_key, db_secondary_key}}}
    """
    global _LIVE_PREVIEW_CACHE
    if _LIVE_PREVIEW_CACHE is not None:
        return _LIVE_PREVIEW_CACHE
    usages = ["office", "retail", "residential", "parking", "storage",
              "hospital", "school", "assembly", "corridor", "mechanical_room", "roof"]
    try:
        if str(MCP_SERVER_PATH) not in sys.path:
            sys.path.insert(0, str(MCP_SERVER_PATH))
        from core.load_generator import _query_live_load_traced
        out = {u: _query_live_load_traced(u) for u in usages}
        resp = {"live_loads": out, "source_code": "KDS 41 12 00 표 3.2-1"}
        # DB 폴백으로 채워진 응답은 캐시하지 않음 (다음 요청에서 재시도)
        if all(v.get("source") == "db_lookup" for v in out.values() if v):
            _LIVE_PREVIEW_CACHE = resp
        return resp
    except Exception as e:
        return {"live_loads": {}, "fallback": True, "error": str(e)}


@app.post("/api/building/analyze")
async def analyze_building_api(input_data: BuildingInput):
    """Run 3D building analysis and return results for the editor"""
    _enforce_demo_size(_estimate_member_count(input_data.config), "요청하신 모델")

    job_id = str(uuid.uuid4())
    job_dir = JOBS_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)

    # Save input config
    with open(job_dir / "config.json", "w", encoding="utf-8") as f:
        json.dump(input_data.config, f, indent=2, ensure_ascii=False)

    jobs_db[job_id] = {
        "job_id": job_id,
        "status": JobStatus.RUNNING,
        "progress": 10,
        "created_at": datetime.now().isoformat(),
        "analysis_type": "building_3d",
    }

    try:
        if str(MCP_SERVER_PATH) not in sys.path:
            sys.path.insert(0, str(MCP_SERVER_PATH))

        from core.building_model import BuildingModel
        from core.load_generator import generate_all_loads
        from core.frame_3d import analyze_frame_3d_multi

        # 1. BuildingModel
        model = BuildingModel.from_json(input_data.config)
        jobs_db[job_id]["progress"] = 20

        # 2. Auto load generation
        load_result = await run_solver(generate_all_loads, model)
        jobs_db[job_id]["progress"] = 40

        # 3. Run 3D analysis
        kwargs = model.to_frame3d_kwargs()
        kwargs["load_cases"] = load_result["load_cases"]
        kwargs["load_combinations"] = load_result["load_combinations"]
        kwargs["modal_analysis"] = True

        # Solver thread: serialises OpenSees global state and keeps the event
        # loop free while this multi-second run proceeds.
        multi = await run_solver(analyze_frame_3d_multi, **kwargs)
        jobs_db[job_id]["progress"] = 70

        # 4. Design check
        dc_result = None
        warnings: list[str] = []
        try:
            from core.design_check import run_design_check
            seismic_rpt = load_result["reports"].get("seismic")
            dc_result = await run_solver(run_design_check, multi, model, seismic_rpt)
        except Exception as dc_err:
            logger.warning("Building design check failed: %s", dc_err, exc_info=True)
            warnings.append(f"design_check_failed: {dc_err}")

        # 5. Result interpretation
        interpretation = None
        if dc_result is not None:
            try:
                from core.result_interpreter import interpret_results
                interpretation = await run_solver(
                    interpret_results, dc_result, multi,
                    modal_analysis=multi.modal_analysis or None,
                )
                # 결과 해설 LLM (역할 #3) — NARRATOR_API_KEY 있으면 Opus 산문, 없으면 identity.
                from core.narrative_interpreter import narrate_interpretation
                from core.narrative_llm import make_narrator_from_env
                interpretation = await run_solver(
                    narrate_interpretation,
                    interpretation, dc_result, llm=make_narrator_from_env(),
                    load_result=load_result,
                    model_info={
                        "num_stories": getattr(multi, "num_stories", None),
                        "total_height_m": getattr(multi, "total_height", None),
                        "material": getattr(multi, "material_name", None),
                        "column_section": getattr(multi, "column_section", None),
                        "beam_x_section": getattr(multi, "beam_x_section", None),
                        "rigid_diaphragm": getattr(model, "rigid_diaphragm", None),
                    },
                    analysis_metadata=getattr(multi, "analysis_metadata", None),
                    seismic_method=input_data.config.get("seismic_method", "ELF"),
                    analysis_id=job_id,
                )
            except Exception as interp_err:
                logger.warning(
                    "Building result interpretation failed: %s",
                    interp_err, exc_info=True,
                )
                warnings.append(f"result_interpretation_failed: {interp_err}")

        jobs_db[job_id]["progress"] = 85

        # 6. Extract 3D model data for viewer
        # Use structural grid nodes (not sub-element internal nodes)
        viewer_nodes = []
        for n in multi.nodes:
            if isinstance(n, dict):
                viewer_nodes.append({
                    "id": n["id"],
                    "x": n.get("x_m", n.get("x", 0)),
                    "y": n.get("y_m", n.get("y", 0)),
                    "z": n.get("z_m", n.get("z", 0)),
                })
            else:
                viewer_nodes.append({"id": n.id, "x": n.x, "y": n.y, "z": n.z})

        # Build structural members from member_info (supports both regular and irregular)
        viewer_elements = []
        _section_map = {
            "column": multi.column_section,
            "beam_x": multi.beam_x_section,
            "beam_y": multi.beam_y_section,
        }
        viewer_node_ids = {vn["id"] for vn in viewer_nodes}
        for minfo in (multi.member_info or []):
            ni_id = minfo.get("ni", 0)
            nj_id = minfo.get("nj", 0)
            etype = minfo.get("elem_type", "column")
            if ni_id in viewer_node_ids and nj_id in viewer_node_ids:
                viewer_elements.append({
                    "id": minfo["member_id"],
                    "ni": ni_id,
                    "nj": nj_id,
                    "type": etype,
                    "section": _section_map.get(etype, ""),
                })

        # 7. Envelope across combos
        env = {"max_dx_mm": 0, "max_dy_mm": 0, "max_dz_mm": 0,
               "max_drift_x": 0, "max_drift_y": 0,
               "max_moment_kNm": 0, "max_axial_kN": 0, "max_shear_kN": 0}
        for cn, cr in multi.combo_results.items():
            if abs(cr.max_displacement_x) > abs(env["max_dx_mm"]):
                env["max_dx_mm"] = round(cr.max_displacement_x, 3)
            if abs(cr.max_displacement_y) > abs(env["max_dy_mm"]):
                env["max_dy_mm"] = round(cr.max_displacement_y, 3)
            if abs(cr.max_displacement_z) > abs(env["max_dz_mm"]):
                env["max_dz_mm"] = round(cr.max_displacement_z, 3)
            if cr.max_drift_x > env["max_drift_x"]:
                env["max_drift_x"] = round(cr.max_drift_x, 6)
            if cr.max_drift_y > env["max_drift_y"]:
                env["max_drift_y"] = round(cr.max_drift_y, 6)
            if cr.max_moment > env["max_moment_kNm"]:
                env["max_moment_kNm"] = round(cr.max_moment, 2)
            if cr.max_axial > env["max_axial_kN"]:
                env["max_axial_kN"] = round(cr.max_axial, 2)
            if cr.max_shear > env["max_shear_kN"]:
                env["max_shear_kN"] = round(cr.max_shear, 2)

        # 7b. Per-case/combo results for frontend case selector
        case_names = list(multi.case_results.keys())
        combo_names = list(multi.combo_results.keys())
        case_data = {}

        def _extract_case_summary(cr):
            """Extract summary + per-node displacements from a case result."""
            summary = {
                "max_dx_mm": round(cr.max_displacement_x, 3),
                "max_dy_mm": round(cr.max_displacement_y, 3),
                "max_dz_mm": round(getattr(cr, 'max_displacement_z', 0), 3),
                "max_drift_x": round(cr.max_drift_x, 6),
                "max_drift_y": round(cr.max_drift_y, 6),
                "max_moment_kNm": round(cr.max_moment, 2),
                "max_axial_kN": round(cr.max_axial, 2),
                "max_shear_kN": round(cr.max_shear, 2),
            }
            displacements = {}
            for nd in cr.nodal_displacements:
                nid = nd.get("node_id", nd.get("node", 0))
                displacements[str(nid)] = [
                    round(nd.get("dx_mm", 0), 3),
                    round(nd.get("dy_mm", 0), 3),
                    round(nd.get("dz_mm", 0), 3),
                ]
            drifts = []
            for sd in cr.story_drifts:
                drifts.append({
                    "story": sd.get("story", 0),
                    "drift_x": round(sd.get("drift_x", 0), 6),
                    "drift_y": round(sd.get("drift_y", 0), 6),
                })
            return {"summary": summary, "displacements": displacements, "story_drifts": drifts}

        for cn, cr in multi.case_results.items():
            case_data[cn] = _extract_case_summary(cr)
        for cn, cr in multi.combo_results.items():
            case_data[cn] = _extract_case_summary(cr)

        # 7c. Response Spectrum Analysis (if requested)
        rsa_result = None
        seismic_method = input_data.config.get("seismic_method", "ELF")
        if seismic_method == "RSA" and multi.modal_analysis:
            try:
                from core.design_spectrum import compute_design_spectrum
                from core.response_spectrum_analysis import (
                    run_response_spectrum_analysis, rsa_result_to_case_data,
                )
                # Get seismic system parameters
                seismic_sys = input_data.config.get("seismic_system", "ordinary_moment_frame")
                R_map = {"special_moment_frame": 8.0, "intermediate_moment_frame": 4.5,
                         "ordinary_moment_frame": 3.5, "special_concentric_braced": 8.0}
                R = R_map.get(seismic_sys, 3.5)
                IE = model.importance_factor
                region = model.region or "서울"
                site_class = model.site_class or "S3"

                spectrum = compute_design_spectrum(region, site_class, IE)
                if spectrum.get("status") == "success":
                    rsa_result = run_response_spectrum_analysis(
                        modal_analysis=multi.modal_analysis,
                        spectrum_result=spectrum,
                        viewer_nodes=viewer_nodes,
                        stories=list(multi.stories),
                        bays_x=list(multi.bays_x),
                        bays_y=list(multi.bays_y),
                        combination_rule=input_data.config.get("rsa_combination", "CQC"),
                        direction_rule=input_data.config.get("rsa_direction", "30pct"),
                        IE=IE, R=R,
                    )
                    # Merge RSA case_data
                    rsa_cases = rsa_result_to_case_data(rsa_result)
                    case_data.update(rsa_cases)
                    case_names.append("__RSA__")  # Marker for frontend
                    combo_names.extend(rsa_cases.keys())
            except Exception as e:
                logger.warning("Building RSA failed: %s", e, exc_info=True)
                warnings.append(f"rsa_failed: {e}")

        # 8. Design check per-member results for coloring
        # Map design check member_id (1-based) to viewer structural member IDs
        member_checks = {}
        if dc_result and "member_check" in dc_result:
            ms = dc_result["member_check"]
            for mem in ms.get("members", []):
                mid = mem.get("member_id", mem.get("element_id", 0))
                ratios = mem.get("ratios", {})
                interaction = ratios.get("interaction", ratios.get("H1", 0))
                check_info = {
                    "status": mem.get("status", "OK"),
                    "interaction_ratio": interaction,
                    "governing": mem.get("governing_combo", ""),
                    "type": mem.get("type", ""),
                    "section": mem.get("section", ""),
                }
                # viewer member IDs match design check member_id (both 1-based structural members)
                member_checks[str(mid)] = check_info

        response = {
            "job_id": job_id,
            "status": "success",
            "building": model.summary(),
            "config": input_data.config,
            "viewer": {
                "nodes": viewer_nodes,
                "elements": viewer_elements,
                "stories": list(multi.stories),
                "bays_x": list(multi.bays_x),
                "bays_y": list(multi.bays_y),
                "total_height": multi.total_height,
                "total_width_x": multi.total_width_x,
                "total_width_y": multi.total_width_y,
                "column_section": multi.column_section,
                "beam_x_section": multi.beam_x_section,
                "beam_y_section": multi.beam_y_section,
                "material_name": multi.material_name,
                "is_irregular": multi.analysis_metadata.get("irregular", False),
                "zones": input_data.config.get("zones"),
            },
            "envelope": env,
            "case_names": case_names,
            "combo_names": combo_names,
            "case_data": case_data,
            "design_check": dc_result,
            "interpretation": interpretation,
            "member_checks": member_checks,
            "modal_analysis": multi.modal_analysis,
            "rsa": {
                "enabled": rsa_result is not None,
                "combination_rule": rsa_result.combination_rule if rsa_result else None,
                "direction_rule": rsa_result.direction_rule if rsa_result else None,
                "num_modes_used": rsa_result.num_modes_used if rsa_result else 0,
                "parameters": rsa_result.parameters if rsa_result else {},
                "story_drifts": rsa_result.story_drifts if rsa_result else [],
            } if seismic_method == "RSA" else None,
        }

        # Generate HTML report
        report_url = None
        try:
            report_path = str(job_dir / "report.html")
            # 문서형 구조계산서 리포트 우선, 실패 시 기존 탭 리포트로 폴백.
            try:
                from core.visualization_calc_report import plot_frame_3d_calc_report
                await run_solver(
                    plot_frame_3d_calc_report,
                    multi,
                    output_path=report_path,
                    design_check=dc_result,
                    interpretation=interpretation,
                    load_result=load_result,
                    cover_info=input_data.config.get("cover_info"),
                    data_out_path=str(job_dir / "calc_data.json"),
                )
            except Exception as calc_err:
                logger.warning(
                    "Calc report failed; fallback to tab report (job %s): %s",
                    job_id, calc_err, exc_info=True,
                )
                from core.visualization_3d import plot_frame_3d_interactive
                plot_frame_3d_interactive(
                    multi, output_path=report_path,
                    design_check=dc_result, interpretation=interpretation,
                )
            report_url = f"/api/jobs/{job_id}/report"
            jobs_db[job_id]["report_url"] = report_url
        except Exception as report_err:
            logger.warning(
                "Building HTML report failed (job %s): %s",
                job_id, report_err, exc_info=True,
            )
            warnings.append(f"report_generation_failed: {report_err}")

        response["report_url"] = report_url

        # Recommendation pipeline foundation (deterministic, no RAG/LLM).
        rec_metadata = {
            "member_info": multi.member_info,
            "building": model.summary() if hasattr(model, "summary") else None,
            "envelope": env,
            "combo_names": combo_names,
            "seismic_method": seismic_method,
            "material_name": getattr(multi, "material_name", None),
        }
        rec_payload = _build_recommendation_block(
            design_check=dc_result,
            raw_warnings=warnings,
            analysis_metadata=rec_metadata,
            stage="building_analyze",
            mode="analyze",
        )
        response["warnings"] = rec_payload["warnings"]
        response["warning_messages"] = rec_payload["warning_messages"]
        response["issues"] = rec_payload["issues"]
        response["recommendation_candidates"] = rec_payload["recommendation_candidates"]
        response["recommendation_summary"] = rec_payload["recommendation_summary"]

        # Cache the compact subset so the chat router's inspect_selection /
        # get_analysis_summary tools can read this baseline. /api/v2/analyze
        # does the same on its job_id (see main_simple.py:1077-1108) — without
        # this block, chat tools returned ``unknown analysis_id`` for legacy
        # /api/building/analyze jobs, which is the path V2 Editor's main
        # "Analyze" button actually uses (editor3d_v2.js runAnalysis ->
        # /api/building/analyze).
        _purge_expired_analysis_contexts()
        compact = build_compact_subset(
            env=env,
            member_info_list=multi.member_info,
            member_check=(dc_result or {}).get("member_check"),
            modal_analysis=multi.modal_analysis,
            material_name=getattr(multi, "material_name", None),
            num_stories=getattr(model, "num_stories", 0) or 0,
            num_elements=len(multi.member_info or []),
        )
        with _ANALYSIS_CONTEXT_LOCK:
            analysis_context_cache[job_id] = {
                **compact,
                # Compact = chat tools can read this entry, recommendation
                # endpoints will reject it with 422. See
                # ``_assert_full_context_for_recommendations``.
                "context_kind": CONTEXT_KIND_COMPACT,
                "expires_at": time.time() + _ANALYSIS_CONTEXT_TTL_SEC,
            }

        # Store for re-analysis
        jobs_db[job_id]["config"] = input_data.config
        jobs_db[job_id]["response"] = response
        jobs_db[job_id]["status"] = JobStatus.DONE
        jobs_db[job_id]["progress"] = 100
        jobs_db[job_id]["completed_at"] = datetime.now().isoformat()

        return response

    except HTTPException:
        raise
    except Exception as e:
        jobs_db[job_id]["status"] = JobStatus.FAILED
        jobs_db[job_id]["error"] = str(e)
        logger.exception("Building analysis failed (job_id=%s)", job_id)
        raise HTTPException(status_code=500, detail=str(e))



@app.get("/api/building/{job_id}")
async def get_building_result(job_id: str):
    """Get stored building analysis result.

    Jobs live in in-process memory only (see ``jobs_db``), so any job_id
    issued before the most recent server restart will return 404. The
    fix is to re-run the analysis — there is no persistent backing
    store yet.
    """
    if job_id not in jobs_db:
        raise HTTPException(
            status_code=404,
            detail=(
                f"Job '{job_id}' not found. Jobs are stored in process memory "
                "and are lost when the server restarts (e.g. Render sleep / "
                "redeploy). Please re-run the analysis to get a fresh job_id."
            ),
        )
    response = jobs_db[job_id].get("response")
    if not response:
        raise HTTPException(status_code=400, detail="No results available")
    return response


# ═══════════════════════════════════════════════════════════════════════════════
# V2 Node-Element Pipeline APIs
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/api/v2/parse-ifc")
async def parse_ifc_v2_api(request: Request, file: UploadFile = File(...)):
    """V2: IFC → Node-Element StructuralModel 변환"""
    import tempfile, os

    if not file.filename.lower().endswith(".ifc"):
        raise HTTPException(status_code=400, detail="IFC 파일(.ifc)만 업로드할 수 있습니다.")

    # Read in chunks with a running limit. `await file.read()` pulled the whole
    # upload into memory BEFORE the size check, so an oversized POST killed the
    # container before the check could reject it.
    max_bytes = IFC_MAX_UPLOAD_MB * 1024 * 1024
    declared = request.headers.get("content-length")
    if declared and declared.isdigit() and int(declared) > max_bytes:
        raise HTTPException(
            status_code=413,
            detail=f"파일 크기가 {IFC_MAX_UPLOAD_MB}MB를 초과합니다.")

    chunks: list[bytes] = []
    total = 0
    while True:
        chunk = await file.read(1024 * 1024)
        if not chunk:
            break
        total += len(chunk)
        if total > max_bytes:
            raise HTTPException(
                status_code=413,
                detail=f"파일 크기가 {IFC_MAX_UPLOAD_MB}MB를 초과합니다.")
        chunks.append(chunk)
    content = b"".join(chunks)

    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".ifc")
    try:
        os.write(tmp_fd, content)
        os.close(tmp_fd)

        if str(MCP_SERVER_PATH) not in sys.path:
            sys.path.insert(0, str(MCP_SERVER_PATH))
        from core.ifc_parser_v2 import parse_ifc_v2
        from core.visualization_v2 import generate_model_viewer

        model, validation = parse_ifc_v2(tmp_path)

        # HTML 뷰어 생성 — 실패는 응답에 warning으로 노출 (success는 유지)
        viewer_path = None
        warnings: list[str] = []
        try:
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            viewer_dir = JOBS_DIR / "v2_viewers"
            viewer_dir.mkdir(parents=True, exist_ok=True)
            # Unguessable name. A second-resolution timestamp is only 86,400
            # candidates a day, and /api/v2/viewer/{filename} has no ownership
            # check — so anyone could enumerate another visitor's uploaded
            # model. The uuid also removes the same-second overwrite collision.
            # The original filename is deliberately NOT put in the title: it
            # leaks the uploader's project name to whoever opens the viewer.
            viewer_path = generate_model_viewer(
                model,
                output_path=str(viewer_dir / f"model_{ts}_{uuid.uuid4().hex}.html"),
                title="IFC Model",
            )
        except Exception as viewer_err:
            logger.warning(
                "IFC viewer generation failed for %s: %s",
                file.filename, viewer_err, exc_info=True,
            )
            warnings.append(f"viewer_generation_failed: {viewer_err}")

        # Structured warnings via the recommendation pipeline helper.
        # parse-ifc is a pre-analysis stage so we run in parse_only mode:
        # missing_design_check is NOT raised, and no recommendation
        # candidates are generated yet.
        rec_payload = _build_recommendation_block(
            design_check=None,
            raw_warnings=warnings,
            stage="parse_ifc",
            mode="parse_only",
        )

        result = {
            "success": validation.is_valid,
            "filename": file.filename,
            "model": model.to_json(),
            "validation": {
                "is_valid": validation.is_valid,
                "summary": validation.summary_text(),
                "extracted_nodes": validation.extracted_nodes,
                "extracted_elements": validation.extracted_elements,
                "failed_elements": validation.failed_elements,
                "issues": [
                    {"severity": i.severity.value, "code": i.code,
                     "message": i.message, "default_value": i.default_value}
                    for i in validation.issues
                ],
                "needs_user_input": [
                    {"code": i.code, "message": i.message, "default_value": i.default_value}
                    for i in validation.needs_user_input
                ],
            },
            "summary": model.summary(),
            "warnings": rec_payload["warnings"],
            "warning_messages": rec_payload["warning_messages"],
        }
        if viewer_path:
            result["viewer_url"] = f"/api/v2/viewer/{Path(viewer_path).name}"

        return result

    except ImportError:
        raise HTTPException(status_code=500, detail="ifcopenshell 패키지가 필요합니다.")
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("IFC V2 parse failed for %s", file.filename)
        raise HTTPException(status_code=500, detail=f"IFC V2 파싱 오류: {str(e)}")
    finally:
        try:
            os.unlink(tmp_path)
        except OSError as cleanup_err:
            logger.warning("Failed to cleanup IFC temp file %s: %s", tmp_path, cleanup_err)


@app.post("/api/v2/snap-joints")
async def snap_joints_api(request: Request):
    """V2: 보-기둥 접합 스냅"""
    body = await request.json()
    model_json = body.get("model")
    snap_tolerance = body.get("snap_tolerance", 0.5)

    if not model_json:
        raise HTTPException(status_code=400, detail="model JSON이 필요합니다.")

    if str(MCP_SERVER_PATH) not in sys.path:
        sys.path.insert(0, str(MCP_SERVER_PATH))
    from core.structural_model import StructuralModel
    from core.ifc_parser_v2 import snap_beams_to_columns

    model = StructuralModel.from_json(model_json)
    nodes_before = len(model.nodes)
    snapped = snap_beams_to_columns(model, snap_tolerance=snap_tolerance)

    return {
        "success": True,
        "snapped_count": snapped,
        "nodes_before": nodes_before,
        "nodes_after": len(model.nodes),
        "model": model.to_json(),
        "summary": model.summary(),
    }


@app.post("/api/v2/analyze")
async def analyze_v2_api(request: Request):
    """V2: 통합 해석 (KDS 하중자동생성 + 해석 + 설계검토 + 리포트).

    V1의 /api/building/analyze와 동일한 기능을 V2 StructuralModel에서 수행.
    """
    body = await request.json()
    model_json = body.get("model")
    user_config = body.get("config", {})
    # 표지·도장란 — 분석 요청에 동반되면 첫 리포트부터 반영. 없으면 None(=placeholder),
    # 이후 /api/jobs/{job_id}/report-cover 로 재해석 없이 주입 가능.
    cover_info = body.get("cover_info") or user_config.get("cover_info")

    if str(MCP_SERVER_PATH) not in sys.path:
        sys.path.insert(0, str(MCP_SERVER_PATH))

    from core.structural_model import StructuralModel
    from core.frame_3d import analyze_from_model

    # 입력 분기: {model} = IFC/편집기 경로(node-element 명시),
    # {config} = 직접입력/NL 경로(high-level config → 자동 그리드 생성)
    if model_json:
        model = StructuralModel.from_json(model_json)
        config_built = False
    else:
        if not user_config:
            raise HTTPException(
                status_code=400,
                detail="model 또는 config 중 하나가 필요합니다.",
            )
        # config-only 진입 시 user_config가 곧 BuildingModel 입력
        model = StructuralModel.from_building_config(user_config)
        config_built = True

    warnings: list[str] = []  # 부분 실패 정보 (success 응답에 포함)

    # An uploaded/edited V2 model carries its elements directly, so count them
    # rather than estimating from a grid config.
    _enforce_demo_size(len(getattr(model, "elements", {}) or {}), "이 모델")

    if not config_built:
        # 요소 분류 확인 + 자동 분류 (IFC는 elem_type이 누락된 경우가 있음)
        from core.structural_model import ElementType as _ET
        type_set = set(e.elem_type for e in model.elements.values())
        logger.info("[V2] Element types before classify: %s", [t.value for t in type_set])
        if len(type_set) == 1 and len(model.elements) > 3:
            model.classify_elements()
            type_set2 = set(e.elem_type for e in model.elements.values())
            logger.info("[V2] Auto-classified: %s", [t.value for t in type_set2])

        # A-1: 근접 노드 병합 (보-기둥 접합 보장)
        merged = model.merge_nearby_nodes()  # 적응형: 단면 높이 기반 자동 계산
        if merged > 0:
            logger.info("[V2] Merged %d nearby nodes", merged)

        # 보-보 교차점 자동 분할
        intersections = model.split_at_intersections()
        if intersections > 0:
            logger.info(
                "[V2] Created %d intersection nodes, model: %d nodes, %d elems",
                intersections, len(model.nodes), len(model.elements),
            )
    else:
        logger.info(
            "[V2] config-built model: %d nodes, %d elems (post-process skipped)",
            len(model.nodes), len(model.elements),
        )

    # 사용자 config 반영
    for key in ["region", "site_class", "importance", "seismic_system", "exposure_category", "geometric_nonlinearity"]:
        if user_config.get(key):
            setattr(model, key, user_config[key])
    if user_config.get("importance"):
        model.importance_factor = {"특": 1.5, "I": 1.2, "II": 1.0}.get(model.importance, 1.0)
    if user_config.get("rigid_diaphragm") is not None:
        model.rigid_diaphragm = user_config["rigid_diaphragm"]

    # 층별 용도/슬래브 설정
    for sc in user_config.get("stories", []):
        s_idx = sc.get("story")
        if s_idx:
            if sc.get("usage"): model.story_usages[s_idx] = sc["usage"]
            if sc.get("slab_thickness"): model.story_slab_thickness[s_idx] = sc["slab_thickness"]
            if sc.get("dead_load_finish") is not None: model.story_dead_load_finish[s_idx] = sc["dead_load_finish"]

    # 기본값 채우기
    for s_idx in range(1, model.num_stories + 1):
        model.story_usages.setdefault(s_idx, "office")
        model.story_slab_thickness.setdefault(s_idx, 0.15)
        model.story_dead_load_finish.setdefault(s_idx, 1.0)

    job_id = str(uuid.uuid4())
    job_dir = JOBS_DIR / job_id
    job_dir.mkdir(parents=True, exist_ok=True)

    try:
        # ── Step 1: 하중 생성 (V1 load_generator 사용) ──
        building_model = model.to_building_model()

        from core.load_generator import generate_all_loads
        load_result = await run_solver(generate_all_loads, building_model)
        load_cases = load_result["load_cases"]
        load_combinations = load_result.get("load_combinations", {})

        # ── Step 2: 해석 ──
        # Solver thread — see the V1 endpoint above.
        multi = await run_solver(analyze_from_model, model, load_cases, load_combinations)

        # ── Step 3: 설계검토 ──
        dc_result = None
        interpretation = None
        seismic_rpt = None  # default so analysis_context cache stays consistent
                            # even when run_design_check raises early
        try:
            from core.design_check import run_design_check
            seismic_rpt = load_result.get("reports", {}).get("seismic")
            dc_result = await run_solver(run_design_check, multi, building_model,
                                         seismic_rpt)
        except Exception as dc_err:
            logger.warning("V2 design check failed: %s", dc_err, exc_info=True)
            warnings.append(f"design_check_failed: {dc_err}")

        try:
            from core.result_interpreter import interpret_results
            if dc_result:
                interpretation = interpret_results(dc_result, multi,
                    modal_analysis=getattr(multi, 'modal_analysis', None) or None)
                # 결과 해설 LLM (역할 #3) — NARRATOR_API_KEY 있으면 Opus 산문, 없으면 identity.
                from core.narrative_interpreter import narrate_interpretation
                from core.narrative_llm import make_narrator_from_env
                interpretation = await run_solver(
                    narrate_interpretation,
                    interpretation, dc_result, llm=make_narrator_from_env(),
                    load_result=load_result,
                    model_info={
                        "num_stories": getattr(multi, "num_stories", None),
                        "total_height_m": getattr(multi, "total_height", None),
                        "material": getattr(multi, "material_name", None),
                        "column_section": getattr(multi, "column_section", None),
                        "beam_x_section": getattr(multi, "beam_x_section", None),
                        "rigid_diaphragm": getattr(model, "rigid_diaphragm", None),
                    },
                    analysis_metadata=getattr(multi, "analysis_metadata", None),
                    seismic_method=user_config.get("seismic_method", "ELF"),
                    analysis_id=job_id,
                )
        except Exception as interp_err:
            logger.warning("V2 result interpretation failed: %s", interp_err, exc_info=True)
            warnings.append(f"result_interpretation_failed: {interp_err}")

        # ── Step 4: HTML 리포트 ──
        report_path = str(job_dir / "report.html")
        try:
            # 문서형 구조계산서 리포트 우선, 실패 시 기존 탭 리포트로 폴백.
            try:
                from core.visualization_calc_report import plot_frame_3d_calc_report
                await run_solver(plot_frame_3d_calc_report, multi,
                                 output_path=report_path,
                                 design_check=dc_result, interpretation=interpretation,
                                 load_result=load_result, cover_info=cover_info,
                                 data_out_path=str(job_dir / "calc_data.json"))
            except Exception as calc_err:
                logger.warning("V2: calc report failed; fallback to tab report: %s",
                               calc_err, exc_info=True)
                from core.visualization_3d import plot_frame_3d_interactive
                plot_frame_3d_interactive(multi, output_path=report_path,
                                          design_check=dc_result, interpretation=interpretation)
        except Exception as viz_err:
            logger.warning(
                "V2: primary HTML report failed (%s); falling back to V2 viewer.",
                viz_err, exc_info=True,
            )
            warnings.append(f"report_primary_failed: {viz_err}")
            try:
                from core.visualization_v2 import generate_model_viewer
                generate_model_viewer(model, result=multi, output_path=report_path,
                                      title="V2 Analysis Results")
            except Exception as fallback_err:
                logger.error(
                    "V2: fallback viewer also failed: %s",
                    fallback_err, exc_info=True,
                )
                warnings.append(f"report_fallback_failed: {fallback_err}")

        report_url = f"/api/jobs/{job_id}/report" if Path(report_path).exists() else None
        if report_url is None:
            warnings.append("report_unavailable: HTML report could not be generated")

        # ── Step 5: 응답 ──
        env = {"max_dx_mm": 0, "max_dy_mm": 0, "max_dz_mm": 0,
               "max_drift_x": 0, "max_drift_y": 0,
               "max_moment_kNm": 0, "max_axial_kN": 0, "max_shear_kN": 0}
        for cr in list(multi.case_results.values()) + list(multi.combo_results.values()):
            env["max_dx_mm"] = max(env["max_dx_mm"], abs(cr.max_displacement_x))
            env["max_dy_mm"] = max(env["max_dy_mm"], abs(cr.max_displacement_y))
            env["max_dz_mm"] = max(env["max_dz_mm"], abs(cr.max_displacement_z))
            env["max_drift_x"] = max(env["max_drift_x"], abs(cr.max_drift_x))
            env["max_drift_y"] = max(env["max_drift_y"], abs(cr.max_drift_y))
            env["max_moment_kNm"] = max(env["max_moment_kNm"], abs(cr.max_moment))
            env["max_axial_kN"] = max(env["max_axial_kN"], abs(cr.max_axial))
            env["max_shear_kN"] = max(env["max_shear_kN"], abs(cr.max_shear))

        case_data = {}
        for cname, cr in {**multi.case_results, **multi.combo_results}.items():
            # 변위 데이터 (에디터 deformed shape용)
            disp_map = {}
            for d in cr.nodal_displacements:
                disp_map[str(d["node"])] = [d["dx_mm"], d["dy_mm"], d["dz_mm"]]
            case_data[cname] = {
                "summary": {"max_dx_mm": cr.max_displacement_x, "max_dy_mm": cr.max_displacement_y,
                            "max_moment_kNm": cr.max_moment, "max_axial_kN": cr.max_axial},
                "story_drifts": cr.story_drifts,
                "reactions": cr.reactions,
                "displacements": disp_map,
            }

        # ── Response Spectrum Analysis (if requested) ──
        rsa_result = None
        seismic_method = user_config.get("seismic_method", "ELF")
        modal_analysis_result = getattr(multi, 'modal_analysis', None) or None
        extra_combo_names = []
        if seismic_method == "RSA" and modal_analysis_result:
            try:
                from core.design_spectrum import compute_design_spectrum
                from core.response_spectrum_analysis import (
                    run_response_spectrum_analysis, rsa_result_to_case_data,
                )
                seismic_sys = user_config.get("seismic_system", model.seismic_system)
                R_map = {"special_moment_frame": 8.0, "intermediate_moment_frame": 4.5,
                         "ordinary_moment_frame": 3.5, "special_concentric_braced": 8.0}
                R = R_map.get(seismic_sys, 3.5)
                IE = model.importance_factor
                region = model.region or "서울"
                site_class = model.site_class or "S3"

                spectrum = compute_design_spectrum(region, site_class, IE)
                if spectrum.get("status") == "success":
                    # V2 viewer_nodes format
                    viewer_nodes_rsa = [
                        {"id": n.id, "x_m": n.x, "y_m": n.y, "z_m": n.z}
                        for n in sorted(model.nodes.values(), key=lambda n: n.id)
                    ]
                    rsa_result = run_response_spectrum_analysis(
                        modal_analysis=modal_analysis_result,
                        spectrum_result=spectrum,
                        viewer_nodes=viewer_nodes_rsa,
                        stories=list(multi.stories),
                        bays_x=list(multi.bays_x) if multi.bays_x else [8.0],
                        bays_y=list(multi.bays_y) if multi.bays_y else [8.0],
                        combination_rule=user_config.get("rsa_combination", "CQC"),
                        direction_rule=user_config.get("rsa_direction", "30pct"),
                        IE=IE, R=R,
                    )
                    # Merge RSA case_data
                    rsa_cases = rsa_result_to_case_data(rsa_result)
                    case_data.update(rsa_cases)
                    extra_combo_names = list(rsa_cases.keys())
            except Exception as rsa_err:
                logger.warning("V2 RSA failed: %s", rsa_err, exc_info=True)
                warnings.append(f"rsa_failed: {rsa_err}")

        member_checks = {}
        if dc_result and "member_check" in dc_result:
            for mem in dc_result["member_check"].get("members", []):
                member_checks[str(mem.get("member_id", ""))] = {
                    "status": mem.get("status", "OK"),
                    "interaction_ratio": mem.get("ratios", {}).get("interaction", 0),
                }

        # Recommendation pipeline foundation — deterministic layer only.
        # Adds `issues`, `recommendation_candidates`, and a structured
        # warning object alongside the legacy string list. NO RAG/LLM.
        #
        # The analysis_metadata payload lets the issue extractor enrich
        # issues with member section/story/connected-nodes context.
        # Whole model + member_forces are intentionally NOT included —
        # they'd inflate the recommendation surface; member_info +
        # updated_model is enough for location resolution.
        rec_metadata = {
            "member_info": multi.member_info,
            "updated_model": model.to_json(),
            "building": model.summary(),
            "envelope": env,
            "combo_names": list(multi.combo_results.keys()) + extra_combo_names,
            "load_summary": load_result.get("summary", {}),
            "seismic_method": seismic_method,
            "material_name": getattr(multi, "material_name", None)
                            or getattr(model, "material_name", None),
        }
        rec_payload = _build_recommendation_block(
            design_check=dc_result,
            raw_warnings=warnings,
            analysis_metadata=rec_metadata,
            stage="v2_analyze",
            mode="analyze",
        )

        response = {
            "job_id": job_id,
            "analysis_id": job_id,  # alias used by /api/v2/recommendations/evaluate
            "status": "success",
            "pipeline": "v2_node_element",
            "building": model.summary(),
            "updated_model": model.to_json(),  # split 후 모델 (JS 동기화용)
            "envelope": env,
            "case_names": list(multi.case_results.keys()) + (["__RSA__"] if rsa_result else []),
            "combo_names": list(multi.combo_results.keys()) + extra_combo_names,
            "case_data": case_data,
            "load_summary": load_result.get("summary", {}),
            "load_cases_raw": load_result.get("load_cases", {}),
            "member_forces": multi.member_forces,
            "member_info": multi.member_info,
            "design_check": dc_result,
            "interpretation": interpretation,
            "member_checks": member_checks,
            "modal_analysis": modal_analysis_result,
            "rsa": rsa_result,
            "seismic_method": seismic_method,
            "report_url": report_url,
            # Structured warnings + legacy string mirror for frontend back-compat.
            "warnings": rec_payload["warnings"],
            "warning_messages": rec_payload["warning_messages"],
            # Recommendation foundation (deterministic, no RAG/LLM yet).
            "issues": rec_payload["issues"],
            "recommendation_candidates": rec_payload["recommendation_candidates"],
            "recommendation_summary": rec_payload["recommendation_summary"],
        }

        # Cache the context needed to reproduce this baseline when the
        # frontend later calls /api/v2/recommendations/evaluate. Keyed by
        # job_id so the client only has to send the candidate ids back.
        _purge_expired_analysis_contexts()
        candidates_by_id = {
            c["candidate_id"]: c for c in rec_payload["recommendation_candidates"]
        }
        # Phase 0 Step 0-2: pre-derive the compact subset the chat router's
        # inspect/summary tools will read. Computed outside the lock since it
        # is pure-functional.
        compact = build_compact_subset(
            env=env,
            member_info_list=multi.member_info,
            member_check=(dc_result or {}).get("member_check"),
            modal_analysis=modal_analysis_result,
            material_name=(
                getattr(multi, "material_name", None)
                or getattr(model, "material_name", None)
            ),
            num_stories=getattr(model, "num_stories", 0) or 0,
            num_elements=len(getattr(model, "elements", []) or []),
        )
        with _ANALYSIS_CONTEXT_LOCK:
            analysis_context_cache[job_id] = {
                "model_json": model.to_json(),
                "load_cases": load_cases,
                "load_combinations": load_combinations,
                "building_model": building_model,
                "seismic_report": seismic_rpt,
                "design_check": dc_result,
                "candidates_by_id": candidates_by_id,
                # Chat-router lookups (see services.analysis_context.build_compact_subset)
                **compact,
                # Full = recommendation endpoints can use this entry.
                # See ``_assert_full_context_for_recommendations``.
                "context_kind": CONTEXT_KIND_FULL,
                "expires_at": time.time() + _ANALYSIS_CONTEXT_TTL_SEC,
            }

        # Job DB 저장
        jobs_db[job_id] = {
            "job_id": job_id, "status": "done",
            "created_at": datetime.now().isoformat(),
            "completed_at": datetime.now().isoformat(),
            "report_url": report_url,
            "summary": {"max_displacement_x_mm": env["max_dx_mm"], "max_displacement_y_mm": env["max_dy_mm"],
                        "max_drift": env["max_drift_x"], "max_drift_story": 1,
                        "max_moment_kNm": env["max_moment_kNm"], "max_shear_kN": env["max_shear_kN"],
                        "max_axial_kN": env["max_axial_kN"], "base_shear_kN": 0,
                        "num_stories": model.num_stories, "num_bays": 0},
        }

        # Export용 결과 디스크 저장 (case_data, member_forces, envelope 등)
        try:
            export_payload = {
                "job_id": job_id,
                "created_at": jobs_db[job_id]["created_at"],
                "envelope": env,
                "case_names": response["case_names"],
                "combo_names": response["combo_names"],
                "case_data": case_data,
                "member_forces": multi.member_forces,
                "member_info": multi.member_info,
                "nodes_info": {str(n.id): [n.x, n.y, n.z] for n in model.nodes.values()},
                "num_stories": model.num_stories,
                "num_nodes": len(model.nodes),
                "num_elements": len(model.elements),
                "analysis_type": "v2_building",
                "seismic_method": seismic_method,
            }
            with open(job_dir / "result.json", "w", encoding="utf-8") as f:
                json.dump(export_payload, f, ensure_ascii=False, default=str)
        except Exception as save_err:
            logger.warning(
                "V2: failed to persist result.json for job %s: %s",
                job_id, save_err, exc_info=True,
            )
            warnings.append(f"export_save_failed: {save_err}")

        return response

    except HTTPException:
        raise
    except Exception as e:
        logger.exception("V2 analyze failed (job_id=%s)", job_id)
        raise HTTPException(status_code=500, detail=f"V2 해석 오류: {str(e)}")


# ═══════════════════════════════════════════════════════════════════════════════
# Recommendation candidate evaluate (reanalyzed verified scoring)
# ═══════════════════════════════════════════════════════════════════════════════

def _make_eval_job_record(analysis_id: str, candidate_ids: list[str]) -> dict:
    return {
        "job_id": None,                  # filled by caller
        "analysis_id": analysis_id,
        "status": "queued",              # queued | running | done | failed
        "progress": {"completed": 0, "total": len(candidate_ids)},
        "candidate_ids_requested": list(candidate_ids),
        "evaluated_candidates": None,
        "rejected_candidates": None,
        "recommendation_summary": None,
        "ranked_order": None,
        "error": None,
        "created_at": datetime.now().isoformat(),
        "started_at": None,
        "completed_at": None,
    }


def _run_evaluation_job(eval_job_id: str) -> None:
    """Background worker: runs through one batch of candidates serially.

    OpenSees has process-global state, so this executes on a single
    worker thread (``MAX_PARALLEL_EVALS=1``). Errors for individual
    candidates are captured as ``rejected_analysis_failed`` entries;
    only an outright job-level failure flips the job status to
    ``failed``.
    """
    if str(MCP_SERVER_PATH) not in sys.path:
        sys.path.insert(0, str(MCP_SERVER_PATH))

    with _EVAL_JOBS_LOCK:
        job = eval_jobs_db.get(eval_job_id)
        if job is None:
            return
        job["status"] = "running"
        job["started_at"] = datetime.now().isoformat()
        analysis_id = job["analysis_id"]
        candidate_ids = list(job["candidate_ids_requested"])

    try:
        ctx = _get_analysis_context(analysis_id)

        if ctx is None:
            with _EVAL_JOBS_LOCK:
                job["status"] = "failed"
                job["error"] = (
                    f"analysis context for '{analysis_id}' is gone "
                    "(expired or server restarted) — re-run /api/v2/analyze"
                )
                job["completed_at"] = datetime.now().isoformat()
            return

        from core.recommendation import (
            BaselineSummary,
            RetrofitCandidate,
            evaluate_candidate,
            rank_evaluated_candidates,
            summarize_evaluations,
            partition_evaluations,
        )

        baseline = BaselineSummary.from_design_check(ctx.get("design_check"))
        candidates_by_id = ctx.get("candidates_by_id", {})

        evaluations = []
        completed = 0
        for cid in candidate_ids:
            cand_dict = candidates_by_id.get(cid)
            if cand_dict is None:
                # Caller asked for an unknown id — fail this one but
                # keep going.
                from core.recommendation import (
                    CandidateEvaluation,
                    STATUS_SKIPPED_INAPPLICABLE,
                )
                evaluations.append(CandidateEvaluation(
                    candidate_id=cid,
                    status=STATUS_SKIPPED_INAPPLICABLE,
                    error="unknown candidate_id — not present in /api/v2/analyze response",
                ))
            else:
                cand = _rehydrate_candidate(cand_dict)
                ev = evaluate_candidate(
                    model_json=ctx["model_json"],
                    load_cases=ctx["load_cases"],
                    load_combinations=ctx.get("load_combinations") or {},
                    candidate=cand,
                    baseline=baseline,
                    building_model=ctx.get("building_model"),
                    seismic_report=ctx.get("seismic_report"),
                )
                evaluations.append(ev)

            completed += 1
            with _EVAL_JOBS_LOCK:
                job["progress"] = {
                    "completed": completed, "total": len(candidate_ids),
                }

        # Rank survivors + partition.
        applicable, rejected = partition_evaluations(evaluations)
        ranked = rank_evaluated_candidates(applicable)
        summary = summarize_evaluations(evaluations)
        summary["analysis_id"] = analysis_id

        with _EVAL_JOBS_LOCK:
            job["evaluated_candidates"] = [r.to_dict() for r in ranked]
            job["rejected_candidates"] = [r.to_dict() for r in rejected]
            job["recommendation_summary"] = summary
            job["ranked_order"] = [r.candidate_id for r in ranked]
            job["status"] = "done"
            job["completed_at"] = datetime.now().isoformat()

    except Exception as e:  # noqa: BLE001 — top-level safety net
        logger.exception("evaluate job %s crashed", eval_job_id)
        with _EVAL_JOBS_LOCK:
            job["status"] = "failed"
            job["error"] = f"{type(e).__name__}: {e}"
            job["completed_at"] = datetime.now().isoformat()


@app.post("/api/v2/recommendations/evaluate")
async def post_recommendations_evaluate(request: Request):
    """Queue a batch reanalysis of candidates from a recent /api/v2/analyze run.

    Request:
        {"analysis_id": "...",  "candidate_ids": ["...", ...]}

    Response (202-style payload):
        {"job_id": "rec_eval_<uuid>", "status": "queued",
         "num_total": int, "num_applicable_requested": int}
    """
    body = await request.json()
    analysis_id = body.get("analysis_id")
    candidate_ids = body.get("candidate_ids") or []

    if not analysis_id:
        raise HTTPException(status_code=400, detail="analysis_id is required")
    if not isinstance(candidate_ids, list) or not candidate_ids:
        raise HTTPException(
            status_code=400,
            detail="candidate_ids must be a non-empty list of strings",
        )

    # Distinguish "never existed" (400 Bad Request) from "expired / evicted"
    # (410 Gone) so a client can tell the user which way to recover. The
    # peek-then-get sequence is intentional: we want to know whether the
    # entry was present at all (regardless of TTL) before we treat the
    # miss as a not-found vs gone.
    with _ANALYSIS_CONTEXT_LOCK:
        ever_existed = analysis_id in analysis_context_cache
    ctx = _get_analysis_context(analysis_id)
    if ctx is None:
        if ever_existed:
            # _get_analysis_context just evicted it because it was past
            # its expires_at — surface that as 410 Gone.
            raise HTTPException(
                status_code=410,
                detail=(
                    f"analysis context for '{analysis_id}' has expired "
                    "(30-minute TTL). Re-run /api/v2/analyze first."
                ),
            )
        raise HTTPException(
            status_code=400,
            detail=(
                f"analysis context for '{analysis_id}' not found. It "
                "may have never been created or the server was restarted. "
                "Re-run /api/v2/analyze first."
            ),
        )

    _assert_full_context_for_recommendations(ctx, analysis_id)

    eval_job_id = f"rec_eval_{uuid.uuid4().hex[:12]}"
    record = _make_eval_job_record(analysis_id, candidate_ids)
    record["job_id"] = eval_job_id

    # Count how many of the requested ids are actually applicable so the
    # caller can show "evaluating 3 of 4" without round-tripping.
    candidates_by_id = ctx.get("candidates_by_id", {})
    num_applicable = sum(
        1 for cid in candidate_ids
        if (candidates_by_id.get(cid) or {})
        .get("proposed_change", {})
        .get("applicable", False)
    )

    with _EVAL_JOBS_LOCK:
        eval_jobs_db[eval_job_id] = record

    _eval_executor.submit(_run_evaluation_job, eval_job_id)

    return {
        "job_id": eval_job_id,
        "status": "queued",
        "num_total": len(candidate_ids),
        "num_applicable_requested": num_applicable,
        "analysis_id": analysis_id,
    }


@app.get("/api/v2/recommendations/evaluate/{eval_job_id}")
async def get_recommendations_evaluate(eval_job_id: str):
    """Poll the state of a batch evaluation job."""
    with _EVAL_JOBS_LOCK:
        job = eval_jobs_db.get(eval_job_id)
        if job is None:
            raise HTTPException(
                status_code=404,
                detail=(
                    f"eval job '{eval_job_id}' not found — server may have "
                    "restarted (eval state is in-memory only)."
                ),
            )
        # Return a shallow copy so the caller can't accidentally mutate
        # the live record by holding the dict reference.
        return {
            "job_id": job["job_id"],
            "analysis_id": job["analysis_id"],
            "status": job["status"],
            "progress": dict(job.get("progress") or {}),
            "evaluated_candidates": job.get("evaluated_candidates"),
            "rejected_candidates": job.get("rejected_candidates"),
            "recommendation_summary": job.get("recommendation_summary"),
            "ranked_order": job.get("ranked_order"),
            "error": job.get("error"),
            "created_at": job.get("created_at"),
            "started_at": job.get("started_at"),
            "completed_at": job.get("completed_at"),
        }


@app.post("/api/v2/recommendations/preview-apply")
async def post_recommendations_preview_apply(request: Request):
    """Materialize a candidate against the cached analysis model.

    Re-runs apply_candidate_to_model on the cached ``model_json`` and
    returns the diff plus the resulting model. The frontend uses the
    returned ``updated_model`` to swap ``window._v2Model`` before
    triggering reanalysis. apply_candidate_to_model deepcopies its
    input, so repeated calls are safe — the cached entry stays pristine.

    Request:
        {"analysis_id": "...", "candidate_id": "..."}

    Response:
        {"candidate_id", "applicable", "diff": {...}, "updated_model": {...}}
    """
    body = await request.json()
    analysis_id = body.get("analysis_id")
    candidate_id = body.get("candidate_id")

    if not analysis_id:
        raise HTTPException(status_code=400, detail="analysis_id is required")
    if not candidate_id:
        raise HTTPException(status_code=400, detail="candidate_id is required")

    # Distinguish "never existed" (400) from "expired and evicted" (410),
    # mirroring the /evaluate endpoint exactly.
    with _ANALYSIS_CONTEXT_LOCK:
        ever_existed = analysis_id in analysis_context_cache
    ctx = _get_analysis_context(analysis_id)
    if ctx is None:
        if ever_existed:
            raise HTTPException(
                status_code=410,
                detail=(
                    f"analysis context for '{analysis_id}' has expired "
                    "(30-minute TTL). Re-run /api/v2/analyze first."
                ),
            )
        raise HTTPException(
            status_code=400,
            detail=(
                f"analysis context for '{analysis_id}' not found. It "
                "may have never been created or the server was restarted. "
                "Re-run /api/v2/analyze first."
            ),
        )

    _assert_full_context_for_recommendations(ctx, analysis_id)

    cand_dict = (ctx.get("candidates_by_id") or {}).get(candidate_id)
    if cand_dict is None:
        raise HTTPException(
            status_code=404,
            detail=f"candidate '{candidate_id}' not found in analysis context",
        )

    proposed = cand_dict.get("proposed_change") or {}
    if not proposed.get("applicable", False):
        raise HTTPException(
            status_code=422,
            detail=(
                f"candidate '{candidate_id}' is not applicable "
                f"(operation='{proposed.get('operation', '?')}', "
                "reason='abstract candidate — manual review required')"
            ),
        )

    from core.recommendation import apply_candidate_to_model
    from core.recommendation.apply_candidate import InapplicableCandidateError

    cand = _rehydrate_candidate(cand_dict)
    try:
        new_model, diff = apply_candidate_to_model(ctx["model_json"], cand)
    except InapplicableCandidateError as e:
        raise HTTPException(
            status_code=422,
            detail=f"apply failed: {e}",
        )

    return {
        "candidate_id": candidate_id,
        "applicable": True,
        "diff": {
            "candidate_id": diff.candidate_id,
            "operation": diff.operation,
            "reason": diff.reason,
            "changed_member_count": diff.changed_member_count,
            "changed_members": diff.changed_members,
        },
        "updated_model": new_model,
    }


@app.get("/api/v2/recommendations/chat-preview/{preview_id}")
async def get_recommendations_chat_preview(preview_id: str):
    """Fetch a chat-staged section-change preview.

    The Phase B chat tool ``propose_section_change`` stages the
    ``updated_model`` server-side and returns only a ``preview_id`` over
    the chat NDJSON stream (the streaming guard rejects ``updated_model``
    /``model_json`` keys at any depth — see
    ``chat/streaming.FORBIDDEN_KEYS``). The frontend bridge fetches the
    full payload here once before opening the rec-diff modal.

    Response shape matches ``/api/v2/recommendations/preview-apply`` so
    the existing ``_renderRecDiffPreview`` / ``applyRecDiff`` code path
    consumes it unchanged.

    Status codes:
        200 — preview present and unexpired.
        410 — preview existed but its TTL elapsed (chat tool ran more
              than 30 minutes ago). Ask the user to repeat the request.
        404 — preview id unknown (typo / wrong server / cache cleared).
    """
    from app.services.chat_preview_cache import get_preview

    entry = get_preview(preview_id)
    if entry is None:
        # We don't keep an audit trail of evicted ids, so 404 here may
        # also mean "expired and already swept". Surface 410 only when
        # the entry is *currently* still there at TTL-edge timing.
        raise HTTPException(
            status_code=404,
            detail=(
                f"chat preview '{preview_id}' not found — it may have "
                "expired (30 min TTL) or never existed. Ask the chatbot "
                "to repeat the section-change request."
            ),
        )
    # Mirror /preview-apply's top-level shape so applyRecDiff
    # (editor3d_v2.js) — which reads ``data.candidate_id`` for its
    # toast — works without a chat-specific branch. Codex review P3.
    candidate_dict = entry.get("candidate") or {}
    return {
        "preview_id": preview_id,
        "analysis_id": entry.get("analysis_id"),
        "candidate_id": candidate_dict.get("candidate_id"),
        "applicable": True,
        "diff": entry.get("diff") or {},
        "candidate": candidate_dict,
        "preview_meta": entry.get("preview_meta") or {},
        "updated_model": entry.get("updated_model") or {},
    }


@app.post("/api/v2/recommendations/explain")
async def post_recommendations_explain(request: Request):
    """Produce a Korean engineer-brief explanation for a candidate.

    The explanation is *advisory*: the deterministic pipeline (candidate
    generation + reanalysis-verified scoring) stays the source of truth.
    This endpoint never mutates the cached model_json — it deep-copies
    before deriving any diff. KDS evidence (when a real retriever is
    configured) is appended verbatim from validated chunks; the
    deterministic explainer never invents clause numbers.

    Request:
        {
            "analysis_id": "...",
            "candidate_id": "...",
            "evaluation": <optional CandidateEvaluation.to_dict()>,
            "diff": <optional ChangeDiff.to_dict()>,
            "language": "ko",
            "style": "engineer_brief"
        }

    Response:
        {
            "candidate_id": "...",
            "explanation": {summary, issue_interpretation, ...},
            "kds_evidence": [{doc_id, title, clause, quote, relevance, score}],
            "confidence": "low|medium|high",
            "source": {deterministic, rag_used, llm_used, score_method},
            "warnings": [...]
        }

    Status codes:
        200  — explanation returned (even if KDS-RAG is unavailable)
        400  — missing/invalid analysis_id or candidate_id, or unknown analysis
        404  — candidate_id not found in the analysis context
        410  — analysis context expired (TTL)
    """
    body = await request.json()
    analysis_id = body.get("analysis_id")
    candidate_id = body.get("candidate_id")
    evaluation_payload = body.get("evaluation")
    diff_payload = body.get("diff")
    language = body.get("language") or "ko"
    style = body.get("style") or "engineer_brief"

    if not analysis_id:
        raise HTTPException(status_code=400, detail="analysis_id is required")
    if not candidate_id:
        raise HTTPException(status_code=400, detail="candidate_id is required")

    # Same 400/410 split as /evaluate and /preview-apply: peek-then-get so
    # we can tell "never existed" apart from "expired and evicted".
    with _ANALYSIS_CONTEXT_LOCK:
        ever_existed = analysis_id in analysis_context_cache
    ctx = _get_analysis_context(analysis_id)
    if ctx is None:
        if ever_existed:
            raise HTTPException(
                status_code=410,
                detail=(
                    f"analysis context for '{analysis_id}' has expired "
                    "(30-minute TTL). Re-run /api/v2/analyze first."
                ),
            )
        raise HTTPException(
            status_code=400,
            detail=(
                f"analysis context for '{analysis_id}' not found. It may "
                "have never been created or the server was restarted. "
                "Re-run /api/v2/analyze first."
            ),
        )

    _assert_full_context_for_recommendations(ctx, analysis_id)

    cand_dict = (ctx.get("candidates_by_id") or {}).get(candidate_id)
    if cand_dict is None:
        raise HTTPException(
            status_code=404,
            detail=f"candidate '{candidate_id}' not found in analysis context",
        )

    candidate = _rehydrate_candidate(cand_dict)

    # Derive diff lazily when caller didn't supply one and the candidate
    # is auto-applicable. ``derive_diff_if_applicable`` deep-copies before
    # applying, so the cached model_json is never touched.
    diff_obj = diff_payload
    derive_warnings: list[str] = []
    if diff_obj is None:
        from core.recommendation import derive_diff_if_applicable
        diff_obj, derive_warnings = derive_diff_if_applicable(
            candidate, ctx.get("model_json"),
        )

    from core.kds_rag import get_default_kds_retriever
    from core.recommendation import explain_candidate

    retriever = get_default_kds_retriever()

    # analysis_summary helps the explainer compose contextual prose. We
    # keep it intentionally small to avoid leaking the full cached entry.
    design_check = ctx.get("design_check") or {}
    summary = (design_check.get("summary") or {}) if isinstance(
        design_check, dict
    ) else {}
    analysis_summary = {
        "max_interaction_ratio": summary.get("max_interaction_ratio"),
        "max_drift_ratio": summary.get("max_drift_ratio"),
        "ng_members": summary.get("ng_members"),
        "ng_stories": summary.get("ng_stories"),
    }

    result = explain_candidate(
        candidate,
        diff=diff_obj,
        evaluation=evaluation_payload,
        analysis_summary=analysis_summary,
        retriever=retriever,
        language=language,
        style=style,
    )

    # Surface diff-derivation warnings on the API response as well.
    for w in derive_warnings:
        if w not in result.warnings:
            result.warnings.append(w)

    return result.to_dict()


@app.get("/api/v2/viewer/{filename}")
async def serve_v2_viewer(filename: str):
    """V2 HTML 뷰어/리포트 서빙"""
    viewer_path = JOBS_DIR / "v2_viewers" / filename
    if not viewer_path.exists():
        raise HTTPException(
            status_code=404,
            detail=(
                "Viewer not found. Reports are written to local disk and are "
                "lost on Render redeploy / sleep. Please re-run the IFC parse."
            ),
        )
    return FileResponse(str(viewer_path), media_type="text/html")


@app.get("/api/jobs/{job_id}/report")
async def serve_job_report(job_id: str):
    """Serve the HTML report written by ``/api/v2/analyze`` (or the legacy
    ``/api/building/analyze``) into ``JOBS_DIR/{job_id}/report.html``.

    Jobs are stored on local disk only; on Render free plan the disk is
    wiped on each cold start / redeploy. A 404 here means the user needs
    to re-run the analysis with their current model.
    """
    report_path = JOBS_DIR / job_id / "report.html"
    if not report_path.exists():
        raise HTTPException(
            status_code=404,
            detail=(
                f"Report for job '{job_id}' not found. Job artifacts live on "
                "the worker's local disk and are lost on server restart "
                "(Render sleep / redeploy). Re-run the analysis to regenerate."
            ),
        )
    return FileResponse(str(report_path), media_type="text/html")


@app.post("/api/jobs/{job_id}/report-cover")
async def update_report_cover(job_id: str, request: Request):
    """표지·도장란(cover_info)을 주입하여 문서형 구조계산서 리포트를 재생성.

    재해석 없이 ``calc_data.json`` 사이드카(분석 시 plot_frame_3d_calc_report가
    저장)만 다시 렌더한다. 사이드카가 없으면(=폴백 탭 리포트가 떴거나 구버전
    잡) 409 — 재해석을 안내한다.

    Body: ``{"cover_info": {...}}`` 또는 cover_info dict 자체(평탄형) 모두 허용.
    """
    try:
        body = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="유효한 JSON 본문이 필요합니다.")
    cover_info = body.get("cover_info") if isinstance(body, dict) and "cover_info" in body else body
    if cover_info is not None and not isinstance(cover_info, dict):
        raise HTTPException(status_code=400, detail="cover_info는 객체(dict)여야 합니다.")

    job_dir = JOBS_DIR / job_id
    data_path = job_dir / "calc_data.json"
    report_path = job_dir / "report.html"

    if not data_path.exists():
        raise HTTPException(
            status_code=409,
            detail=(
                "표지 정보를 주입할 문서형 리포트 데이터(calc_data.json)가 없습니다. "
                "문서형 구조계산서가 생성된 해석에만 적용되며, 서버 재시작 등으로 "
                "잡 산출물이 사라졌을 수 있습니다. 해석을 다시 실행해 주세요."
            ),
        )

    if str(MCP_SERVER_PATH) not in sys.path:
        sys.path.insert(0, str(MCP_SERVER_PATH))

    try:
        with open(data_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        from core.visualization_calc_report import render_calc_report_from_data
        render_calc_report_from_data(data, cover_info, output_path=str(report_path))
    except Exception as e:
        logger.error("report-cover regeneration failed (job %s): %s", job_id, e, exc_info=True)
        raise HTTPException(status_code=500, detail=f"리포트 재생성 실패: {e}")

    if job_id in jobs_db:
        jobs_db[job_id]["cover_info"] = cover_info

    return {"status": "success", "report_url": f"/api/jobs/{job_id}/report"}


# ═══════════════════════════════════════════════════════════════════════════════
# Excel Export
# ═══════════════════════════════════════════════════════════════════════════════

# 케이스 그룹별 헤더 배경/글자 색 (hex, 6자리)
_GROUP_COLORS = {
    'Info':   ('37474F', 'FFFFFF'),  # 짙은 회색
    'DL':     ('0097A7', 'FFFFFF'),  # 시안
    'LL':     ('F57C00', 'FFFFFF'),  # 주황
    'EQ':     ('C62828', 'FFFFFF'),  # 빨강
    'Wind':   ('6A1B9A', 'FFFFFF'),  # 보라
    'RSA':    ('1565C0', 'FFFFFF'),  # 파랑
    'Combo0': ('E8F5E9', '000000'),  # 연녹 1
    'Combo1': ('F1F8E9', '000000'),  # 연녹 2
}


def _case_group(cname: str) -> str:
    """케이스 이름 → 그룹 분류."""
    if cname == 'Info':
        return 'Info'
    cu = cname.upper()
    if 'RSA' in cu:
        return 'RSA'
    if 'WIND' in cu or cu.startswith('W'):
        return 'Wind'
    if 'EQ' in cu or cu.startswith('S_'):
        return 'EQ'
    if cu.startswith('DL') or cu == 'DEAD':
        return 'DL'
    if cu.startswith('LL') or cu == 'LIVE':
        return 'LL'
    return 'Combo'


def _style_wide_header(ws, case_order, info_col_names, n_metrics):
    """2행 MultiIndex 헤더에 그룹별 색 적용.

    case_order: Info 제외한 케이스 이름 리스트
    info_col_names: ['node_id', 'x_m', ...] (Info 레벨 컬럼 이름들)
    n_metrics: 각 case 블록의 컬럼 수
    """
    from openpyxl.styles import PatternFill, Font, Alignment

    col = 1
    combo_idx = 0
    # Info 컬럼 먼저
    bg, fg = _GROUP_COLORS['Info']
    info_fill = PatternFill('solid', fgColor=bg)
    info_font = Font(bold=True, color=fg)
    for _ in info_col_names:
        for r in (1, 2):
            cell = ws.cell(row=r, column=col)
            cell.fill = info_fill
            cell.font = info_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        col += 1

    # 각 case 블록
    for cname in case_order:
        grp = _case_group(cname)
        if grp == 'Combo':
            fill_key = f'Combo{combo_idx % 2}'
            combo_idx += 1
        else:
            fill_key = grp
        bg, fg = _GROUP_COLORS[fill_key]
        fill = PatternFill('solid', fgColor=bg)
        font = Font(bold=True, color=fg)
        for r in (1, 2):
            for c in range(col, col + n_metrics):
                cell = ws.cell(row=r, column=c)
                cell.fill = fill
                cell.font = font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        col += n_metrics

    ws.freeze_panes = 'A3'


def _write_wide_sheet(writer, sheet_name, case_names, row_keys,
                      info_col_names, info_values, metric_keys, fetch_value):
    """openpyxl로 직접 wide sheet 쓰기 (MultiIndex + index=False 제약 회피).

    헤더 1행: Info 컬럼은 각 컬럼명 + case 블록은 첫 셀에 케이스명(merge)
    헤더 2행: Info 컬럼은 비움 + case 블록은 metric 이름들
    3행~: 데이터
    """
    from openpyxl import Workbook
    # pandas ExcelWriter.sheets는 기존 워크북 공유 → 새 시트 생성
    wb = writer.book
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    writer.sheets[sheet_name] = ws

    n_info = len(info_col_names)
    n_metrics = len(metric_keys)

    # 1행 헤더: Info 컬럼명 + case별 merge 셀
    for i, name in enumerate(info_col_names):
        ws.cell(row=1, column=i + 1, value=name)
    col = n_info + 1
    for cname in case_names:
        ws.cell(row=1, column=col, value=cname)
        if n_metrics > 1:
            ws.merge_cells(start_row=1, start_column=col,
                           end_row=1, end_column=col + n_metrics - 1)
        col += n_metrics

    # 2행 헤더: Info 컬럼은 이미 1행이 이름이라 빈 셀(또는 동일)도 가능 —
    # 보기 좋게 Info는 1행 이름 그대로 두고 2행도 동일값으로 유지 (merge 없이)
    for i, name in enumerate(info_col_names):
        ws.cell(row=2, column=i + 1, value=name)
    col = n_info + 1
    for _ in case_names:
        for m in metric_keys:
            ws.cell(row=2, column=col, value=m)
            col += 1

    # Info 컬럼 1~2행 세로 병합
    for i in range(n_info):
        ws.merge_cells(start_row=1, start_column=i + 1,
                       end_row=2, end_column=i + 1)

    # 3행~ 데이터
    for ri, rk in enumerate(row_keys):
        r = ri + 3
        for ci, name in enumerate(info_col_names):
            ws.cell(row=r, column=ci + 1, value=info_values[name][ri])
        col = n_info + 1
        for cname in case_names:
            for m in metric_keys:
                ws.cell(row=r, column=col, value=fetch_value(cname, rk, m))
                col += 1


@app.get("/api/export/excel/{job_id}")
async def export_excel(job_id: str):
    """해석 결과를 Excel 다중 시트(.xlsx)로 다운로드. Wide Pivot 구조."""
    result_path = JOBS_DIR / job_id / "result.json"
    if not result_path.exists():
        raise HTTPException(
            status_code=404,
            detail=(
                "해석 결과를 찾을 수 없습니다. 결과 파일은 서버 로컬 디스크에만 "
                "저장되며 Render 슬립/재배포 시 사라집니다. 해석을 다시 실행해 "
                "주세요."
            ),
        )

    try:
        import pandas as pd
        from io import BytesIO
    except ImportError:
        raise HTTPException(status_code=500, detail="pandas/openpyxl 미설치. `pip install pandas openpyxl` 실행 필요.")

    with open(result_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    case_data = data.get("case_data", {}) or {}
    member_forces = data.get("member_forces", {}) or {}
    envelope = data.get("envelope", {}) or {}
    nodes_info = data.get("nodes_info", {}) or {}

    def _coord(nid, axis):
        """nid의 x/y/z (없으면 None)."""
        v = nodes_info.get(str(nid))
        if not v:
            return None
        return v[axis] if axis < len(v) else None

    def _nid_sort_key(nid):
        try:
            return (0, int(nid))
        except (ValueError, TypeError):
            return (1, str(nid))

    # ─── 1) Displacements (Wide) ─────────────────────────────
    disp_metrics = ["dx_mm", "dy_mm", "dz_mm", "rx_rad", "ry_rad", "rz_rad"]
    disp_case_order = list(case_data.keys())
    all_disp_nodes = set()
    for cd in case_data.values():
        all_disp_nodes.update((cd.get("displacements") or {}).keys())
    all_disp_nodes = sorted(all_disp_nodes, key=_nid_sort_key)

    def _disp_get(cname, nid, m):
        vals = (case_data.get(cname, {}).get("displacements") or {}).get(str(nid))
        if not vals:
            return None
        idx = disp_metrics.index(m)
        return vals[idx] if idx < len(vals) else None

    disp_info_cols = ["node_id", "x_m", "y_m", "z_m"]
    disp_info_values = {
        "node_id": list(all_disp_nodes),
        "x_m": [_coord(n, 0) for n in all_disp_nodes],
        "y_m": [_coord(n, 1) for n in all_disp_nodes],
        "z_m": [_coord(n, 2) for n in all_disp_nodes],
    }

    # ─── 2) Member Forces (Wide) ─────────────────────────────
    force_metrics = ["N_kN", "Vy_kN", "Vz_kN", "T_kNm", "My_kNm", "Mz_kNm"]
    force_case_order = list(member_forces.keys())

    # 모든 부재 수집 + 첫 등장 case에서 info 추출
    member_info_map = {}  # member_id -> {ni, nj, type, section}
    for cname in force_case_order:
        mf_list = member_forces.get(cname, [])
        if not isinstance(mf_list, list):
            continue
        for mf in mf_list:
            mid = mf.get("member_id")
            if mid is None or mid in member_info_map:
                continue
            member_info_map[mid] = {
                "ni": mf.get("ni"),
                "nj": mf.get("nj"),
                "type": mf.get("type"),
                "section": mf.get("section"),
            }
    all_members = sorted(member_info_map.keys(), key=lambda x: int(x) if str(x).isdigit() else 0)

    # member_id → case → {metric: value} 맵 사전 구축 (O(N) lookup)
    force_lookup = {}  # (cname, mid) -> mf dict
    for cname in force_case_order:
        mf_list = member_forces.get(cname, [])
        if not isinstance(mf_list, list):
            continue
        for mf in mf_list:
            mid = mf.get("member_id")
            if mid is not None:
                force_lookup[(cname, mid)] = mf

    def _signed_peak(v):
        """스칼라면 그대로, 배열이면 절댓값 최대(부호 유지)."""
        if isinstance(v, (list, tuple)):
            if not v:
                return None
            vals = [x for x in v if x is not None]
            if not vals:
                return None
            return max(vals, key=lambda x: abs(x))
        return v

    def _force_get(cname, mid, m):
        mf = force_lookup.get((cname, mid))
        if not mf:
            return None
        return _signed_peak(mf.get(m))

    force_info_cols = ["member_id", "ni", "nj", "type", "section"]
    force_info_values = {
        "member_id": list(all_members),
        "ni": [member_info_map[m]["ni"] for m in all_members],
        "nj": [member_info_map[m]["nj"] for m in all_members],
        "type": [member_info_map[m]["type"] for m in all_members],
        "section": [member_info_map[m]["section"] for m in all_members],
    }

    # ─── 3) Reactions (Wide) ─────────────────────────────────
    react_metrics = ["RX_kN", "RY_kN", "RZ_kN", "MX_kNm", "MY_kNm", "MZ_kNm"]
    react_case_order = list(case_data.keys())

    # 지점 노드 수집 + 각 case의 lookup 구축
    react_lookup = {}  # (cname, node_id) -> r dict
    all_react_nodes = set()
    for cname in react_case_order:
        reactions = case_data.get(cname, {}).get("reactions") or []
        for r in reactions:
            nid = r.get("node_id") if "node_id" in r else r.get("node")
            if nid is None:
                continue
            all_react_nodes.add(str(nid))
            react_lookup[(cname, str(nid))] = r
    all_react_nodes = sorted(all_react_nodes, key=_nid_sort_key)

    def _react_get(cname, nid, m):
        r = react_lookup.get((cname, str(nid)))
        if not r:
            return None
        return _signed_peak(r.get(m))

    def _react_coord(nid, axis):
        """반력 노드 좌표: nodes_info 우선, fallback은 첫 case의 reaction 항목."""
        v = nodes_info.get(str(nid))
        if v:
            return v[axis] if axis < len(v) else None
        # fallback: 첫 등장 case의 좌표
        for cname in react_case_order:
            r = react_lookup.get((cname, str(nid)))
            if r:
                keys = ["x_m", "y_m", "z_m"]
                return r.get(keys[axis])
        return None

    react_info_cols = ["node_id", "x_m", "y_m", "z_m"]
    react_info_values = {
        "node_id": list(all_react_nodes),
        "x_m": [_react_coord(n, 0) for n in all_react_nodes],
        "y_m": [_react_coord(n, 1) for n in all_react_nodes],
        "z_m": [_react_coord(n, 2) for n in all_react_nodes],
    }

    # ─── 4) Envelope (기존 long-form 유지) ───────────────────
    env_rows = [{"metric": k, "value": v} for k, v in envelope.items()]
    df_env = pd.DataFrame(env_rows)

    # ─── 5) Metadata ─────────────────────────────────────────
    meta_rows = [
        {"key": "job_id", "value": data.get("job_id", "")},
        {"key": "created_at", "value": data.get("created_at", "")},
        {"key": "analysis_type", "value": data.get("analysis_type", "")},
        {"key": "seismic_method", "value": data.get("seismic_method", "")},
        {"key": "num_stories", "value": data.get("num_stories", 0)},
        {"key": "num_nodes", "value": data.get("num_nodes", 0)},
        {"key": "num_elements", "value": data.get("num_elements", 0)},
        {"key": "n_cases", "value": len(data.get("case_names", []))},
        {"key": "n_combos", "value": len(data.get("combo_names", []))},
    ]
    df_meta = pd.DataFrame(meta_rows)

    # ─── Excel 쓰기 ──────────────────────────────────────────
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer:
        # Envelope/Metadata는 단순 long-form → pandas로
        df_env.to_excel(writer, sheet_name='Envelope', index=False)
        df_meta.to_excel(writer, sheet_name='Metadata', index=False)

        # Wide sheets: openpyxl 직접 쓰기 (MultiIndex+index=False 미지원 회피)
        _write_wide_sheet(writer, 'Displacements', disp_case_order, all_disp_nodes,
                          disp_info_cols, disp_info_values, disp_metrics, _disp_get)
        _write_wide_sheet(writer, 'Member_Forces', force_case_order, all_members,
                          force_info_cols, force_info_values, force_metrics, _force_get)
        _write_wide_sheet(writer, 'Reactions', react_case_order, all_react_nodes,
                          react_info_cols, react_info_values, react_metrics, _react_get)

        # 헤더 그룹별 색상 + freeze pane
        _style_wide_header(writer.sheets['Displacements'],
                           disp_case_order, disp_info_cols, len(disp_metrics))
        _style_wide_header(writer.sheets['Member_Forces'],
                           force_case_order, force_info_cols, len(force_metrics))
        _style_wide_header(writer.sheets['Reactions'],
                           react_case_order, react_info_cols, len(react_metrics))

        # 시트 순서: Displacements, Member_Forces, Reactions, Envelope, Metadata
        wb = writer.book
        desired = ['Displacements', 'Member_Forces', 'Reactions', 'Envelope', 'Metadata']
        wb._sheets = [wb[n] for n in desired if n in wb.sheetnames]

        # 열 너비 자동 조정 (MergedCell 회피: 인덱스 기반 순회)
        from openpyxl.utils import get_column_letter
        from openpyxl.cell.cell import MergedCell
        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            max_col = ws.max_column
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                max_len = 0
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if isinstance(cell, MergedCell):
                        continue
                    val = str(cell.value) if cell.value is not None else ""
                    if len(val) > max_len:
                        max_len = len(val)
                ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 32)

    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename=analysis_{job_id}.xlsx'}
    )


# ═══════════════════════════════════════════════════════════════════════════════
# DXF Export (층별 평면도)
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/api/export/dxf")
async def export_dxf(request: Request):
    """V2 모델을 층별 평면 DXF로 내보내기."""
    import ezdxf
    from io import BytesIO

    body = await request.json()
    model = body.get("model")
    if not model:
        return JSONResponse({"error": "모델 데이터 없음"}, status_code=400)

    nodes = model.get("nodes", [])
    elements = model.get("elements", [])
    elevations = model.get("story_elevations", [])

    if not nodes or not elements:
        return JSONResponse({"error": "노드/요소 없음"}, status_code=400)

    node_map = {n["id"]: n for n in nodes}
    n_stories = len(elevations) - 1 if len(elevations) > 1 else 1

    # ── 모델 크기 산정 (배치 offset용) ──
    xs = [n["x"] for n in nodes]
    ys = [n["y"] for n in nodes]
    model_w = max(xs) - min(xs) if xs else 20
    model_h = max(ys) - min(ys) if ys else 20
    x_min = min(xs) if xs else 0
    x_max = max(xs) if xs else 20
    y_min = min(ys) if ys else 0
    y_max = max(ys) if ys else 20

    # 테두리 박스 margin
    # 좌측: 그리드연장(5%) + 축번호(3%) + 개별치수선(5%) + 전체치수선(5%) + 여유(4%) = 22%
    # 하단: 동일 = 22%
    # 상단: 제목박스(8%) + 여유(4%) = 12%
    # 우측: 여유(5%)
    md = max(model_w, model_h)
    margin_left = md * 0.28
    margin_right = md * 0.08
    margin_bottom = md * 0.26
    margin_top = md * 0.18
    plan_h = model_h + margin_bottom + margin_top
    plan_w = model_w + margin_left + margin_right
    spacing = plan_h + md * 0.10  # 평면도 간 Y간격

    # ── DXF 문서 생성 ──
    doc = ezdxf.new("R2010", setup=True)
    msp = doc.modelspace()

    # 색상 (DXF ACI color index)
    COL_GRID = 8       # 회색 (그리드선)
    COL_COLUMN = 1     # 빨강
    COL_BEAM = 3       # 초록
    COL_TEXT = 7       # 흰/검
    COL_DIM = 2        # 노랑
    COL_SUPPORT = 6    # 마젠타
    COL_FRAME = 5      # 파랑 (테두리)

    col_sym_size = 0.3  # m

    # 치수 스타일 설정 (모델 크기 비례)
    dim_scale = max(model_w, model_h) * 0.025
    # 소수점 구분자를 "."로 강제 (기본은 locale에 따라 "," 사용)
    doc.header["$DIMDSEP"] = ord(".")
    if "STRUCT_DIM" not in doc.dimstyles:
        dimstyle = doc.dimstyles.new("STRUCT_DIM")
        dimstyle.dxf.dimtxt = dim_scale       # 치수 텍스트 높이
        dimstyle.dxf.dimasz = dim_scale * 0.8 # 화살표 크기
        dimstyle.dxf.dimexe = dim_scale * 0.4 # 연장선 돌출
        dimstyle.dxf.dimexo = dim_scale * 0.3 # 원점 오프셋
        dimstyle.dxf.dimgap = dim_scale * 0.3
        dimstyle.dxf.dimclrt = COL_DIM
        dimstyle.dxf.dimdec = 2               # 소수 자릿수
        dimstyle.dxf.dimdsep = ord(".")       # 소수점 = "."

    # 층 루프: 0=GL, 1=1F, 2=2F, ...
    for story_idx in range(0, n_stories + 1):
        is_gl = (story_idx == 0)
        z_level = elevations[story_idx] if story_idx < len(elevations) else story_idx * 3.5
        z_tol = 0.01
        label = "GL" if is_gl else f"{story_idx}F"

        layer_frame = f"{label}_FRAME"
        layer_grid = f"{label}_GRID"
        layer_col = f"{label}_COL"
        layer_beam = f"{label}_BEAM"
        layer_text = f"{label}_TEXT"
        layer_dim = f"{label}_DIM"

        for lname, lcolor in [
            (layer_frame, COL_FRAME), (layer_grid, COL_GRID),
            (layer_col, COL_COLUMN), (layer_beam, COL_BEAM),
            (layer_text, COL_TEXT), (layer_dim, COL_DIM),
        ]:
            if lname not in doc.layers:
                doc.layers.add(lname, color=lcolor)

        # 배치 offset: 아래에서 위로 GL → 1F → 2F → ...
        ox = 0
        oy = story_idx * spacing

        # 해당 층 노드
        story_nodes = [n for n in nodes if abs(n["z"] - z_level) < z_tol]
        story_node_ids = {n["id"] for n in story_nodes}

        # 해당 층 요소 (GL은 요소 없음)
        story_elements = [
            e for e in elements
            if e.get("node_i") in story_node_ids and e.get("node_j") in story_node_ids
        ] if not is_gl else []

        # 기둥: 하부 노드가 아래층, 상부 노드가 이 층 (GL은 없음)
        column_elements = []
        if not is_gl:
            z_below = elevations[story_idx - 1] if story_idx - 1 >= 0 else 0
            below_ids = {n["id"] for n in nodes if abs(n["z"] - z_below) < z_tol}
            column_elements = [
                e for e in elements
                if e.get("elem_type") == "column"
                and ((e["node_i"] in below_ids and e["node_j"] in story_node_ids)
                     or (e["node_j"] in below_ids and e["node_i"] in story_node_ids))
            ]

        # ── 테두리 박스 ──
        fx1 = ox + x_min - margin_left
        fy1 = oy + y_min - margin_bottom
        fx2 = ox + x_max + margin_right
        fy2 = oy + y_max + margin_top
        msp.add_lwpolyline(
            [(fx1, fy1), (fx2, fy1), (fx2, fy2), (fx1, fy2)],
            close=True,
            dxfattribs={"layer": layer_frame, "lineweight": 50}
        )
        # 내부 박스 (여백)
        inner_m = md * 0.012
        msp.add_lwpolyline(
            [(fx1+inner_m, fy1+inner_m), (fx2-inner_m, fy1+inner_m),
             (fx2-inner_m, fy2-inner_m), (fx1+inner_m, fy2-inner_m)],
            close=True,
            dxfattribs={"layer": layer_frame, "lineweight": 25}
        )
        # 제목 박스 (우상단) — 텍스트가 충분히 들어갈 크기
        title_text = f"{label} PLAN  EL. {z_level:.1f}m"
        title_text_h = md * 0.028  # 텍스트 높이
        title_box_h = title_text_h * 2.2  # 텍스트 높이의 2.2배 (상하 여유)
        # 대략적인 텍스트 너비 계산 (AutoCAD txt 폰트: 글자당 폭 ≈ height × 0.9)
        text_width_est = len(title_text) * title_text_h * 0.9
        title_box_w = max(md * 0.5, text_width_est + title_text_h * 3.0)
        tb_x2 = fx2 - inner_m
        tb_y2 = fy2 - inner_m
        tb_x1 = tb_x2 - title_box_w
        tb_y1 = tb_y2 - title_box_h
        msp.add_lwpolyline(
            [(tb_x1, tb_y1), (tb_x2, tb_y1), (tb_x2, tb_y2), (tb_x1, tb_y2)],
            close=True,
            dxfattribs={"layer": layer_frame}
        )
        t = msp.add_text(
            title_text,
            dxfattribs={
                "layer": layer_text, "height": title_text_h,
                "color": COL_TEXT,
            }
        )
        t.set_placement(
            ((tb_x1 + tb_x2) / 2, (tb_y1 + tb_y2) / 2),
            align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER,
        )

        # ── 그리드선 + 축번호 ──
        story_xs = sorted(set(round(n["x"], 3) for n in (story_nodes if story_nodes else nodes)))
        story_ys = sorted(set(round(n["y"], 3) for n in (story_nodes if story_nodes else nodes)))
        grid_ext = max(model_w, model_h) * 0.05

        for i, gx in enumerate(story_xs):
            msp.add_line(
                (ox + gx, oy + y_min - grid_ext),
                (ox + gx, oy + y_max + grid_ext),
                dxfattribs={"layer": layer_grid, "linetype": "CENTER"}
            )
            t = msp.add_text(
                f"X{i+1}",
                dxfattribs={"layer": layer_text, "height": dim_scale * 1.2, "color": COL_DIM}
            )
            t.set_placement(
                (ox + gx, oy + y_min - grid_ext - dim_scale * 1.5),
                align=ezdxf.enums.TextEntityAlignment.CENTER,
            )

        for j, gy in enumerate(story_ys):
            msp.add_line(
                (ox + x_min - grid_ext, oy + gy),
                (ox + x_max + grid_ext, oy + gy),
                dxfattribs={"layer": layer_grid, "linetype": "CENTER"}
            )
            t = msp.add_text(
                f"Y{j+1}",
                dxfattribs={"layer": layer_text, "height": dim_scale * 1.2, "color": COL_DIM}
            )
            t.set_placement(
                (ox + x_min - grid_ext - dim_scale * 0.8, oy + gy),
                align=ezdxf.enums.TextEntityAlignment.MIDDLE_RIGHT,
            )

        # ── 보 ──
        for e in story_elements:
            if e.get("elem_type") == "column":
                continue
            ni = node_map.get(e["node_i"])
            nj = node_map.get(e["node_j"])
            if not ni or not nj:
                continue
            msp.add_line(
                (ox + ni["x"], oy + ni["y"]),
                (ox + nj["x"], oy + nj["y"]),
                dxfattribs={"layer": layer_beam}
            )

        # ── 기둥 심볼 ──
        col_positions = set()
        for e in column_elements:
            ni = node_map.get(e["node_i"])
            nj = node_map.get(e["node_j"])
            top_node = nj if (nj and nj["id"] in story_node_ids) else ni
            if not top_node:
                continue
            pos_key = (round(top_node["x"], 3), round(top_node["y"], 3))
            if pos_key in col_positions:
                continue
            col_positions.add(pos_key)
            cx, cy = ox + top_node["x"], oy + top_node["y"]
            hs = col_sym_size / 2
            msp.add_lwpolyline(
                [(cx-hs, cy-hs), (cx+hs, cy-hs), (cx+hs, cy+hs), (cx-hs, cy+hs)],
                close=True,
                dxfattribs={"layer": layer_col}
            )

        # ── GL: 지점 심볼 (삼각형) ──
        if is_gl:
            for n in story_nodes:
                if not n.get("support"):
                    continue
                sx, sy = ox + n["x"], oy + n["y"]
                ts = 0.25
                msp.add_lwpolyline(
                    [(sx, sy), (sx - ts, sy - ts*1.5), (sx + ts, sy - ts*1.5)],
                    close=True,
                    dxfattribs={"layer": layer_col, "color": COL_SUPPORT}
                )

        # ── 치수선 (bay 간격) ──
        # X방향 치수 (하단): 연속 치수
        if len(story_xs) >= 2:
            dim_y = oy + y_min - grid_ext - dim_scale * 3.5
            for i in range(len(story_xs) - 1):
                x1 = ox + story_xs[i]
                x2 = ox + story_xs[i + 1]
                dim = msp.add_linear_dim(
                    base=(0, dim_y),
                    p1=(x1, oy + y_min),
                    p2=(x2, oy + y_min),
                    angle=0,
                    dimstyle="STRUCT_DIM",
                    dxfattribs={"layer": layer_dim},
                )
                dim.render()
            # 전체 치수 (overall)
            dim_y_all = dim_y - dim_scale * 2.5
            dim = msp.add_linear_dim(
                base=(0, dim_y_all),
                p1=(ox + story_xs[0], oy + y_min),
                p2=(ox + story_xs[-1], oy + y_min),
                angle=0,
                dimstyle="STRUCT_DIM",
                dxfattribs={"layer": layer_dim},
            )
            dim.render()

        # Y방향 치수 (좌측)
        if len(story_ys) >= 2:
            dim_x = ox + x_min - grid_ext - dim_scale * 3.5
            for j in range(len(story_ys) - 1):
                y1 = oy + story_ys[j]
                y2 = oy + story_ys[j + 1]
                dim = msp.add_linear_dim(
                    base=(dim_x, 0),
                    p1=(ox + x_min, y1),
                    p2=(ox + x_min, y2),
                    angle=90,
                    dimstyle="STRUCT_DIM",
                    dxfattribs={"layer": layer_dim},
                )
                dim.render()
            # 전체 치수
            dim_x_all = dim_x - dim_scale * 2.5
            dim = msp.add_linear_dim(
                base=(dim_x_all, 0),
                p1=(ox + x_min, oy + story_ys[0]),
                p2=(ox + x_min, oy + story_ys[-1]),
                angle=90,
                dimstyle="STRUCT_DIM",
                dxfattribs={"layer": layer_dim},
            )
            dim.render()

    # ═══════════════════════════════════════════════════════════
    # 입면도 (Elevations)
    # ═══════════════════════════════════════════════════════════
    total_z = elevations[-1] if elevations else 0
    model_h_z = total_z  # 입면도의 세로 크기 = 전체 높이
    # 평면도 우측으로 offset (평면 폭 + margin)
    elev_col1_ox = plan_w + md * 0.20  # X방향 입면도 컬럼
    elev_col2_ox = elev_col1_ox + (model_w + margin_left + margin_right) + md * 0.20  # Y방향 입면도 컬럼

    # 입면도 그리기 공통 헬퍼
    def _draw_elevation(ox, oy, label_text, horiz_coords, hmin, hmax,
                        draw_columns, draw_beams, layer_prefix, axis_prefix):
        """입면도 1개 그리기 (공통 로직).
        horiz_coords: 수평축 그리드 좌표 (축번호용)
        hmin, hmax: 수평축 범위
        draw_columns: [(x_pos, z_start, z_end)] 기둥 리스트
        draw_beams: [(x_start, x_end, z)] 보 리스트
        """
        # 레이어
        lf = f"{layer_prefix}_FRAME"
        lg = f"{layer_prefix}_GRID"
        lc = f"{layer_prefix}_COL"
        lb = f"{layer_prefix}_BEAM"
        lt = f"{layer_prefix}_TEXT"
        ld = f"{layer_prefix}_DIM"
        for name, cc in [(lf, COL_FRAME), (lg, COL_GRID), (lc, COL_COLUMN),
                         (lb, COL_BEAM), (lt, COL_TEXT), (ld, COL_DIM)]:
            if name not in doc.layers:
                doc.layers.add(name, color=cc)

        # 테두리
        fx1 = ox + hmin - margin_left
        fy1 = oy - margin_bottom
        fx2 = ox + hmax + margin_right
        fy2 = oy + total_z + margin_top
        msp.add_lwpolyline(
            [(fx1, fy1), (fx2, fy1), (fx2, fy2), (fx1, fy2)],
            close=True,
            dxfattribs={"layer": lf, "lineweight": 50}
        )
        # 내부 박스
        inner_mm = md * 0.012
        msp.add_lwpolyline(
            [(fx1+inner_mm, fy1+inner_mm), (fx2-inner_mm, fy1+inner_mm),
             (fx2-inner_mm, fy2-inner_mm), (fx1+inner_mm, fy2-inner_mm)],
            close=True,
            dxfattribs={"layer": lf, "lineweight": 25}
        )
        # 제목 박스
        t_title_h = md * 0.028
        tb_h = t_title_h * 2.2
        tw_est = len(label_text) * t_title_h * 0.9
        tb_w = max(md * 0.5, tw_est + t_title_h * 3.0)
        tbx2 = fx2 - inner_mm
        tby2 = fy2 - inner_mm
        tbx1 = tbx2 - tb_w
        tby1 = tby2 - tb_h
        msp.add_lwpolyline(
            [(tbx1, tby1), (tbx2, tby1), (tbx2, tby2), (tbx1, tby2)],
            close=True,
            dxfattribs={"layer": lf}
        )
        tt = msp.add_text(
            label_text,
            dxfattribs={"layer": lt, "height": t_title_h, "color": COL_TEXT}
        )
        tt.set_placement(
            ((tbx1 + tbx2) / 2, (tby1 + tby2) / 2),
            align=ezdxf.enums.TextEntityAlignment.MIDDLE_CENTER,
        )

        gext = md * 0.05

        # 수직 그리드 (수평 위치: horiz_coords) + 축번호
        for i, hx in enumerate(horiz_coords):
            msp.add_line(
                (ox + hx, oy - gext),
                (ox + hx, oy + total_z + gext),
                dxfattribs={"layer": lg, "linetype": "CENTER"}
            )
            t = msp.add_text(
                f"{axis_prefix}{i+1}",  # X1/X2... or Y1/Y2...
                dxfattribs={"layer": lt, "height": dim_scale * 1.2, "color": COL_DIM}
            )
            t.set_placement(
                (ox + hx, oy - gext - dim_scale * 1.5),
                align=ezdxf.enums.TextEntityAlignment.CENTER,
            )

        # 수평 그리드 (각 층 z) + 층 라벨 (GL/1F/2F...)
        for s_idx, ez in enumerate(elevations):
            msp.add_line(
                (ox + hmin - gext, oy + ez),
                (ox + hmax + gext, oy + ez),
                dxfattribs={"layer": lg, "linetype": "CENTER"}
            )
            story_label = "GL" if s_idx == 0 else f"{s_idx}F"
            t = msp.add_text(
                story_label,
                dxfattribs={"layer": lt, "height": dim_scale * 1.2, "color": COL_DIM}
            )
            t.set_placement(
                (ox + hmin - gext - dim_scale * 0.8, oy + ez),
                align=ezdxf.enums.TextEntityAlignment.MIDDLE_RIGHT,
            )

        # 기둥 (수직 LINE)
        for hx, z_start, z_end in draw_columns:
            msp.add_line(
                (ox + hx, oy + z_start),
                (ox + hx, oy + z_end),
                dxfattribs={"layer": lc}
            )

        # 보 (수평 LINE)
        for hx_start, hx_end, z in draw_beams:
            msp.add_line(
                (ox + hx_start, oy + z),
                (ox + hx_end, oy + z),
                dxfattribs={"layer": lb}
            )

        # 수평 치수선 (bay 간격)
        if len(horiz_coords) >= 2:
            d_y = oy - gext - dim_scale * 3.5
            for i in range(len(horiz_coords) - 1):
                dim = msp.add_linear_dim(
                    base=(0, d_y),
                    p1=(ox + horiz_coords[i], oy),
                    p2=(ox + horiz_coords[i + 1], oy),
                    angle=0,
                    dimstyle="STRUCT_DIM",
                    dxfattribs={"layer": ld},
                )
                dim.render()
            d_y_all = d_y - dim_scale * 2.5
            dim = msp.add_linear_dim(
                base=(0, d_y_all),
                p1=(ox + horiz_coords[0], oy),
                p2=(ox + horiz_coords[-1], oy),
                angle=0,
                dimstyle="STRUCT_DIM",
                dxfattribs={"layer": ld},
            )
            dim.render()

        # 수직 치수선 (층고)
        if len(elevations) >= 2:
            d_x = ox + hmin - gext - dim_scale * 3.5
            for s_idx in range(len(elevations) - 1):
                dim = msp.add_linear_dim(
                    base=(d_x, 0),
                    p1=(ox + hmin, oy + elevations[s_idx]),
                    p2=(ox + hmin, oy + elevations[s_idx + 1]),
                    angle=90,
                    dimstyle="STRUCT_DIM",
                    dxfattribs={"layer": ld},
                )
                dim.render()
            d_x_all = d_x - dim_scale * 2.5
            dim = msp.add_linear_dim(
                base=(d_x_all, 0),
                p1=(ox + hmin, oy + elevations[0]),
                p2=(ox + hmin, oy + elevations[-1]),
                angle=90,
                dimstyle="STRUCT_DIM",
                dxfattribs={"layer": ld},
            )
            dim.render()

    # X방향 입면도 (Y_j 라인별) — 좌→우: X1, X2... / 상→하: Y_j
    all_xs = sorted(set(round(n["x"], 3) for n in nodes))
    all_ys = sorted(set(round(n["y"], 3) for n in nodes))
    y_tol = 0.01

    elev_x_h = total_z + margin_bottom + margin_top
    elev_x_spacing = elev_x_h + md * 0.10

    for j, y_line in enumerate(all_ys):
        # 해당 Y 라인 상의 노드
        at_line_nodes = [n for n in nodes if abs(n["y"] - y_line) < y_tol]
        at_line_ids = {n["id"] for n in at_line_nodes}
        if len(at_line_nodes) < 2:
            continue

        # 기둥: node_i, node_j 모두 이 라인, elem_type=column
        draw_columns = []
        for e in elements:
            if e.get("elem_type") != "column":
                continue
            if e["node_i"] not in at_line_ids or e["node_j"] not in at_line_ids:
                continue
            ni = node_map[e["node_i"]]
            nj = node_map[e["node_j"]]
            if abs(ni["x"] - nj["x"]) > y_tol:
                continue  # 경사 기둥 배제
            draw_columns.append((ni["x"], min(ni["z"], nj["z"]), max(ni["z"], nj["z"])))

        # 보: node_i, node_j 모두 이 라인, elem_type=beam_x or beam
        draw_beams = []
        for e in elements:
            et = e.get("elem_type", "")
            if et == "column":
                continue
            if e["node_i"] not in at_line_ids or e["node_j"] not in at_line_ids:
                continue
            ni = node_map[e["node_i"]]
            nj = node_map[e["node_j"]]
            if abs(ni["z"] - nj["z"]) > y_tol:
                continue  # 수평 보만
            draw_beams.append((min(ni["x"], nj["x"]), max(ni["x"], nj["x"]), ni["z"]))

        # 배치
        ox = elev_col1_ox
        oy = j * elev_x_spacing
        label_text = f"Y{j+1} ELEVATION"
        layer_prefix = f"ELEV_Y{j+1}"
        _draw_elevation(ox, oy, label_text, all_xs, x_min, x_max,
                        draw_columns, draw_beams, layer_prefix, "X")

    # Y방향 입면도 (X_i 라인별)
    elev_y_w = model_h  # Y방향 입면도의 가로 = Y치수
    for i, x_line in enumerate(all_xs):
        at_line_nodes = [n for n in nodes if abs(n["x"] - x_line) < y_tol]
        at_line_ids = {n["id"] for n in at_line_nodes}
        if len(at_line_nodes) < 2:
            continue

        draw_columns = []
        for e in elements:
            if e.get("elem_type") != "column":
                continue
            if e["node_i"] not in at_line_ids or e["node_j"] not in at_line_ids:
                continue
            ni = node_map[e["node_i"]]
            nj = node_map[e["node_j"]]
            if abs(ni["y"] - nj["y"]) > y_tol:
                continue
            draw_columns.append((ni["y"], min(ni["z"], nj["z"]), max(ni["z"], nj["z"])))

        draw_beams = []
        for e in elements:
            et = e.get("elem_type", "")
            if et == "column":
                continue
            if e["node_i"] not in at_line_ids or e["node_j"] not in at_line_ids:
                continue
            ni = node_map[e["node_i"]]
            nj = node_map[e["node_j"]]
            if abs(ni["z"] - nj["z"]) > y_tol:
                continue
            draw_beams.append((min(ni["y"], nj["y"]), max(ni["y"], nj["y"]), ni["z"]))

        ox = elev_col2_ox
        oy = i * elev_x_spacing
        label_text = f"X{i+1} ELEVATION"
        layer_prefix = f"ELEV_X{i+1}"
        _draw_elevation(ox, oy, label_text, all_ys, y_min, y_max,
                        draw_columns, draw_beams, layer_prefix, "Y")

    # ── Extents 및 뷰 자동 설정 (AutoCAD 열 때 도면이 보이도록) ──
    try:
        from ezdxf import zoom
        zoom.extents(msp, factor=1.1)
    except Exception as ex:
        logger.warning("DXF zoom.extents failed: %s", ex)

    # ── 파일 출력 (ezdxf는 파일 경로 필요) ──
    import tempfile, os
    tmp = tempfile.NamedTemporaryFile(suffix='.dxf', delete=False)
    tmp.close()
    doc.saveas(tmp.name)
    buf = BytesIO(open(tmp.name, 'rb').read())
    os.unlink(tmp.name)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/dxf",
        headers={"Content-Disposition": "attachment; filename=structural_plan.dxf"}
    )


# ═══════════════════════════════════════════════════════════════════════════════
# Health check
# ═══════════════════════════════════════════════════════════════════════════════

@app.get("/health")
async def health():
    return {"status": "ok", "mode": "simple (no Redis)"}


@app.get("/api/health/demo")
async def health_demo():
    """Is it worth sending a visitor into a live run right now?

    The landing page calls this before opening the "run it yourself" door. A
    visitor whose first click ends in a timeout or a raw error concludes the
    tool does not work — better to say "busy, come back" than to hand them a
    request that will sit behind three others.
    """
    depth = queue_depth()
    kds_offline = None
    try:
        if str(MCP_SERVER_PATH) not in sys.path:
            sys.path.insert(0, str(MCP_SERVER_PATH))
        from core import kds_cache
        kds_offline = kds_cache.is_offline()
        kds_mode = kds_cache.mode()
    except Exception:
        kds_mode = "unknown"

    # Busy is not broken: the run still works, it just queues.
    ready = depth <= 2
    return {
        "ready": ready,
        "queue_depth": depth,
        "kds_source": "snapshot" if (kds_mode in ("prefer", "only") or kds_offline)
                      else "database",
        "kds_mode": kds_mode,
        "demo_max_members": DEMO_MAX_MEMBERS or None,
        "expected_seconds": 15,
        "message": (
            "지금 바로 실행할 수 있습니다." if ready else
            f"현재 {depth}건이 대기 중입니다. 잠시 후 다시 시도해 주세요."
        ),
    }


if __name__ == "__main__":
    import uvicorn
    logger.info("=" * 50)
    logger.info("OpenSees Structural Analysis Platform")
    logger.info("=" * 50)
    uvicorn.run(app, host="0.0.0.0", port=8000)
